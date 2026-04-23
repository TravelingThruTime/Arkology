"""
monthly_tasks.py — Run on the 1st of each month.
=================================================
Computes N/M from raw input columns (no formula dependency).
Balances A19. Accumulates Taxes Owed. Sends email report.

Usage:
  python monthly_tasks.py           → dry run / preview
  python monthly_tasks.py --apply   → write changes to Excel + send email
"""
import sys, os, math, datetime, smtplib

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Ark_Database_v6-1.xlsx")
DRY_RUN    = "--apply" not in sys.argv

# ── Email config (fill in to enable reports) ──────────────────────────────────
EMAIL_FROM    = "ark@arkology.org"
EMAIL_TO      = "admin@arkology.org"
SMTP_HOST     = "smtp.gmail.com"
SMTP_PORT     = 587
SMTP_USER     = ""          # set this to enable email
SMTP_PASSWORD = ""
# ─────────────────────────────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("[ERROR] openpyxl not installed."); input("Press Enter..."); sys.exit(1)

_UBI       = {3:360,4:480,5:600,6:720,7:840,8:960,9:1080,10:1200,
              11:1320,12:1440,13:1560,14:1680,15:1800}
_CHILD_TAX = {0:0,1:30,2:120,3:360,4:720,5:1200,
              6:1800,7:2520,8:3360,9:4320,10:5400}
SUNSHINE_RATE=600; DEPUTY_PAY=480; DISABILITY_PA=1200

def excel_round(x,d=0):
    f=10**d; return math.floor(abs(x)*f+0.5)*math.copysign(1,x)/f

def as_date(v):
    if isinstance(v,datetime.datetime): return v.date()
    if isinstance(v,datetime.date):    return v
    return None

def calc_ubi(age):
    if age<15: return 0
    key=int(excel_round((age-2)/5)); key=max(3,min(15,key))
    return _UBI.get(key,0)

def compute_n(vals, charity_rate, bh96, today):
    yr  = int(vals.get("year_born") or 1990)
    sex = str(vals.get("sex") or "M")
    dis = bool(vals.get("disabled")); dep = bool(vals.get("deputy"))
    s1  = float(vals.get("sqft1") or 0); s2 = float(vals.get("sqft2") or 0)
    sd  = int(vals.get("hex_sides") or 0); sh = int(vals.get("shared") or 1)
    wm  = float(vals.get("wealth_m") or 0)
    marr= as_date(vals.get("marriage_date")); mort = as_date(vals.get("mort_date"))
    kids= [as_date(vals.get(f"child{i}")) for i in range(1,6)]
    kids= [d for d in kids if d]
    age = today.year - yr

    ubi = calc_ubi(age)
    disability = DISABILITY_PA if dis else 0
    deputy_pay = DEPUTY_PAY if dep else 0

    # Maternity
    mat=0.0
    if sex.upper() in('W','F'):
        for d in kids:
            mo=(today-d).days/30
            if mo<=37: mat+=2400/(mo/6+1)
    mat=round(mat,2)

    # Marriage bonus
    marr_bonus=0.0
    if sex.upper()=='M' and isinstance(marr,datetime.date):
        mo=(today-marr).days/180
        if mo<1: marr_bonus=2400.0
        elif mo<2: marr_bonus=1200.0

    # Mortgage
    mort_assist=0.0
    if isinstance(mort,datetime.date) and today<mort+datetime.timedelta(days=15*365):
        mort_assist=round(bh96,2)

    # Taxes
    nch=len(kids); ct=_CHILD_TAX.get(min(nch,10),0)
    t=s1+s2
    if t>=16000:  prop=round(0.22*t,2)
    elif t>600:   prop=round(((t-600)**0.8/10000)*t,2)
    else:         prop=0.0
    pub=round(t*0.0741,2)
    sun=round(sd*SUNSHINE_RATE/max(1,sh),2)
    wrate=0.0; wt=0.0
    if wm>=1:
        wrate=math.log10(wm)+1
        wt=round(wm*1e6*(wrate/100)/12,2)
    charity=round((prop+ct+sun+wt)*charity_rate/100,2)

    n=round(ubi-ct-sun-prop-wt-charity-pub+mort_assist+marr_bonus+mat+disability+deputy_pay,2)
    return n, {
        "age":age,"ubi":ubi,"mat":mat,"marr":marr_bonus,"mort":mort_assist,
        "dis":disability,"dep":deputy_pay,"prop":prop,"sun":sun,"ct":ct,
        "pub":pub,"wt":wt,"wrate":wrate,"charity":charity,"n":n,
    }

def send_email(subject, body):
    if not SMTP_USER or not SMTP_PASSWORD:
        print("  [EMAIL] SMTP not configured — skipping.")
        print("  Set SMTP_USER / SMTP_PASSWORD at top of monthly_tasks.py")
        return
    try:
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        msg=MIMEMultipart(); msg["From"]=EMAIL_FROM; msg["To"]=EMAIL_TO; msg["Subject"]=subject
        msg.attach(MIMEText(body,"plain"))
        with smtplib.SMTP(SMTP_HOST,SMTP_PORT) as s:
            s.starttls(); s.login(SMTP_USER,SMTP_PASSWORD); s.send_message(msg)
        print(f"  [EMAIL] Report sent to {EMAIL_TO}")
    except Exception as e:
        print(f"  [EMAIL] Failed: {e}")

def main():
    today=datetime.date.today()
    print(f"\n{'='*62}")
    print(f"  ARK Monthly Tasks — {today.strftime('%B %d, %Y')}")
    print(f"  Mode: {'DRY RUN' if DRY_RUN else '*** APPLYING CHANGES ***'}")
    print(f"{'='*62}\n")
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] File not found: {EXCEL_FILE}"); input("Press Enter..."); return

    wb  = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws  = wb["Res"]

    # ── Read constants from source sheets ─────────────────────────────────────
    # A9 = 'Budget History'!E3 + 'Budget History'!F3
    try:
        ws_bh = wb["Budget History"]
        e3 = ws_bh.cell(3,5).value or 0
        f3 = ws_bh.cell(3,6).value or 0
        a9 = float(e3) + float(f3)
    except: a9 = 800.0

    # Charity rate = Charity!W6 or fallback
    # Only apply charity deduction if at least one charity has a BCH address
    try:
        charity_rate_raw = float(wb["Charity"].cell(6,23).value or 11.1111)
        ws_ch = wb["Charity"]
        any_charity_has_bch = False
        for cr in range(2, ws_ch.max_row + 1):
            bch_val = str(ws_ch.cell(cr, 7).value or "").strip()
            if bch_val.startswith("bitcoincash:"):
                any_charity_has_bch = True
                break
        if any_charity_has_bch:
            charity_rate = charity_rate_raw
        else:
            charity_rate = 0.0
            print("  [INFO] No charities have a BCH address — charity deduction set to 0%")
    except: charity_rate = 11.1111

    # BL14 = sum of all ACTIVE gov employee salaries from Gov Employees sheet
    try:
        if "Gov Employees" in wb.sheetnames:
            ws_gov = wb["Gov Employees"]
            bl14 = 0.0
            for r in range(2, ws_gov.max_row + 1):
                active  = ws_gov.cell(r, 8).value   # col H = Active
                salary  = ws_gov.cell(r, 6).value   # col F = Monthly Salary
                if str(active or "").strip().lower() in ("yes","y","true","1") and isinstance(salary,(int,float)) and salary > 0:
                    bl14 += float(salary)
        else:
            # Fallback: Councils sheet K33:K60
            ws_co = wb["Councils"]
            bl14 = sum(float(ws_co.cell(r,11).value) for r in range(33,61)
                       if isinstance(ws_co.cell(r,11).value,(int,float)) and ws_co.cell(r,11).value>0)
    except: bl14 = 870.0

    # Mortgage monthly from BH96 (data_only — this one IS cached in the original file)
    try: bh96 = float(ws.cell(96,60).value or 1081.14)
    except: bh96 = 1081.14

    # A19 current value
    a19 = float(ws.cell(19,1).value or 0.75)

    print(f"  Sin/Elec revenue (A9) : ${a9:,.2f}")
    print(f"  Council budget (BL14) : ${bl14:,.2f}")
    print(f"  Charity rate          : {charity_rate:.4f}%")
    print(f"  Current A19           : {a19:.4f}")
    print(f"  Mortgage assist/mo    : ${bh96:,.2f}\n")

    # ── Read all residents from INPUT columns ─────────────────────────────────
    INP = {"res_num":2,"first_name":6,"last_name":5,"email":4,"alive":16,"bch":53,
           "year_born":8,"sex":11,"disabled":24,"deputy":46,"hex_sides":25,"shared":26,
           "sqft1":39,"sqft2":40,"wealth_m":49,
           "child1":29,"child2":30,"child3":31,"child4":32,"child5":33,
           "marriage_date":36,"mort_date":44,"taxes_owed":97}  # CS=97 matches server.py

    residents=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        rn=row[INP["res_num"]-1]
        if not isinstance(rn,(int,float)): continue
        if row[INP["alive"]-1]==0: continue  # None=active, only 0=dead/banned
        vals={k:row[c-1] for k,c in INP.items()}
        n,bd=compute_n(vals,charity_rate,bh96,today)
        m=round(n*a19,2) if n>0 else n
        residents.append({
            "res_id":int(rn),
            "name":f"{vals.get('first_name') or ''} {vals.get('last_name') or ''}".strip(),
            "email":str(vals.get("email") or "").strip(),
            "bch":  str(vals.get("bch")   or "").strip(),
            "n":n,"m":m,"bd":bd,
            "taxes_owed_now":float(vals.get("taxes_owed") or 0) if isinstance(vals.get("taxes_owed"),(int,float)) else 0.0,
        })
    if not residents:
        print("[WARN] No active residents found."); input("Press Enter..."); return

    # ── Compute A16 from Python-computed charity values ───────────────────────
    a16 = round(sum(r["bd"]["charity"] for r in residents), 2)

    # ── Balanced A19 ──────────────────────────────────────────────────────────
    # A21 = (SUM(M)*-1) - A16 + A9 - BL14 = 0
    # SUM(M) = pos_N*A19 + neg_N  →  A19 = (A9 - A16 - BL14 - sum_neg_N) / sum_pos_N
    sum_pos_n = sum(r["n"] for r in residents if r["n"]>0)
    sum_neg_n = sum(r["n"] for r in residents if r["n"]<0)
    sum_m_cur = sum(r["m"] for r in residents)
    a21_cur   = (-sum_m_cur) - a16 + a9 - bl14

    if sum_pos_n>0:
        new_a19=round(max(0.50,min(1.0,(a9-a16-bl14-sum_neg_n)/sum_pos_n)),6)
    else:
        new_a19=a19

    sum_m_new = sum((r["n"]*new_a19 if r["n"]>0 else r["n"]) for r in residents)
    a21_new   = (-sum_m_new) - a16 + a9 - bl14
    reduction_pct = round((1-new_a19)*100,2)

    print(f"  ── Budget Balance ─────────────────────────────────────────")
    print(f"  Computed A16 (total charity)  = ${a16:>10,.2f}")
    print(f"  Sum positive N (UBI payments) = ${sum_pos_n:>10,.2f}")
    print(f"  Sum negative N (taxes owed)   = ${sum_neg_n:>10,.2f}")
    print(f"  Before: A19={a19:.4f}  A21=${a21_cur:>10,.2f}")
    print(f"  After : A19={new_a19:.4f}  A21=${a21_new:>10,.2f}  (target $0.00)")
    print(f"  Tax payers get a {reduction_pct:.1f}% discount on payments this month\n")

    ubi_rows   =[r for r in residents if r["n"]>0]
    owing_rows =[r for r in residents if r["n"]<0]

    print(f"  ── UBI Recipients ({len(ubi_rows)}) ─────────────────────────────────")
    ubi_total=0.0
    for r in ubi_rows:
        m_new=round(r["n"]*new_a19,2); ubi_total+=m_new
        bch_ok="✓ BCH" if r["bch"].startswith("bitcoincash:") else "  (no BCH)"
        print(f"  #{r['res_id']:4d} {r['name']:20s}  N=${r['n']:8.2f}  pays ${m_new:8.2f}  {bch_ok}")
    print(f"  Total payout: ${ubi_total:,.2f}")

    print(f"\n  ── Tax Accruals ({len(owing_rows)}) ──────────────────────────────────")
    for r in owing_rows:
        new_owed=r["taxes_owed_now"]+abs(r["n"])
        print(f"  #{r['res_id']:4d} {r['name']:20s}  owes ${abs(r['n']):8.2f}/mo  total accrued ${new_owed:,.2f}")

    if DRY_RUN:
        print(f"\n  *** DRY RUN — nothing written ***")
        print(f"  Run:  python monthly_tasks.py --apply")
        input("\nPress Enter to close..."); return

    # ── Write changes ──────────────────────────────────────────────────────────
    wb2=openpyxl.load_workbook(EXCEL_FILE)
    ws2=wb2["Res"]
    ws2.cell(19,1).value=new_a19; ws2.cell(19,1).font=Font(name="Arial",size=10)
    if ws2.cell(1,97).value is None:
        ws2.cell(1,97).value="Taxes Owed"; ws2.cell(1,97).font=Font(name="Arial",size=10,bold=True)
    owed_map={r["res_id"]:r for r in owing_rows}
    for row2 in ws2.iter_rows(min_row=2):
        rn=row2[INP["res_num"]-1].value
        if not isinstance(rn,(int,float)): continue
        if int(rn) not in owed_map: continue
        r=owed_map[int(rn)]
        cur=float(row2[INP["taxes_owed"]-1].value or 0)
        c=row2[INP["taxes_owed"]-1]; c.value=round(cur+abs(r["n"]),2)
        c.font=Font(name="Arial",size=10); c.number_format='#,##0.00'
    try:
        wb2.save(EXCEL_FILE)
        print(f"\n  [OK] A19 updated → {new_a19:.6f}")
        print(f"  [OK] Taxes Owed updated for {len(owing_rows)} residents.")
    except PermissionError:
        print("[ERROR] Close Excel first."); input("Press Enter..."); return

    # ── Build email ────────────────────────────────────────────────────────────
    now=datetime.datetime.now(); ts=now.strftime('%B %d, %Y %H:%M')
    lines=[
        f"ARK COMMUNITY — MONTHLY UBI & BUDGET REPORT",
        f"="*50,
        f"Date          : {ts}",
        f"Period        : {today.strftime('%B %Y')}",
        "",
        f"BUDGET BALANCE",
        f"-"*40,
        f"  Sin+Electric revenue : ${a9:>10,.2f}",
        f"  Charity expenses     : -${a16:>9,.2f}",
        f"  Council budget       : -${bl14:>9,.2f}",
        f"  Previous A19         : {a19:.4f}  ({(1-a19)*100:.1f}% payment discount)",
        f"  Balanced A19         : {new_a19:.4f}  ({reduction_pct:.1f}% payment discount this month)",
        f"  Net Budget (A21)     : ${a21_new:>10,.4f}  (target $0.00)",
        "",
        f"UBI / BENEFIT PAYMENTS — {len(ubi_rows)} residents",
        f"-"*60,
        f"{'Res#':<6}{'Name':<22}{'N':>10}{'Payment':>10}  Breakdown",
        f"-"*60,
    ]
    for r in ubi_rows:
        m_new=round(r["n"]*new_a19,2); bd=r["bd"]
        lines.append(f"#{r['res_id']:<5}{r['name']:<22}${r['n']:>9,.2f} ${m_new:>9,.2f}")
        lines.append(f"      Age={bd['age']} | UBI=${bd['ubi']} Mat=${bd['mat']:.0f} Marr=${bd['marr']:.0f} Mort=${bd['mort']:.0f} Dep=${bd['dep']:.0f} Dis=${bd['dis']:.0f}")
        lines.append(f"      -Prop=${bd['prop']:.0f} -Sun=${bd['sun']:.0f} -Child=${bd['ct']:.0f} -Pub=${bd['pub']:.0f} -Wealth=${bd['wt']:.0f} -Charity=${bd['charity']:.0f}")
        lines.append("")
    lines.append(f"TOTAL UBI PAYOUT: ${ubi_total:,.2f}")
    lines.append("")
    lines.append(f"TAX ACCRUALS — {len(owing_rows)} residents owe community")
    lines.append(f"-"*50)
    for r in owing_rows:
        new_owed=r["taxes_owed_now"]+abs(r["n"])
        lines.append(f"#{r['res_id']:<5}{r['name']:<22}  owes ${abs(r['n']):,.2f}/mo  total: ${new_owed:,.2f}")
    lines.append(f"\n-- Ark Admin System --")
    body="\n".join(lines)
    print(body)
    subject=f"Ark {today.strftime('%b %Y')} Report — A19={new_a19:.4f} ({reduction_pct:.1f}% discount)"
    send_email(subject,body)
    input("\nPress Enter to close...")

if __name__=="__main__":
    main()
