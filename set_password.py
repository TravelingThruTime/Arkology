"""
ARK — Set Resident Password (Admin Tool)
Assigns a hashed password (SHA-256) to any existing resident in column C.
Run via: python set_password.py  or from ARK.bat option 8.
"""
import os, hashlib, getpass
try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("\n[ERROR] openpyxl not found. Run SETUP_FIRST_TIME.bat first.\n"); raise

BASE_DIR   = os.path.dirname(__file__)
EXCEL_FILE = os.path.join(BASE_DIR, "Ark_Database_v6-1.xlsx")

COL_RES_NUM = 2   # B — Res #
COL_PASSWORD = 3  # C — Password hash (used for Citizen Portal login)
COL_FIRST   = 6   # F — First Name
COL_LAST    = 5   # E — Last Name

def hash_pw(p):
    return hashlib.sha256(p.encode()).hexdigest()

def set_password():
    print()
    print("  ══════════════════════════════════════════")
    print("   ARK — Set Resident Password (Admin Tool)")
    print("  ══════════════════════════════════════════")
    print()
    print("  This sets the PASSWORD used for Citizen Portal login.")
    print("  Passwords are SHA-256 hashed before storage.")
    print()

    if not os.path.exists(EXCEL_FILE):
        print(f"  [ERROR] Database not found: {EXCEL_FILE}")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Res"]
    except Exception as e:
        print(f"  [ERROR] Could not open database: {e}")
        return

    res_id_raw = input("  Resident Number: ").strip()
    try:
        res_id = int(res_id_raw)
    except ValueError:
        print("  [ERROR] Invalid resident number.")
        return

    # Find the resident
    target_row = None
    resident_name = ""
    for row in ws.iter_rows(min_row=2):
        rn = row[COL_RES_NUM - 1].value
        if isinstance(rn, (int, float)) and int(rn) == res_id:
            fn = str(row[COL_FIRST - 1].value or "")
            ln = str(row[COL_LAST  - 1].value or "")
            resident_name = f"{fn} {ln}".strip()
            target_row = row[0].row
            break

    if not target_row:
        print(f"  [ERROR] Resident #{res_id} not found.")
        return

    print(f"\n  Found: #{res_id} — {resident_name}")
    print()

    # Get new password
    pw1 = getpass.getpass("  New password: ")
    if len(pw1) < 6:
        print("  [ERROR] Password must be at least 6 characters.")
        return
    pw2 = getpass.getpass("  Confirm password: ")
    if pw1 != pw2:
        print("  [ERROR] Passwords do not match.")
        return

    # Hash and save
    hashed = hash_pw(pw1)
    cell = ws.cell(row=target_row, column=COL_PASSWORD)
    cell.value = hashed
    cell.font = Font(name="Arial", size=10)

    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        print("  [ERROR] Cannot save — close the Excel file first.")
        return

    print(f"\n  [OK] Password set for #{res_id} — {resident_name}")
    print("       They can now log in to the Citizen Portal with this password.")
    print()

if __name__ == "__main__":
    set_password()
