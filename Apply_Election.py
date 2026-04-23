"""
Apply_Election.py
=================
Reads the current Arkology election results from the SQLite database and:
  1. Updates the Gov Employees sheet — sets the 6 top-rated Hexarchy candidates
     to their ranked Hexarchy positions.
  2. Updates the Charity sheet — writes each charity's election rating into the
     current year's rating column.
  3. Stores a snapshot of the results in the previous_elections DB table so
     the /api/previous-election endpoint can serve them.
  4. Resets election votes so the next cycle starts fresh (optional, prompted).

Run from the ark_server directory (same folder as ARK.bat):
    python Apply_Election.py

Or from ARK.bat menu option "17".
"""
import sys
import os
import json
import sqlite3
import datetime
import openpyxl

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE  = os.path.join(BASE_DIR, "Ark_Database_v6-1.xlsx")
ELECTION_DB = os.path.join(BASE_DIR, "arkology_election.db")

COL = {
    "res_num":   2,
    "last_name": 5,
    "first_name":6,
    "alive":     16,
    "bch":       50,
}

# ── Helpers ─────────────────────────────────────────────────────────────────────
def db_connect():
    conn = sqlite3.connect(ELECTION_DB)
    conn.row_factory = sqlite3.Row
    return conn

def get_state(conn, key):
    row = conn.execute("SELECT value FROM election_state WHERE key=?", (key,)).fetchone()
    return row["value"] if row else None

def banner(text):
    print("\n" + "─" * 60)
    print(f"  {text}")
    print("─" * 60)

# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    banner("ARKOLOGY — Apply Election Results")

    if not os.path.exists(ELECTION_DB):
        print("[ERROR] election.db not found. Start the registry server first.")
        sys.exit(1)
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] Spreadsheet not found: {EXCEL_FILE}")
        sys.exit(1)

    conn = db_connect()

    election_name = get_state(conn, "election_name") or "Arkology Election"

    # ── 1. Hexarchy results ────────────────────────────────────────────────────
    banner("HEXARCHY RESULTS")
    hex_rows = conn.execute("""
        SELECT c.id, c.name, c.res_num,
               COALESCE(SUM(v.rating), 0) as rating_sum,
               COUNT(v.id) as vote_count
        FROM candidates c
        LEFT JOIN votes v ON v.target_id=c.id AND v.vote_type='hexarchy'
        WHERE c.active=1
        GROUP BY c.id
        ORDER BY rating_sum DESC
    """).fetchall()

    if not hex_rows:
        print("[WARN] No candidate votes found. Hexarchy will not be updated.")
    else:
        print(f"  Total candidates: {len(hex_rows)}")
        for i, r in enumerate(hex_rows):
            seat = f"SEAT {i+1}" if i < 6 else "  out"
            res  = f"Res#{r['res_num']}" if r['res_num'] else "no res#"
            print(f"  {seat}  {r['name']:<28} ({res})  {r['rating_sum']:.1f} pts / {r['vote_count']} votes")

    # Break ties at the 6th/7th seat boundary
    import random as _random
    if len(hex_rows) > 6:
        # Check if 6th and 7th are tied
        sorted_hex = list(hex_rows)
        sixth_rating = sorted_hex[5]["rating_sum"] if len(sorted_hex) > 5 else None
        seventh_rating = sorted_hex[6]["rating_sum"] if len(sorted_hex) > 6 else None
        if sixth_rating is not None and seventh_rating is not None and sixth_rating == seventh_rating:
            print(f"  [TIE] Seats 6/7 tied at {sixth_rating:.1f} pts — using random tiebreak")
            # Find all tied at boundary
            tied_at = sixth_rating
            before_6 = [r for r in sorted_hex if r["rating_sum"] > tied_at]
            tied_group = [r for r in sorted_hex if r["rating_sum"] == tied_at]
            _random.shuffle(tied_group)
            needed = 6 - len(before_6)
            winners = before_6 + tied_group[:needed]
            losers  = tied_group[needed:]
            # Give winner +0.01 to rating_sum to mark tiebreak
            for r in tied_group[:needed]:
                print(f"    → Tiebreak winner: {r['name']} (random)")
            sorted_hex = winners + losers + [r for r in sorted_hex if r["rating_sum"] < tied_at]
            hex_rows = sorted_hex
    hex_results = [
        {"name": r["name"], "res_num": r["res_num"], "rating_sum": r["rating_sum"], "vote_count": r["vote_count"]}
        for r in hex_rows
    ]

    # ── 2. Charity results ─────────────────────────────────────────────────────
    banner("CHARITY RATINGS")
    total_charity_rating = conn.execute(
        "SELECT COALESCE(SUM(rating), 0) as tot FROM votes WHERE vote_type='charity'"
    ).fetchone()["tot"]

    charity_votes = conn.execute("""
        SELECT target_id, COALESCE(SUM(rating), 0) as total_rating, COUNT(*) as vote_count
        FROM votes WHERE vote_type='charity'
        GROUP BY target_id
    """).fetchall()
    vote_map = {r["target_id"]: {"total_rating": r["total_rating"], "vote_count": r["vote_count"]} for r in charity_votes}

    # Load spreadsheet charities
    wb_read = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_charity = wb_read["Charity"] if "Charity" in wb_read.sheetnames else None
    spreadsheet_charities = []
    if ws_charity:
        for row in range(2, ws_charity.max_row + 1):
            cid   = ws_charity.cell(row, 2).value
            cname = ws_charity.cell(row, 3).value
            if not cid or not cname: continue
            if not isinstance(cid, (int, float)): continue
            spreadsheet_charities.append({"row": row, "id": int(cid), "name": str(cname)})
    wb_read.close()

    year = datetime.date.today().year
    rating_col = 9 + (year - 2025)  # 2025=col9, 2026=col10, ...

    charity_results = []
    for ch in spreadsheet_charities:
        v = vote_map.get(ch["id"], {"total_rating": 0, "vote_count": 0})
        pct = (v["total_rating"] / total_charity_rating * 100) if total_charity_rating > 0 else 0
        # Estimate monthly amount (placeholder — uses same logic as api_charities)
        monthly = (v["total_rating"] / total_charity_rating * total_charity_rating * 12.5) if total_charity_rating > 0 else 0
        charity_results.append({
            "id": ch["id"],
            "name": ch["name"],
            "row": ch["row"],
            "rating": v["total_rating"],
            "vote_count": v["vote_count"],
            "pct": pct,
            "monthly_usd": round(total_charity_rating * 12.5 * pct / 100, 2) if pct > 0 else 0,
        })
        print(f"  {ch['name']:<40} {pct:5.1f}%  rating={v['total_rating']:.1f}  votes={v['vote_count']}")

    # ── Confirm ────────────────────────────────────────────────────────────────
    banner("CONFIRM APPLICATION")
    print("  This will:")
    if hex_rows:
        print("  • Update Gov Employees sheet — set top 6 to Hexarchy positions")
    print(f"  • Write charity ratings to spreadsheet column {rating_col} (year {year})")
    print("  • Save results to previous_elections database table")
    print()
    confirm = input("  Apply these results? Type YES to confirm: ").strip()
    if confirm.upper() != "YES":
        print("\n  Aborted — no changes made.")
        return

    # ── 3. Write to spreadsheet ────────────────────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Update charity ratings
    if "Charity" in wb.sheetnames:
        ws_c = wb["Charity"]
        for ch in charity_results:
            ws_c.cell(ch["row"], rating_col).value = round(ch["rating"], 4)
        print(f"\n  [OK] Charity ratings written to column {rating_col} (year {year})")

    # Update Gov Employees — Hexarchy positions
    if hex_rows and "Gov Employees" in wb.sheetnames:
        ws_emp = wb["Gov Employees"]
        positions = [
            "Hexarchy (1st Place)", "Hexarchy (2nd Place)", "Hexarchy (3rd Place)",
            "Hexarchy (4th Place)", "Hexarchy (5th Place)", "Hexarchy (6th Place)",
        ]
        today_str = datetime.date.today().strftime("%Y-%m-%d")

        # Step 1: DELETE all existing hexarchy rows (history is preserved in previous_elections DB)
        rows_to_delete = []
        for row in ws_emp.iter_rows(min_row=2):
            pos = str(row[4].value or "").strip()
            if "Hexarchy" in pos:
                rows_to_delete.append(row[0].row)
        # Delete from bottom up so row numbers don't shift
        for row_num in sorted(rows_to_delete, reverse=True):
            ws_emp.delete_rows(row_num)

        # Step 2: Find next emp_num and insert new hexarchy rows
        max_emp_num = max(
            (int(row[0]) for row in ws_emp.iter_rows(min_row=2, values_only=True)
             if row[0] and isinstance(row[0], (int, float))),
            default=0
        )

        for i, cand in enumerate(list(hex_rows)[:6]):
            pos = positions[i]
            res_num = cand["res_num"]
            first_name = str(cand["name"]).split()[0] if cand["name"] else ""
            last_name = " ".join(str(cand["name"]).split()[1:]) if cand["name"] else ""
            max_emp_num += 1
            ws_emp.append([max_emp_num, res_num, last_name, first_name, pos, 72 - i*6, today_str, "Yes", "Yes"])
            print(f"  [OK] {pos}: {cand['name']} (Res#{res_num})")

    wb.save(EXCEL_FILE)
    print(f"\n  [OK] Spreadsheet saved: {EXCEL_FILE}")

    # ── 4. Store in previous_elections ────────────────────────────────────────
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS previous_elections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                election_name TEXT NOT NULL,
                applied_at TEXT DEFAULT (datetime('now')),
                hexarchy_results TEXT,
                charity_results TEXT,
                proposal_results TEXT
            )
        """)
        # Fetch proposal results from votes
        prop_results = []
        try:
            for prop in conn.execute("SELECT * FROM proposals WHERE active=1 ORDER BY id").fetchall():
                p = dict(prop)
                import json as _j
                options = _j.loads(p.get("options") or "null")
                if options:
                    # Alliance/rated: sum ratings per option, with tiebreak
                    import random as _rand
                    opts_data = []
                    for opt in options:
                        rows_opt = conn.execute(
                            "SELECT COUNT(*) as cnt, COALESCE(SUM(rating),0) as total "
                            "FROM alliance_votes WHERE proposal_id=? AND option_name=?",
                            (p["id"], opt)).fetchone()
                        cnt = rows_opt["cnt"] if rows_opt else 0
                        total = rows_opt["total"] if rows_opt else 0
                        opts_data.append({"option": opt, "total_rating": float(total), "vote_count": cnt,
                                          "avg": total/cnt if cnt > 0 else 0})
                    opts_data.sort(key=lambda x: x["total_rating"], reverse=True)
                    # Tiebreak: if top two are tied, random winner gets +0.01
                    if len(opts_data) >= 2 and opts_data[0]["total_rating"] == opts_data[1]["total_rating"]:
                        tied_top = opts_data[0]["total_rating"]
                        tied_group = [o for o in opts_data if o["total_rating"] == tied_top]
                        winner_opt = _rand.choice(tied_group)
                        winner_opt["total_rating"] += 0.01
                        winner_opt["tiebreak_winner"] = True
                        print(f"  [TIE] Alliance tiebreak: {winner_opt['option']} wins (random)")
                        opts_data.sort(key=lambda x: x["total_rating"], reverse=True)
                    winner = opts_data[0]["option"] if opts_data else None
                    prop_results.append({"id": p["id"], "title": p["title"], "type": p["type"],
                                         "threshold": p.get("threshold"), "options": opts_data,
                                         "winner": winner, "passed": True})
                else:
                    yes = conn.execute("SELECT COUNT(*) FROM votes WHERE vote_type='democracy' AND target_id=? AND choice='yes'", (p["id"],)).fetchone()[0]
                    no = conn.execute("SELECT COUNT(*) FROM votes WHERE vote_type='democracy' AND target_id=? AND choice='no'", (p["id"],)).fetchone()[0]
                    total = yes + no
                    threshold_val = p.get("threshold_val") or 0.667
                    passed = (yes / total >= threshold_val) if total > 0 else False
                    prop_results.append({"id": p["id"], "title": p["title"], "type": p["type"],
                                         "threshold": p.get("threshold"), "yes_votes": yes, "no_votes": no,
                                         "passed": passed})
        except Exception as ep:
            print(f"  [WARN] Could not fetch proposal results: {ep}")

        conn.execute("""
            INSERT INTO previous_elections (election_name, hexarchy_results, charity_results, proposal_results)
            VALUES (?, ?, ?, ?)
        """, (
            election_name,
            json.dumps(hex_results),
            json.dumps([{k: v for k, v in c.items() if k != "row"} for c in charity_results]),
            json.dumps(prop_results),
        ))
        conn.commit()
        print("  [OK] Results saved to previous_elections table")
    except Exception as e:
        print(f"  [WARN] Could not save to DB: {e}")

    # ── 5. Optional: reset votes ───────────────────────────────────────────────
    print()
    reset = input("  Reset all election votes for the next cycle? (Y/N): ").strip().upper()
    if reset == "Y":
        conn.execute("DELETE FROM votes")
        conn.execute("DELETE FROM alliance_votes")
        conn.execute("DELETE FROM voter_sessions")
        conn.execute("DELETE FROM resident_ballots")
        conn.execute("UPDATE election_state SET value='0' WHERE key IN ('hexarchy_open','charity_open','democracy_open')")
        conn.commit()
        print("  [OK] Votes reset. Election sections closed.")

    conn.close()
    banner("DONE — Election results applied successfully")

if __name__ == "__main__":
    main()
