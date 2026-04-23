#!/usr/bin/env python3
"""
CHARITY_PAYMENTS.py  —  Ark Community Charity Budget Calculator
================================================================
Reads the Charity sheet from the database and calculates each
charity's monthly budget based on their election ratings.

Formula: charity_budget = (charity_rating / sum_all_ratings) * total_charity_market

The total charity market budget is the sum of all charity taxes collected
from residents this month (computed from the charity tax rate applied to
property tax, child tax, sunshine tax, and wealth tax).

Usage:
  python CHARITY_PAYMENTS.py         -- compute and display budgets
  python CHARITY_PAYMENTS.py --dry   -- display only, no write
"""

import os, sys, datetime

DRY_RUN = "--dry" in sys.argv
EXCEL   = os.path.join(os.path.dirname(__file__), "Ark_Database_v6-1.xlsx")

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("[ERROR] openpyxl not installed.")
    input("Press Enter..."); sys.exit(1)


def main():
    today = datetime.date.today()
    year  = today.year
    print()
    print("=" * 62)
    print("  ARK COMMUNITY — CHARITY BUDGET CALCULATOR")
    print(f"  {today.strftime('%B %d, %Y')}")
    print(f"  Mode: {'DRY RUN' if DRY_RUN else 'LIVE — will write budgets'}")
    print("=" * 62)
    print()

    if not os.path.exists(EXCEL):
        print(f"[ERROR] File not found: {EXCEL}")
        input("Press Enter..."); return

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["Charity"]

    # Determine rating column for current year
    rating_col = 9 + (year - 2025)  # 2025=col9, 2026=col10, ...
    budget_col = 16 + (year - 2025)  # 2025=col16, 2026=col17, ...

    print(f"  Rating column: {rating_col} (year {year})")
    print(f"  Budget column: {budget_col}")
    print()

    # Read charities
    charities = []
    for r in range(2, ws.max_row + 1):
        cid  = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        if not cid or not name: continue
        try:
            cid_int = int(cid)
        except (ValueError, TypeError):
            continue  # skip header row or any non-integer ID
        _emp_raw = ws.cell(r, 4).value
        try:
            emp = int(_emp_raw) if isinstance(_emp_raw, (int, float)) else 0
        except (ValueError, TypeError):
            emp = 0
        rating = float(ws.cell(r, rating_col).value or 0)
        bch    = str(ws.cell(r, 7).value or "").strip()
        charities.append({"row": r, "id": cid_int, "name": str(name),
                          "employees": emp, "rating": rating, "bch": bch})

    total_ratings = sum(c["rating"] for c in charities)
    print(f"  Charities found: {len(charities)}")
    print(f"  Total ratings:   {total_ratings}")

    # Compute total charity market from Res sheet resident charity taxes
    # charity_tax per resident = (prop + child_tax + sun + wealth_tax) * charity_rate / 100
    import math as _math
    CHILD_TAX_LOCAL = {0:0,1:30,2:120,3:360,4:720,5:1200,6:1800,7:2520,8:3360,9:4320,10:5400}
    try:
        charity_rate = float(wb["Charity"].cell(6, 23).value or 11.1111)
    except:
        charity_rate = 11.1111

    try:
        ws_r = wb["Res"]
        total_market = 0.0
        today_yr = datetime.date.today().year
        for row in ws_r.iter_rows(min_row=2, values_only=True):
            rn    = row[1]   # col B
            alive = row[15]  # col P
            if not isinstance(rn, (int, float)): continue
            if alive == 0: continue  # skip deceased/banned
            s1    = float(row[38] or 0); s2 = float(row[39] or 0)  # sqft1/sqft2 (cols 39,40)
            t     = s1 + s2
            # Property tax
            if t >= 16000:  prop = round(0.22 * t, 2)
            elif t > 600:   prop = round(((t - 600) ** 0.8 / 10000) * t, 2)
            else:           prop = 0.0
            # Child tax
            nch = sum(1 for k in [28, 29, 30, 31, 32] if row[k] is not None)  # child cols 29-33
            ct = CHILD_TAX_LOCAL.get(min(nch, 10), 0)
            # Sunlight tax
            sides  = int(row[24] or 0); shared = max(1, int(row[25] or 1))
            sun    = round(sides * 600 / shared, 2)
            # Wealth tax
            wm  = float(row[48] or 0)  # col 49
            wt  = round(wm * 1e6 * (_math.log10(wm) + 1) / 100 / 12, 2) if wm >= 1 else 0.0
            # Charity tax
            total_market += round((prop + ct + sun + wt) * charity_rate / 100, 2)
        total_market = round(total_market, 2)
        if total_market > 0:
            print(f"  Monthly charity market: ${total_market:,.2f}  (computed from resident taxes)")
        else:
            total_market = 870.0
            print(f"  Monthly charity market: ${total_market:,.2f}  (fallback estimate)")
    except Exception as e:
        total_market = 870.0
        print(f"  Monthly charity market: ${total_market:,.2f}  (fallback — {e})")
    print()

    wb.close()

    # Calculate budgets
    print(f"  {'Charity':<20} {'Rating':>8} {'%':>8} {'Monthly $':>12} {'Employees':>10}  BCH")
    print(f"  {'-'*20} {'-'*8} {'-'*8} {'-'*12} {'-'*10}  {'-'*15}")

    budgets = []
    for c in sorted(charities, key=lambda x: x["rating"], reverse=True):
        if total_ratings > 0:
            pct   = c["rating"] / total_ratings
            budget = round(pct * total_market, 2)
        else:
            pct    = 0
            budget = 0

        addr = c["bch"][:15] + "…" if len(c["bch"]) > 15 else (c["bch"] or "— none —")
        print(f"  {c['name']:<20} {c['rating']:>8.0f} {pct*100:>7.1f}% ${budget:>10,.2f} {c['employees']:>10}  {addr}")
        budgets.append({**c, "pct": pct, "budget": budget})

    total_budget = sum(b["budget"] for b in budgets)
    print(f"  {'TOTAL':<20} {total_ratings:>8.0f} {'100.0%':>8} ${total_budget:>10,.2f}")
    print()

    if DRY_RUN:
        print("  [DRY RUN] Budgets not written. Run without --dry to apply.")
    else:
        try:
            wb2 = openpyxl.load_workbook(EXCEL)
            ws2 = wb2["Charity"]
            fnt = Font(name="Arial", size=10)
            for b in budgets:
                ws2.cell(b["row"], budget_col).value = b["budget"]
                ws2.cell(b["row"], budget_col).font = fnt
                ws2.cell(b["row"], 5).value = b["budget"] * 12  # Annual budget
                ws2.cell(b["row"], 5).font = fnt
                if b["employees"] > 0:
                    ws2.cell(b["row"], 6).value = round(b["budget"] / b["employees"], 2)
                    ws2.cell(b["row"], 6).font = fnt
            wb2.save(EXCEL)
            print(f"  [OK] Budgets written to column {budget_col} for year {year}.")
        except PermissionError:
            print("  [ERROR] Close Excel first, then run again.")
        except Exception as e:
            print(f"  [ERROR] {e}")

    print()
    input("Press Enter to close...")


if __name__ == "__main__":
    main()
