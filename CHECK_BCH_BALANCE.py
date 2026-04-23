#!/usr/bin/env python3
"""
CHECK_BCH_BALANCE.py  -  Ark Community A19 Updater
====================================================
Uses ONLY viable residents (alive=1 + ref_checked=1 + immigration=1).
Includes charities and gov employees WITH BCH addresses in A19 calculation.

A19 = (revenue_in + |taxes_in|) / (UBI_BCH_out + charity_BCH + salary_BCH)

Usage:
  python CHECK_BCH_BALANCE.py         -- update A19
  python CHECK_BCH_BALANCE.py --dry   -- report only, no write
"""
import os, sys, datetime, math, urllib.request, json

DRY_RUN            = "--dry" in sys.argv
EXCEL_FILE         = os.path.join(os.path.dirname(__file__), "Ark_Database_v6-1.xlsx")
BCH_SYSTEM_ADDRESS = "bitcoincash:qqvvwjw8rns37te6r4a0jzvtu86s3kx6lsykymqj5k"
TX_FEE_RESERVE     = 0.99

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("[ERROR] openpyxl not installed."); input("Press Enter..."); sys.exit(1)

# Column indices (1-based)
COL = {
    "res_num":2,"first_name":6,"last_name":5,"alive":16,"bch":50,
    "ref_checked":96,"immigration":100,
    "year_born":8,"sex":11,"disabled":24,"deputy":46,
    "hex_sides":25,"shared":26,"sqft1":39,"sqft2":40,"wealth_m":49,
    "child1":29,"child2":30,"child3":31,"child4":32,"child5":33,
    "marriage_date":36,"mort_date":44,
}

_UBI={3:360,4:480,5:600,6:720,7:840,8:960,9:1080,10:1200,
      11:1320,12:1440,13:1560,14:1680,15:1800}
_CHILD_TAX={0:0,1:30,2:120,3:360,4:720,5:1200,6:1800,7:2520,8:3360,9:4320,10:5400}
SUNSHINE_RATE=600; DEPUTY_PAY=480; DISABILITY_PA=1200

def excel_round(x,d=0):
    f=10**d; return math.floor(abs(x)*f+0.5)*math.copysign(1,x)/f

def as_date(v):
    if isinstance(v,datetime.datetime): return v.date()
    if isinstance(v,datetime.date): return v
    return None

def compute_n(vals,charity_rate,bh96,today):
    yr=int(vals.get("year_born") or 1990)
    sex=str(vals.get("sex") or "M"); dis=bool(vals.get("disabled")); dep=bool(vals.get("deputy"))
    s1=float(vals.get("sqft1") or 0); s2=float(vals.get("sqft2") or 0)
    sd=int(vals.get("hex_sides") or 0); sh=int(vals.get("shared") or 1)
    wm=float(vals.get("wealth_m") or 0)
    mort=as_date(vals.get("mort_date"))
    kids=[as_date(vals.get(f"child{i}")) for i in range(1,6)]
    kids=[d for d in kids if d]
    age=today.year-yr
    if age<15: ubi=0
    else:
        key=int(excel_round((age-2)/5)); key=max(3,min(15,key))
        ubi=_UBI.get(key,0)
    dis_pay=DISABILITY_PA if dis else 0; dep_pay=DEPUTY_PAY if dep else 0
    mat=0.0
    if sex.upper() in('W','F'):
        for d in kids:
            mo=(today-d).days/30
            if mo<=37: mat+=2400/(mo/6+1)
    mat=round(mat,2)
    mort_assist=round(bh96,2) if(mort and today<mort+datetime.timedelta(days=15*365)) else 0.0
    nch=len(kids); ct=_CHILD_TAX.get(min(nch,10),0)
    t=s1+s2
    if t>=16000: prop=round(0.22*t,2)
    elif t>600:  prop=round(((t-600)**0.8/10000)*t,2)
    else:        prop=0.0
    pub=round(t*0.0741,2); sun=round(sd*SUNSHINE_RATE/max(1,sh),2)
    wt=0.0
    if wm>=1: wt=round(wm*1e6*(math.log10(wm)+1)/100/12,2)
    charity=round((prop+ct+sun+wt)*charity_rate/100,2)
    return round(ubi-ct-sun-prop-wt-charity-pub+mort_assist+mat+dis_pay+dep_pay,2)

def get_bch_rate():
    try:
        url="https://api.coingecko.com/api/v3/simple/price?ids=bitcoin-cash&vs_currencies=usd"
        req=urllib.request.Request(url,headers={"User-Agent":"ArkRegistry/1.0"})
        with urllib.request.urlopen(req,timeout=8) as r:
            data=json.loads(r.read())
        rate=float(data["bitcoin-cash"]["usd"])
        print(f"  BCH/USD rate: ${rate:,.2f}  (CoinGecko live)"); return rate
    except Exception as e:
        print(f"  [WARN] CoinGecko failed ({e}), using fallback $462"); return 462.0

def get_wallet_balance_bch():
    apis=[
        f"https://api.blockchain.info/haskoin-store/bch/address/{BCH_SYSTEM_ADDRESS}/balance",
        f"https://bch-chain.api.btc.com/v3/address/{BCH_SYSTEM_ADDRESS}",
    ]
    for url in apis:
        try:
            req=urllib.request.Request(url,headers={"User-Agent":"ArkRegistry/1.0","Accept":"application/json"})
            with urllib.request.urlopen(req,timeout=10) as r:
                data=json.loads(r.read())
            if "confirmed" in data:
                return(int(data.get("confirmed",0))+int(data.get("unconfirmed",0)))/1e8
            if "data" in data and "balance" in data["data"]:
                return int(data["data"]["balance"])/1e8
        except Exception as e:
            print(f"  [WARN] Wallet API ({url.split('/')[2]}): {e}"); continue
    raise Exception("All wallet APIs failed")

def load_charity_bch(wb):
    charities=[]; total=0.0
    try:
        ws=wb["Charity"]; year=datetime.date.today().year
        budget_col=16+(year-2025)
        for r in range(2,ws.max_row+1):
            cid=ws.cell(r,2).value; name=ws.cell(r,3).value
            bch=str(ws.cell(r,7).value or "").strip()
            budget=ws.cell(r,budget_col).value
            if not cid or not name: continue
            if not bch.startswith("bitcoincash:"): continue
            if not isinstance(budget,(int,float)) or budget<=0: continue
            total+=float(budget)
            charities.append({"name":str(name),"bch":bch,"budget":float(budget)})
    except Exception as e: print(f"  [WARN] Charity sheet: {e}")
    return total,charities

def load_employee_bch(wb):
    employees=[]; total=0.0
    try:
        ws_emp=wb["Gov Employees"]; ws_res=wb["Res"]
        res_bch={}
        for row in ws_res.iter_rows(min_row=2,values_only=True):
            rn=row[COL["res_num"]-1]; bch=row[COL["bch"]-1]
            if isinstance(rn,(int,float)) and bch and str(bch).strip().startswith("bitcoincash:"):
                res_bch[int(rn)]=str(bch).strip()
        for row in ws_emp.iter_rows(min_row=2,values_only=True):
            emp_num,res_num,last,first,position,salary,date_hired,active,hexarchy=(row+(None,)*9)[:9]
            if not isinstance(salary,(int,float)) or salary<=0: continue
            if str(active or "").strip().lower() not in("yes","y","true","1"): continue
            if not isinstance(res_num,(int,float)): continue
            ri=int(res_num)
            if ri not in res_bch: continue
            total+=float(salary)
            name=f"{first or ''} {last or ''}".strip() or f"Res#{ri}"
            employees.append({"name":name,"res_num":ri,"bch":res_bch[ri],
                              "salary":float(salary),"position":str(position or "")})
    except Exception as e: print(f"  [WARN] Gov Employees sheet: {e}")
    return total,employees

def main():
    print(); print("="*65)
    print("  ARK COMMUNITY - BCH BALANCE CHECK + A19 UPDATER")
    print(f"  {datetime.date.today().strftime('%B %d, %Y')}")
    print("  Mode:", "DRY RUN" if DRY_RUN else "LIVE - will write A19")
    print("  Only VIABLE residents: alive + ref checked + immigration passed")
    print("="*65); print()

    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] {EXCEL_FILE}"); input("Press Enter..."); return

    wb=openpyxl.load_workbook(EXCEL_FILE,data_only=True)
    ws=wb["Res"]; today=datetime.date.today()
    a19_current=float(ws.cell(19,1).value or 0.75)
    try: charity_rate=float(wb["Charity"].cell(6,23).value or 11.1111)
    except: charity_rate=11.1111
    try: bh96=float(ws.cell(96,60).value or 1081.14)
    except: bh96=1081.14
    try:
        ws_bh=wb["Budget History"]
        a9=float(ws_bh.cell(3,5).value or 0)+float(ws_bh.cell(3,6).value or 0)
    except: a9=800.0

    # Load viable residents
    viable_pos=[]; viable_neg=[]; skipped=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        rn=row[COL["res_num"]-1]
        if not isinstance(rn,(int,float)): continue
        alive=row[COL["alive"]-1]
        ref=row[COL["ref_checked"]-1]
        immig=row[COL["immigration"]-1]
        bch=str(row[COL["bch"]-1] or "").strip()
        first=str(row[COL["first_name"]-1] or ""); last=str(row[COL["last_name"]-1] or "")
        name=f"{first} {last}".strip() or f"Res#{int(rn)}"
        if alive==0: continue
        if ref!=1:   skipped.append((int(rn),name,"ref not checked")); continue
        if immig!=1: skipped.append((int(rn),name,"immigration not passed")); continue
        vals={k:row[c-1] for k,c in COL.items()}
        n=compute_n(vals,charity_rate,bh96,today)
        rec={"res_id":int(rn),"name":name,"bch":bch,"has_bch":bch.startswith("bitcoincash:"),"n":n}
        if n>0: viable_pos.append(rec)
        elif n<0: viable_neg.append(rec)

    charity_total,charities_bch=load_charity_bch(wb)
    employee_total,employees_bch=load_employee_bch(wb)
    wb.close()

    sum_pos=sum(r["n"] for r in viable_pos)
    sum_neg=sum(r["n"] for r in viable_neg)  # negative
    payable=[r for r in viable_pos if r["has_bch"]]
    sum_payable=sum(r["n"] for r in payable)

    total_denom=sum_payable+charity_total+employee_total
    total_numer=a9+abs(sum_neg)

    print(f"  Viable UBI receivers:  {len(viable_pos)} (sum N=${sum_pos:,.2f})")
    print(f"  Viable tax payers:     {len(viable_neg)} (taxes in=${abs(sum_neg):,.2f})")
    print(f"  BCH-payable residents: {len(payable)} (sum N=${sum_payable:,.2f})")
    if skipped: print(f"  Skipped (not viable):  {len(skipped)}")
    print()
    print(f"  UBI payments (BCH):  ${sum_payable:>10,.2f}")
    print(f"  Charity BCH budgets: ${charity_total:>10,.2f}  ({len(charities_bch)} charities)")
    print(f"  Employee BCH salaries:${employee_total:>10,.2f}  ({len(employees_bch)} employees)")
    print(f"  Total outflow denom: ${total_denom:>10,.2f}")
    print()
    print(f"  Revenue A9:          ${a9:>10,.2f}")
    print(f"  Taxes in:            ${abs(sum_neg):>10,.2f}")
    print(f"  Total inflow numer:  ${total_numer:>10,.2f}")
    print()

    if total_denom>0:
        budget_a19=round(max(0.0,min(1.0,total_numer/total_denom)),6)
    else:
        budget_a19=a19_current
    print(f"  Budget-balanced A19: {budget_a19:.6f}  ({budget_a19*100:.2f}%)")

    bch_rate=get_bch_rate(); wallet_bch=0.0; bch_ok=False
    try:
        wallet_bch=get_wallet_balance_bch()
        print(f"  Wallet: {wallet_bch:.5f} BCH  (${wallet_bch*bch_rate:,.2f})")
        bch_ok=True
    except Exception as e:
        print(f"  [WARN] Wallet fetch failed: {e}")
        print("  [NOTE] Using budget A19 as fallback (wallet unknown).")

    if bch_ok and total_denom>0:
        avail_usd=wallet_bch*TX_FEE_RESERVE*bch_rate
        bch_cap_a19=round(max(0.0,min(1.0,avail_usd/total_denom)),6)
        print(f"  BCH-capacity A19: {bch_cap_a19:.6f}  ({bch_cap_a19*100:.2f}%)")
    else:
        bch_cap_a19=budget_a19

    new_a19=round(min(budget_a19,bch_cap_a19),6)
    print()
    print(f"  Final A19: {new_a19:.6f}  ({new_a19*100:.2f}%)")
    if new_a19!=a19_current:
        d="UP" if new_a19>a19_current else "DOWN"
        print(f"  Change: {d}  {a19_current:.6f} -> {new_a19:.6f}")

    # Payment table
    print(); print(f"  -- Payments at A19={new_a19:.4f} --")
    total_out=0.0
    for r in viable_pos:
        pay=round(r["n"]*new_a19,2) if r["has_bch"] else 0
        bch_amt=pay/bch_rate; total_out+=pay
        addr=r["bch"][12:20]+"..." if r["has_bch"] else "- no BCH -"
        print(f"  UBI  #{r['res_id']:<4} {r['name']:<20} N=${r['n']:>8.2f} pay=${pay:>8.2f} {bch_amt:.5f}BCH  {addr}")
    for c in charities_bch:
        pay=round(c["budget"]*new_a19,2); bch_amt=pay/bch_rate; total_out+=pay
        print(f"  CHR  {'':5} {c['name']:<20} budget=${c['budget']:>6.2f} pay=${pay:>8.2f} {bch_amt:.5f}BCH")
    for e in employees_bch:
        pay=round(e["salary"]*new_a19,2); bch_amt=pay/bch_rate; total_out+=pay
        print(f"  EMP  #{e['res_num']:<4} {e['name']:<20} sal=${e['salary']:>8.2f} pay=${pay:>8.2f} {bch_amt:.5f}BCH")
    print(f"  TOTAL OUT: ${total_out:,.2f}  /  {total_out/bch_rate:.5f} BCH")
    if bch_ok:
        surplus=wallet_bch*TX_FEE_RESERVE-total_out/bch_rate
        print(f"  After payout: {surplus:+.5f} BCH  [{'OK' if surplus>=0 else 'INSUFFICIENT'}]")

    print()
    if DRY_RUN:
        print("  [DRY RUN] A19 not written.")
    elif new_a19==a19_current:
        print(f"  A19 unchanged - no write needed.")
    else:
        try:
            wb2=openpyxl.load_workbook(EXCEL_FILE)
            wb2["Res"].cell(19,1).value=new_a19
            wb2["Res"].cell(19,1).font=Font(name="Arial",size=10)
            if "Budget History" in wb2.sheetnames:
                ws_bh=wb2["Budget History"]
                td=datetime.date.today(); this_year,this_month=td.year,td.month
                target_row=None
                for r in range(2,ws_bh.max_row+2):
                    yr=ws_bh.cell(r,1).value; mo=ws_bh.cell(r,2).value
                    if yr is None: target_row=r; break
                    if int(yr)==this_year and int(mo)==this_month: target_row=r; break
                if target_row is None: target_row=ws_bh.max_row+1
                n_living=len(viable_pos)
                total_bud=round(sum(r["n"]*new_a19 for r in payable),2)
                per_cap=round(total_bud/n_living,2) if n_living>0 else 0
                label=td.strftime("%b %Y"); fnt=Font(name="Arial",size=10)
                ws_bh.cell(target_row,1,this_year).font=fnt
                ws_bh.cell(target_row,2,this_month).font=fnt
                ws_bh.cell(target_row,3,label).font=fnt
                ws_bh.cell(target_row,4,new_a19).font=fnt  # store 0-1 directly
                ws_bh.cell(target_row,5,total_bud).font=fnt
                ws_bh.cell(target_row,6,n_living).font=fnt
                ws_bh.cell(target_row,7,per_cap).font=fnt
            wb2.save(EXCEL_FILE)
            print(f"  [OK] A19 written: {a19_current:.6f} -> {new_a19:.6f}")
        except PermissionError: print("  [ERROR] Close Excel first.")
        except Exception as e: print(f"  [ERROR] {e}")

    print(); input("Press Enter to close...")

if __name__=="__main__": main()
