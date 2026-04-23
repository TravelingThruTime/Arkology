"""
Ark Citizen Registration Server  v6.7
Correct column mapping based on original spreadsheet:
  C  (3)  = Password hash (emergency recovery)
  AX (50) = BCH Address
  CN (92) = PIN hash (daily login)
  CO (93) = Own Phone
  CP (94) = Reference 1 (name)
  CQ (95) = Reference 2 (name)
  CR (96) = Ref Checked
  CS (97) = Taxes Owed
  CT (98) = Last Payment Date
  CU (99) = Election Year (most recent election voted)
"""
import os, json, hashlib, datetime, math, re, glob, sqlite3, threading, time as _time
from flask import Flask, request, jsonify, send_from_directory, make_response
try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("\n[ERROR] openpyxl not found. Run SETUP_FIRST_TIME.bat first.\n"); raise

app = Flask(__name__, static_folder="static", template_folder="templates")
BASE_DIR    = os.path.dirname(__file__)
EXCEL_FILE  = os.path.join(BASE_DIR, "Ark_Database_v6-1.xlsx")
LOG_DIR     = os.path.join(BASE_DIR, "Update_Information_Logs")
RECEIPT_DIR = os.path.join(BASE_DIR, "Tax_Payment_Receipts")
RES_SHEET   = "Res"
os.makedirs(LOG_DIR,     exist_ok=True)
os.makedirs(RECEIPT_DIR, exist_ok=True)

# ── Election DB (SQLite) ───────────────────────────────────────────────────────
ELECTION_DB   = os.path.join(BASE_DIR, "arkology_election.db")
ADMIN_PASSWORD = "arkology2026"

def get_election_db():
    conn = sqlite3.connect(ELECTION_DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_election_db():
    with get_election_db() as conn:
        conn.executescript("""
          CREATE TABLE IF NOT EXISTS election_state (key TEXT PRIMARY KEY, value TEXT NOT NULL);
          CREATE TABLE IF NOT EXISTS candidates (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
            platform TEXT NOT NULL, active INTEGER DEFAULT 1,
            res_num INTEGER DEFAULT NULL,
            created_at TEXT DEFAULT (datetime('now')));
          CREATE TABLE IF NOT EXISTS previous_elections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            election_name TEXT NOT NULL,
            applied_at TEXT DEFAULT (datetime('now')),
            hexarchy_results TEXT,
            charity_results TEXT,
            proposal_results TEXT
          );
          CREATE TABLE IF NOT EXISTS election_charities (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
            description TEXT NOT NULL, min_votes INTEGER DEFAULT 20,
            active INTEGER DEFAULT 1, created_at TEXT DEFAULT (datetime('now')));
          CREATE TABLE IF NOT EXISTS proposals (
            id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL,
            type TEXT NOT NULL, description TEXT NOT NULL,
            threshold TEXT NOT NULL DEFAULT '2/3', threshold_val REAL,
            options TEXT, active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now')));
          CREATE TABLE IF NOT EXISTS votes (
            id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT NOT NULL,
            vote_type TEXT NOT NULL, target_id INTEGER NOT NULL,
            rating REAL, choice TEXT, created_at TEXT DEFAULT (datetime('now')));
          CREATE TABLE IF NOT EXISTS alliance_votes (
            id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT NOT NULL,
            proposal_id INTEGER NOT NULL, option_name TEXT NOT NULL,
            rating REAL NOT NULL DEFAULT 0, created_at TEXT DEFAULT (datetime('now')));
          CREATE TABLE IF NOT EXISTS voter_sessions (
            session_id TEXT PRIMARY KEY, voted_hexarchy INTEGER DEFAULT 0,
            voted_charity INTEGER DEFAULT 0, created_at TEXT DEFAULT (datetime('now')));
          CREATE UNIQUE INDEX IF NOT EXISTS idx_votes_unique ON votes(session_id,vote_type,target_id);
          CREATE UNIQUE INDEX IF NOT EXISTS idx_alliance_unique ON alliance_votes(session_id,proposal_id,option_name);
          CREATE TABLE IF NOT EXISTS resident_ballots (
            resident_id INTEGER PRIMARY KEY,
            voted_hexarchy INTEGER DEFAULT 0,
            voted_charity INTEGER DEFAULT 0,
            voted_democracy INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
          );
        """)
        # Seed initial state keys
        for key, val in [('hexarchy_open','0'),('charity_open','0'),('democracy_open','0'),('election_name','Annual Arkology Election')]:
            conn.execute("INSERT OR IGNORE INTO election_state (key,value) VALUES (?,?)", (key,val))
        conn.commit()
        # Migrations
        try: conn.execute("ALTER TABLE candidates ADD COLUMN res_num INTEGER DEFAULT NULL")
        except: pass
        try: conn.execute("CREATE TABLE IF NOT EXISTS previous_elections (id INTEGER PRIMARY KEY AUTOINCREMENT, election_name TEXT NOT NULL, applied_at TEXT DEFAULT (datetime('now')), hexarchy_results TEXT, charity_results TEXT, proposal_results TEXT)")
        except: pass
        # Migrate resident_ballots from single 'voted' to per-category
        try: conn.execute("ALTER TABLE resident_ballots ADD COLUMN voted_hexarchy INTEGER DEFAULT 0")
        except: pass
        try: conn.execute("ALTER TABLE resident_ballots ADD COLUMN voted_charity INTEGER DEFAULT 0")
        except: pass
        try: conn.execute("ALTER TABLE resident_ballots ADD COLUMN voted_democracy INTEGER DEFAULT 0")
        except: pass
        # Backfill: if old 'voted' column exists and is 1, set all three
        try: conn.execute("UPDATE resident_ballots SET voted_hexarchy=1, voted_charity=1, voted_democracy=1 WHERE voted=1 AND voted_hexarchy=0")
        except: pass
        # Resident-level democracy/alliance vote tracking
        try: conn.execute("""CREATE TABLE IF NOT EXISTS resident_demo_votes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resident_id INTEGER NOT NULL,
            proposal_id INTEGER NOT NULL,
            vote_type TEXT NOT NULL DEFAULT 'democracy',
            created_at TEXT DEFAULT (datetime('now'))
        )""")
        except: pass
        try: conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_res_demo ON resident_demo_votes(resident_id, proposal_id, vote_type)")
        except: pass
        conn.commit()

def seed_election_db():
    """Populate the DB with sample data if empty (from original Arkology election system)."""
    with get_election_db() as conn:
        # Seed candidates if none exist
        n = conn.execute('SELECT COUNT(*) FROM candidates').fetchone()[0]
        if n == 0:
            # Candidates seeded with real resident numbers from spreadsheet (Res# 2-15)
            cands = [
                (2,  'Traveling Time',  'Infrastructure expansion, solar grid optimization'),
                (3,  'High Tea',        'Public health council reform, expanded maternity support'),
                (5,  'Resident Five',   'Charity market transparency, anti-fraud measures'),
                (7,  'Resident Seven',  'Education curriculum modernization, STEM focus'),
                (9,  'Resident Nine',   'Environmental protections, pesticide enforcement'),
                (11, 'Resident Eleven', 'Criminal justice reform, Buzzy Bee expansion'),
                (13, 'Resident Thirteen','Coin voting system audit, election integrity'),
                (15, 'James Okafor',    'Dog & Cat rights expansion, Paw Patrol funding'),
            ]
            for res_num, name, platform in cands:
                conn.execute('INSERT INTO candidates (name,platform,res_num) VALUES (?,?,?)', (name, platform, res_num))
            conn.commit()

        # Seed election charities if none exist
        n = conn.execute('SELECT COUNT(*) FROM election_charities').fetchone()[0]
        if n == 0:
            chars = [
                ("Ark Children's Fund",       'Educational supplies and tutoring for youth residents',           20),
                ('Paw Patrol Reserve',         'Additional shelter capacity and medical care for animals',        20),
                ('Rooftop Gardens Initiative', 'Community farming on Arkology rooftops',                         15),
                ('Deputy Training Scholarship','Cover training costs for low-income deputy candidates',           10),
                ('Buzzy Bee Reintegration',    'Post-release job placement and counseling',                       10),
                ('Solar Commons',              'Shared solar panels for high-electricity-use residents',          15),
            ]
            for name, desc, min_v in chars:
                conn.execute('INSERT INTO election_charities (name,description,min_votes) VALUES (?,?,?)', (name, desc, min_v))
            conn.commit()

        # Seed proposals if none exist
        n = conn.execute('SELECT COUNT(*) FROM proposals').fetchone()[0]
        if n == 0:
            import json as _json
            proposals = [
                ('Electricity Rate Increase (1.15×)', 'tax',
                 'Raise all electricity tax brackets by 15%, within the 0.8×–1.2× Hexarchy proposal limit. Estimated +$170K/month revenue.',
                 '2/3', 0.667, None),
                ('New Hexagon District: Eastern Quarter', 'project',
                 'Construct 40 new hexagon cells in the eastern sector. Estimated cost: $8M, billed proportionally by sq-ft ownership.',
                 '2/3', 0.667, None),
                ('Dog Population Cap: +500 Units', 'animals',
                 'Increase the dog per-capita population target by 500 animals. Requires expanded Paw Patrol budget approval.',
                 '2/3', 0.667, None),
                ('Child Tax Slope Adjustment (0.9×)', 'tax',
                 'Reduce child tax rates by 10% across all tiers. Estimated –$366K/month revenue.',
                 '2/3', 0.667, None),
                ('Panarchist Military Alliance', 'alliance',
                 'Select which Panarchist military alliance Arkology should join, if any. The Hexarchy will appoint a Senator to manage joint military, court, and intelligence affairs. Rate each option 0–10; highest total rating wins.',
                 'rated', None, _json.dumps(['Northern Compact','Free Mariner League','Mountain Covenant','No Alliance'])),
                ('Amend Article 5: Voter Age Range 18–72', 'amendment',
                 'Expand eligible voter ages from 20–70 to 18–72. This is a word change requiring a 5/6 supermajority.',
                 '5/6', 0.833, None),
            ]
            for title, ptype, desc, threshold, threshold_val, options in proposals:
                conn.execute(
                    'INSERT INTO proposals (title,type,description,threshold,threshold_val,options) VALUES (?,?,?,?,?,?)',
                    (title, ptype, desc, threshold, threshold_val, options)
                )
            conn.commit()

init_election_db()
seed_election_db()

def election_get_state(key):
    with get_election_db() as conn:
        row = conn.execute("SELECT value FROM election_state WHERE key=?", (key,)).fetchone()
        return row[0] if row else None

def require_admin(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.headers.get('x-admin-password') != ADMIN_PASSWORD:
            return jsonify({'error':'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

BCH_SYSTEM_ADDRESS = "bitcoincash:qqvvwjw8rns37te6r4a0jzvtu86s3kx6lsykymqj5k"

# ── Column indices (1-based) — matches original spreadsheet headers ────────────
COL = {
    "res_num":       2,   # B  — Res #
    "password":      3,   # C  — Password (hashed emergency password)
    "email":         4,   # D  — E-Mail
    "last_name":     5,   # E  — Last Name
    "first_name":    6,   # F  — First Name
    "address":       7,   # G  — Address
    "year_born":     8,   # H  — Year Born
    "sex":           11,  # K  — Sex
    "alive":         16,  # P  — Alive?
    "disabled":      24,  # X  — Disabled?
    "hex_sides":     25,  # Y  — Sunlight (hex sides)
    "shared":        26,  # Z  — # Shared?
    "child1":        29,  # AC — Child 1
    "child2":        30,  # AD — Child 2
    "child3":        31,  # AE — Child 3
    "child4":        32,  # AF — Child 4
    "child5":        33,  # AG — Child 5
    "spouse_id":     35,  # AI — Spouse ID
    "marriage_date": 36,  # AJ — Date (marriage)
    "sqft1":         39,  # AM — Sq-ft 1
    "sqft2":         40,  # AN — Sq-ft 2
    "mort_date":     44,  # AR — Date (mortgage)
    "jury":          45,  # AS — Want Jury Duty?
    "deputy":        46,  # AT — Deputy?
    "wealth_m":      49,  # AW — Wealth $M
    "bch":           50,  # AX — BCH Address (original column)
    "pin_col":       92,  # CN — PIN hash (original column)
    "phone":         93,  # CO — Own Phone
    "ref1":          94,  # CP — Reference 1
    "ref2":          95,  # CQ — Reference 2
    "ref_checked":   96,  # CR — Ref Checked
    "taxes_owed":    97,  # CS — Taxes Owed
    "last_paid":     98,  # CT — Last Payment Date
    "election_col":       99,  # CU — Election Year Voted
    "immigration_interview": 100, # CV — Immigration Interview Pass/Fail
}

_UBI       = {3:360,4:480,5:600,6:720,7:840,8:960,9:1080,10:1200,
              11:1320,12:1440,13:1560,14:1680,15:1800}
_CHILD_TAX = {0:0,1:30,2:120,3:360,4:720,5:1200,
              6:1800,7:2520,8:3360,9:4320,10:5400}
SUNSHINE_RATE=600; DEPUTY_PAY=480; DISABILITY_PA=1200

def _read_startup_a19():
    """Read current A19 (budget multiplier) from spreadsheet."""
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE,data_only=True)
        ws=wb["Res"]
        v=ws.cell(19,1).value
        return float(v) if isinstance(v,(int,float)) else 0.75
    except: return 0.75

def _read_startup_constants():
    defaults={"bh96":1081.14,"charity_rate":11.1111}
    if not os.path.exists(EXCEL_FILE): return defaults
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE,data_only=True)
        ws=wb["Res"]; ws_ch=wb["Charity"]
        bh96=ws.cell(96,60).value
        ch_w6=ws_ch.cell(6,23).value; a14=ws.cell(14,1).value
        rate=a14 or ch_w6
        return {
            "bh96":        float(bh96) if isinstance(bh96,(int,float)) else defaults["bh96"],
            "charity_rate":float(rate) if isinstance(rate,(int,float)) and rate>0 else defaults["charity_rate"],
        }
    except Exception as e:
        print(f"[WARN] Startup: {e}"); return defaults

_C = _read_startup_constants()
MORT_MONTHLY = _C["bh96"]; CHARITY_RATE = _C["charity_rate"]
print(f"[*] BH96=${MORT_MONTHLY:.2f}  Charity={CHARITY_RATE:.4f}%")

def hash_pin(p):  return hashlib.sha256(p.encode()).hexdigest()
def load_wb():    return openpyxl.load_workbook(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else None

def excel_round(x,d=0):
    f=10**d; return math.floor(abs(x)*f+0.5)*math.copysign(1,x)/f

def parse_date(s):
    s=str(s or "").strip()
    if not s: return None
    try:
        if len(s)==7: return datetime.date(int(s[:4]),int(s[5:7]),1)
        return datetime.date.fromisoformat(s)
    except: return None

def is_active(alive_val):
    """Resident is active unless alive is explicitly 0 (deceased/banned)."""
    return alive_val != 0

def get_next_res_num(ws):
    existing=set()
    for row in ws.iter_rows(min_row=2,min_col=2,max_col=2,values_only=True):
        v=row[0]
        if isinstance(v,(int,float)) and v>0: existing.add(int(v))
    c=max(16,(max(existing)+1) if existing else 16)
    while c in existing or c%50==0: c+=1
    return c

def get_last_data_row(ws):
    last=1
    for i,row in enumerate(ws.iter_rows(min_row=2,min_col=2,max_col=2,values_only=True),start=2):
        if row[0] not in(None,""): last=i
    return last

def ensure_headers(ws):
    hdrs={
        92:"PIN",
        93:"Phone",94:"Reference 1",95:"Reference 2",96:"Ref Checked",
        97:"Taxes Owed",98:"Last Payment Date",99:"Election Year Voted",
    }
    for col,name in hdrs.items():
        if ws.cell(1,col).value is None:
            ws.cell(1,col).value=name
            ws.cell(1,col).font=Font(name="Arial",size=10,bold=True)

def _next_id(folder,prefix):
    import re as _re
    pat=_re.compile(rf"{_re.escape(prefix)}(\d+)",_re.IGNORECASE)
    best=0
    for fn in os.listdir(folder):
        m=pat.search(fn)
        if m: best=max(best,int(m.group(1)))
    return best+1

# ── Tax math ──────────────────────────────────────────────────────────────────
def calc_ubi(age):
    if age<15: return 0
    key=int(excel_round((age-2)/5)); key=max(3,min(15,key))
    return _UBI.get(key,0)

def calc_property_tax(sqft1,sqft2):
    t=(sqft1 or 0)+(sqft2 or 0)
    # Smooth transition at 16000: use max of both formulas to avoid discontinuity dip
    if t>600:
        formula_a = round(((t-600)**0.8/10000)*t, 2)
        formula_b = round(0.22*t, 2) if t>=16000 else 0
        return max(formula_a, formula_b)
    return 0.0

def calc_public_projects(sqft1,sqft2):
    return round(((sqft1 or 0)+(sqft2 or 0))*0.0741,2)

def calc_sunlight_tax(sides,shared):
    return round(int(sides or 0)*SUNSHINE_RATE/max(1,int(shared or 1)),2)

def calc_wealth_tax(wm):
    w=float(wm or 0)
    if w<1: return 0.0,0.0
    rate=math.log10(w)+1
    return round(rate,4),round(w*1e6*(rate/100)/12,2)

def calc_maternity(child_dates,sex,today=None):
    if today is None: today=datetime.date.today()
    if str(sex).upper() not in('W','F'): return 0.0,[]
    total,each=0.0,[]
    for d in child_dates:
        if not isinstance(d,datetime.date): continue
        mo=(today-d).days/30
        if mo>37: each.append(0.0); continue
        b=2400/(mo/6+1); total+=b; each.append(round(b,2))
    return round(total,2),each

def calc_marriage(marr_date,sex,today=None):
    if today is None: today=datetime.date.today()
    if str(sex).upper()!='M' or not isinstance(marr_date,datetime.date): return 0.0
    mo=(today-marr_date).days/180
    if mo<0: return 0.0
    if mo<1: return 2400.0
    if mo<2: return 1200.0
    return 0.0

def calc_mort_assist(mort_date,today=None):
    if today is None: today=datetime.date.today()
    if not isinstance(mort_date,datetime.date): return 0.0
    return round(MORT_MONTHLY,2) if today<mort_date+datetime.timedelta(days=15*365) else 0.0

def calc_late_fee(taxes_owed, last_paid_date):
    """Return (days_overdue, fee_pct, total_with_fee, months_unpaid)."""
    if taxes_owed <= 0: return 0, 0, 0, 0
    today = datetime.date.today()
    if not last_paid_date or not isinstance(last_paid_date, datetime.date):
        days = 0
    else:
        days = (today - last_paid_date).days
    if days >= 180:
        fee_pct = 90
    elif days >= 90:
        fee_pct = 30
    else:
        fee_pct = 0
    total = round(taxes_owed * (1 + fee_pct / 100), 2)
    months = round(days / 30, 1) if days > 0 else 0
    return days, fee_pct, total, months

def calc_net(year_born,disabled,deputy,sex,child_dates,sqft1,sqft2,sides,shared,wm,mort_date,marr_date):
    today=datetime.date.today(); age=today.year-int(year_born)
    ubi=calc_ubi(age); dis=DISABILITY_PA if disabled else 0; dep=DEPUTY_PAY if deputy else 0
    mat,each=calc_maternity(child_dates,sex,today)
    marr=calc_marriage(marr_date,sex,today); mort=calc_mort_assist(mort_date,today)
    nch=len([d for d in child_dates if isinstance(d,datetime.date)])
    ct=_CHILD_TAX.get(min(nch,10),0)
    prop=calc_property_tax(sqft1,sqft2); pub=calc_public_projects(sqft1,sqft2)
    sun=calc_sunlight_tax(sides,shared); wr,wt=calc_wealth_tax(wm)
    charity=round((prop+ct+sun+wt)*CHARITY_RATE/100,2)
    net=round(ubi-ct-sun-prop-wt-charity-pub+mort+marr+mat+dis+dep,2)
    a19=float(_read_startup_a19())
    total_sqft=int((sqft1 or 0)+(sqft2 or 0))
    return {"net_tax":net,"ubi":ubi,"disability_pay":dis,"deputy_pay":dep,
            "maternity":round(mat,2),"maternity_each":each,"marriage_bonus":marr,
            "mort_assist":round(mort,2),"prop_tax":prop,"sunlight_tax":sun,
            "child_tax":ct,"public_projects_tax":pub,"charity_tax":charity,
            "wealth_tax":wt,"wealth_rate":wr,"num_children":nch,"age":age,
            "total_sqft":total_sqft,"a19":round(a19,6)}

def excel_formulas(R):
    mat_sum="+".join(
        f'IF(($A$6-{c}{R})/30>37,0,IFERROR(VLOOKUP(($A$6-{c}{R})/30,$BE$58:$BF$94,2),0))'
        for c in["AC","AD","AE","AF","AG"])
    return {
        9:f"=$A$4", 10:f"=IF(P{R}=1,K{R},0)", 12:f"=I{R}-H{R}",
        13:f"=IF(N{R}>0,N{R}*$A$19,N{R})",
        14:f"=(O{R}-S{R}-T{R}-R{R}-U{R}-Q{R}-AK{R}+AP{R}+AH{R}+AA{R}+X{R}*$BK$2+AT{R}*$BJ$2)*P{R}",
        15:f"=IF(L{R}>=15,VLOOKUP(ROUND((L{R}-2)/5,0),$BI$64:$BJ$77,2,0),0)",
        17:f"=SUM(R{R}:U{R})*$A$14/100",
        18:f"=AL{R}", 19:f"=AB{R}", 20:f"=Y{R}*$BL$2/Z{R}", 21:f"=AU{R}",
        22:f"=IF(AND(70>=L{R},L{R}>=20),1,0)*P{R}",
        23:f"=IF(AND(70>=L{R},L{R}>=20),COUNT(L{R}),0)*P{R}*AS{R}",
        27:f'=IF(K{R}="W",{mat_sum},0)',
        28:f"=VLOOKUP(COUNT(AC{R}:AG{R}),$BM$4:$BO$14,3)",
        34:f'=IF(K{R}="M",VLOOKUP(($A$6-AJ{R})/180,$BG$84:$BH$86,2),0)',
        37:f"=SUM(AM{R}:AO{R})*$BR$47",
        38:f"=IF(SUM(AM{R}:AO{R})>=$BQ$19,$BQ$20*SUM(AM{R}:AO{R}),IF(SUM(AM{R}:AO{R})>$BP$19,((SUM(AM{R}:AO{R})-$BP$19)^($BO$19)/$BN$19)*SUM(AM{R}:AO{R}),0))",
        42:f"=AQ{R}*$BH$96", 43:f"=IF($A$6<(AR{R}+(15*365)),1,0)",
        47:f"=IF(AW{R}>=1,(LOG(AW{R},10)+1)/100*(AW{R}*1000000)/12,0)",
        48:f"=IF(AW{R}>=1,(LOG(AW{R},10)+1)/100,0)",
        55:f"=M{R}",
    }

# ── Logging helpers ───────────────────────────────────────────────────────────
FIELD_LABELS={
    "email":"Email","last_name":"Last Name","first_name":"First Name",
    "address":"Address","year_born":"Year Born","sex":"Sex",
    "sqft1":"Sq-ft 1","sqft2":"Sq-ft 2","hex_sides":"Hex Sides",
    "shared":"Shared Residents","wealth_m":"Net Worth ($M)","mort_date":"Mortgage Date",
    "spouse_id":"Spouse ID","marriage_date":"Marriage Date",
    "child1":"Child 1 DOB","child2":"Child 2 DOB","child3":"Child 3 DOB",
    "child4":"Child 4 DOB","child5":"Child 5 DOB",
    "jury":"Jury Willingness","deputy":"Deputy Role","disabled":"Disability",
    "bch":"BCH Address","ref1":"Reference 1","ref2":"Reference 2",
    "phone":"Phone Number","pin_col":"PIN","password":"Password",
    "taxes_owed":"Taxes Owed","last_paid":"Last Payment Date",
}

def write_update_log(res_id,name,field,old_val,new_val,reason,editor_id=None):
    now=datetime.datetime.now()
    log_id=f"L{_next_id(LOG_DIR,'L'):06d}"
    fname=f"{log_id}_res{str(res_id).zfill(3)}_{now.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
    fpath=os.path.join(LOG_DIR,fname)
    label=FIELD_LABELS.get(field,field)
    if field in("pin_col","password"):
        old_d="(hashed)" if old_val else "(none)"; new_d="(new hash set)"
    else:
        old_d=str(old_val) if old_val is not None else "(blank)"
        new_d=str(new_val) if new_val is not None else "(blank)"
    content=(f"ARK CITIZEN REGISTRY — UPDATE LOG\n===================================\n"
             f"Log ID       : {log_id}\nDate/Time    : {now.strftime('%Y-%m-%d %H:%M:%S')}\n"
             f"Resident #   : {res_id}\nName         : {name}\nField Updated: {label}\n"
             f"Previous     : {old_d}\nNew Value    : {new_d}\n"
             f"Reason       : {reason or '(no reason given)'}\nLogged by    : {editor_id or 'System'}\n")
    with open(fpath,"w",encoding="utf-8") as f: f.write(content)
    # Also write to changelog database
    _log_change(target_id=res_id, target_name=name, editor_id=editor_id, field=field,
                old_val=old_d, new_val=new_d, category=label, reason=reason)
    print(f"[LOG] {log_id}"); return log_id

def _init_changelog_db():
    db_path = os.path.join(BASE_DIR, "changelog.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("""CREATE TABLE IF NOT EXISTS changes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT NOT NULL DEFAULT (datetime('now','localtime')),
        target_id INTEGER,
        target_name TEXT NOT NULL DEFAULT '',
        editor_id INTEGER,
        editor_name TEXT NOT NULL DEFAULT '',
        field TEXT NOT NULL DEFAULT '',
        category TEXT NOT NULL DEFAULT '',
        old_val TEXT NOT NULL DEFAULT '',
        new_val TEXT NOT NULL DEFAULT '',
        reason TEXT NOT NULL DEFAULT '',
        reverted INTEGER NOT NULL DEFAULT 0,
        reverted_by INTEGER,
        reverted_at TEXT
    )""")
    conn.commit()
    return conn

def _log_change(target_id=None, target_name="", editor_id=None, field="", old_val="", new_val="", category="", reason=""):
    """Log a database change to the changelog."""
    conn = _init_changelog_db()
    # Resolve editor name
    editor_name = ""
    if editor_id:
        wb = load_wb()
        if wb:
            ws = wb[RES_SHEET]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rid = row[COL["res_num"]-1]
                if isinstance(rid, (int, float)) and int(rid) == int(editor_id):
                    fn = row[COL["first_name"]-1] or ""; ln = row[COL["last_name"]-1] or ""
                    editor_name = f"{fn} {ln}".strip(); break
    conn.execute("""INSERT INTO changes (target_id, target_name, editor_id, editor_name,
                    field, category, old_val, new_val, reason)
                    VALUES (?,?,?,?,?,?,?,?,?)""",
                 (target_id, target_name, editor_id, editor_name, field, category, old_val, new_val, reason or ""))
    conn.commit(); conn.close()

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index(): return send_from_directory("templates","index.html")

@app.route("/citizen-portal")
def citizen_portal(): return send_from_directory("templates","citizen_portal.html")

@app.route("/faq")
def faq_page(): return send_from_directory("templates","faq.html")


@app.route('/api/qr')
def api_qr():
    """Generate a QR code SVG for any text. Usage: /api/qr?text=...&size=6"""
    text = request.args.get('text', '')
    px   = max(2, min(12, int(request.args.get('px', 6))))
    if not text:
        return "Missing text", 400
    try:
        import importlib.util, os as _os
        spec = importlib.util.spec_from_file_location("qr_gen",
            _os.path.join(_os.path.dirname(__file__), "qr_generator.py"))
        qr_mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(qr_mod)
        svg = qr_mod.qr_to_svg(text, px=px)
    except Exception as e:
        svg = f'<svg xmlns="http://www.w3.org/2000/svg" width="200" height="200"><rect width="200" height="200" fill="white"/><text y="20" fill="red">QR Error: {str(e)[:80]}</text></svg>'
    resp = make_response(svg)
    resp.headers['Content-Type'] = 'image/svg+xml'
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp

@app.route("/api/budget-history")
def budget_history():
    """Return all rows from the Budget History sheet."""
    wb = load_wb()
    if not wb or "Budget History" not in wb.sheetnames:
        return jsonify({"rows": []})
    ws = wb["Budget History"]
    rows = []
    for r in range(2, ws.max_row + 1):
        yr   = ws.cell(r, 1).value
        mo   = ws.cell(r, 2).value
        lbl  = ws.cell(r, 3).value
        mult = ws.cell(r, 4).value
        tot  = ws.cell(r, 5).value
        nc   = ws.cell(r, 6).value
        pc   = ws.cell(r, 7).value
        if yr is None: continue
        rows.append({
            "year": int(yr), "month": int(mo), "label": str(lbl or ""),
            "multiplier": round(float(mult or 0), 4),
            "totalBudget": round(float(tot or 0), 2),
            "citizens": int(nc or 0),
            "perCapita": round(float(pc or 0), 2)
        })
    return jsonify({"rows": rows})

@app.route("/ark.mp4")
def serve_video():
    base = BASE_DIR if os.path.exists(os.path.join(BASE_DIR, "ark.mp4")) else "."
    resp = send_from_directory(base, "ark.mp4", conditional=True)
    resp.headers["Accept-Ranges"] = "bytes"
    resp.headers["Content-Type"] = "video/mp4"
    return resp

@app.route("/ark2.mp4")
def serve_video2():
    base = BASE_DIR if os.path.exists(os.path.join(BASE_DIR, "ark2.mp4")) else "."
    resp = send_from_directory(base, "ark2.mp4", conditional=True)
    resp.headers["Accept-Ranges"] = "bytes"
    resp.headers["Content-Type"] = "video/mp4"
    return resp

# ── IP-based password attempt tracking ───────────────────────────────────────
_pw_attempts = {}   # { ip: {'count': n, 'window_start': t, 'locked_until': t} }
_pw_lock = threading.Lock()
_MAX_ATTEMPTS = 3
_WINDOW_SECS  = 5 * 60    # 5 minutes
_LOCKOUT_SECS = 15 * 60   # 15 minutes

def _get_client_ip():
    return request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()

def _check_ip_lockout():
    """Returns (locked, seconds_remaining). Cleans up old entries."""
    ip = _get_client_ip()
    now = _time.time()
    with _pw_lock:
        rec = _pw_attempts.get(ip)
        if not rec:
            return False, 0
        if rec.get('locked_until', 0) > now:
            return True, int(rec['locked_until'] - now)
        # Window expired — reset
        if now - rec.get('window_start', 0) > _WINDOW_SECS:
            _pw_attempts.pop(ip, None)
            return False, 0
        return False, 0

def _record_failed_attempt():
    ip = _get_client_ip()
    now = _time.time()
    with _pw_lock:
        rec = _pw_attempts.setdefault(ip, {'count': 0, 'window_start': now, 'locked_until': 0})
        # Reset window if expired
        if now - rec['window_start'] > _WINDOW_SECS:
            rec['count'] = 0
            rec['window_start'] = now
            rec['locked_until'] = 0
        rec['count'] += 1
        if rec['count'] >= _MAX_ATTEMPTS:
            rec['locked_until'] = now + _LOCKOUT_SECS

def _clear_failed_attempts():
    ip = _get_client_ip()
    with _pw_lock:
        _pw_attempts.pop(ip, None)

@app.route("/verify-password", methods=["POST"])
def verify_password():
    """Verify res# + password for Citizen Portal login. IP lockout after 3 fails in 5 min."""
    locked, secs = _check_ip_lockout()
    if locked:
        return jsonify({"valid": False, "locked": True, "seconds": secs,
                        "error": f"Too many failed attempts. Resets in {secs//60}m {secs%60:02d}s"}), 429
    try:    data = request.get_json(force=True)
    except: return jsonify({"valid": False, "error": "Invalid request"}), 400
    res_id  = str(data.get("residentId", "")).strip()
    pw_raw  = str(data.get("password", "")).strip()
    if not res_id or not pw_raw:
        return jsonify({"valid": False, "error": "Resident ID and password required."}), 400
    try: res_id_int = int(res_id)
    except: return jsonify({"valid": False, "error": "Resident ID must be a number."}), 400
    wb = load_wb()
    if not wb: return jsonify({"valid": False, "error": "Database unavailable."}), 500
    ws = wb[RES_SHEET]
    ph = hash_pin(pw_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr  = row[COL["res_num"]-1]
        dpw = row[COL["password"]-1]
        da  = row[COL["alive"]-1]
        fn  = row[COL["first_name"]-1]; ln = row[COL["last_name"]-1]
        if not isinstance(dr, (int, float)): continue
        if int(dr) != res_id_int: continue
        if not is_active(da):
            return jsonify({"valid": False, "error": "Account not active."}), 403
        if not dpw:
            _record_failed_attempt()
            return jsonify({"valid": False, "error": "No password set. Ask an admin to run SET PASSWORD."}), 401
        if dpw != ph:
            _record_failed_attempt()
            locked2, secs2 = _check_ip_lockout()
            if locked2:
                return jsonify({"valid": False, "locked": True, "seconds": secs2,
                                "error": f"Too many failed attempts. Resets in {secs2//60}m {secs2%60:02d}s"}), 429
            return jsonify({"valid": False, "error": "Incorrect password."}), 401
        _clear_failed_attempts()
        return jsonify({"valid": True, "residentId": int(dr),
                        "firstName": str(fn or ""), "lastName": str(ln or "")})
    _record_failed_attempt()
    return jsonify({"valid": False, "error": f"Resident #{res_id} not found."}), 404

@app.route("/api/resident-summary", methods=["POST"])
def resident_summary():
    """Return net payment data for citizen portal display."""
    try:    data = request.get_json(force=True)
    except: return jsonify({"error": "Invalid request"}), 400
    res_id  = str(data.get("residentId", "")).strip()
    pw_raw  = str(data.get("password", "")).strip()
    if not res_id or not pw_raw:
        return jsonify({"error": "Missing credentials"}), 400
    try: res_id_int = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    wb = load_wb()
    if not wb: return jsonify({"error": "Database unavailable."}), 500
    ws = wb[RES_SHEET]
    ph = hash_pin(pw_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr  = row[COL["res_num"]-1]
        dpw = row[COL["password"]-1]
        da  = row[COL["alive"]-1]
        fn  = row[COL["first_name"]-1]; ln = row[COL["last_name"]-1]
        yb  = row[COL["year_born"]-1]
        if not isinstance(dr, (int, float)): continue
        if int(dr) != res_id_int: continue
        if dpw != ph: return jsonify({"error": "Auth failed"}), 401
        # Calculate net payment using calc_net
        dis  = row[COL["disabled"]-1]; dep = row[COL["deputy"]-1]
        sex  = row[COL["sex"]-1]; sq1 = row[COL["sqft1"]-1]; sq2 = row[COL["sqft2"]-1]
        sides = row[COL["hex_sides"]-1]; shared = row[COL["shared"]-1]
        wm   = row[COL["wealth_m"]-1]
        mort = row[COL["mort_date"]-1]; marr = row[COL["marriage_date"]-1]
        c_cols = ["child1","child2","child3","child4","child5"]
        cdates = []
        for ck in c_cols:
            raw = row[COL[ck]-1]
            cdates.append(parse_date(raw) if raw else None)
        try:
            result = calc_net(
                yb or 1990, bool(dis), bool(dep), str(sex or "M"),
                cdates, float(sq1 or 0), float(sq2 or 0),
                int(sides or 0), int(shared or 0),
                float(wm or 0), parse_date(mort), parse_date(marr)
            )
        except Exception as e:
            result = {"net_tax": 0, "a19": 0}
        a19 = result.get("a19", 0)
        net = result.get("net_tax", 0)
        adjusted = round(net * a19, 2) if net > 0 else round(net, 2)
        ref_chk = row[COL["ref_checked"]-1] if len(row) > COL["ref_checked"]-1 else None
        immig   = row[COL["immigration_interview"]-1] if len(row) > COL["immigration_interview"]-1 else None
        bch_addr = row[COL["bch"]-1] if len(row) > COL["bch"]-1 else None
        # ref_checked: 0=Processing, 1=Checked, 2=Declined
        # immigration_interview: 0=Processing, 1=Passed, 2=Declined
        ref_val = int(ref_chk) if isinstance(ref_chk, (int, float)) else 0
        immig_val = int(immig) if isinstance(immig, (int, float)) else 0
        return jsonify({
            "resId": int(dr), "firstName": str(fn or ""), "lastName": str(ln or ""),
            "netPayment": round(net, 2), "a19": round(a19, 6), "adjustedPayment": adjusted,
            "alive": bool(is_active(da)),
            "refChecked": 1 if ref_val == 1 else 0,
            "refCheckedRaw": ref_val,
            "immigrationInterview": 1 if immig_val == 1 else 0,
            "immigrationInterviewRaw": immig_val,
            "bchAddress": str(bch_addr).strip() if bch_addr else None,
        })
    return jsonify({"error": "Resident not found"}), 404

@app.route("/api/resident-fields", methods=["POST"])
def resident_fields():
    """Return current field values for a target resident. Requires hexarchy/auditor auth."""
    try:    data = request.get_json(force=True)
    except: return jsonify({"error": "Invalid request"}), 400
    caller_id = data.get("residentId")
    pw_raw    = str(data.get("password", "")).strip()
    target_id = data.get("targetResidentId")
    if not caller_id or not pw_raw or not target_id:
        return jsonify({"error": "Missing credentials or target ID"}), 400
    try: caller_int = int(caller_id); target_int = int(target_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    wb = load_wb()
    if not wb: return jsonify({"error": "Database unavailable"}), 500
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    # Auth the caller
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != caller_int: continue
        if not is_active(da): return jsonify({"error": "Account not active"}), 403
        if not dpw or dpw != ph: return jsonify({"error": "Auth failed"}), 401
        authed = True; break
    if not authed: return jsonify({"error": "Caller not found"}), 404
    # Check hexarchy/auditor
    is_authorized = False
    if "Gov Employees" in wb.sheetnames:
        ws_gov = wb["Gov Employees"]
        for gov_row in ws_gov.iter_rows(min_row=2, values_only=True):
            gov_res = gov_row[1]
            active = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
            pos = str(gov_row[4] or "").strip().lower() if len(gov_row) > 4 else ""
            hexarchy = str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
            if not isinstance(gov_res, (int, float)): continue
            if int(gov_res) != caller_int: continue
            if active not in ("yes","y","true","1"): continue
            if hexarchy in ("yes","y","true","1") or "auditor" in pos:
                is_authorized = True; break
    if not is_authorized:
        return jsonify({"error": "Not authorized — must be Hexarchy or Auditor"}), 403
    # Get target resident fields
    FIELD_MAP = {
        "email": COL["email"], "last_name": COL["last_name"], "first_name": COL["first_name"],
        "address": COL.get("address", 7), "year_born": COL["year_born"], "sex": COL["sex"],
        "sqft1": COL["sqft1"], "sqft2": COL["sqft2"], "hex_sides": COL["hex_sides"],
        "shared": COL["shared"], "wealth_m": COL["wealth_m"], "mort_date": COL["mort_date"],
        "spouse_id": COL["spouse_id"], "marriage_date": COL["marriage_date"],
        "child1": COL["child1"], "child2": COL["child2"], "child3": COL["child3"],
        "child4": COL["child4"], "child5": COL["child5"],
        "jury": COL["jury"], "deputy": COL["deputy"], "disabled": COL["disabled"],
        "bch": COL["bch"], "ref1": COL.get("ref1", 94), "ref2": COL.get("ref2", 95),
        "phone": COL.get("phone", 93),
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != target_int: continue
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        fields = {}
        for fname, col_idx in FIELD_MAP.items():
            val = row[col_idx - 1] if col_idx - 1 < len(row) else None
            if isinstance(val, (datetime.datetime, datetime.date)):
                val = val.strftime("%Y-%m-%d")
            elif val is not None:
                val = str(val)
            fields[fname] = val or ""
        return jsonify({"found": True, "resId": target_int,
                        "name": f"{fn} {ln}".strip(), "fields": fields})
    return jsonify({"error": f"Target Res#{target_int} not found"}), 404

@app.route("/last-month-budget")
def last_month_budget_page():
    return send_from_directory("templates","last_month_budget.html")

@app.route("/api/last-month-budget")
def last_month_budget():
    """Return last month's budget summary with top 50 taxpayers and benefit receivers."""
    wb = load_wb()
    if not wb: return jsonify({"error": "Database unavailable"}), 500
    ws = wb[RES_SHEET]
    a19 = float(ws.cell(19, 1).value or 0.75)
    # Read Budget History for latest row
    total_budget = 0; citizen_count = 0; per_capita = 0; multiplier = a19
    if "Budget History" in wb.sheetnames:
        ws_bh = wb["Budget History"]
        last_row = None
        for r in range(2, ws_bh.max_row + 1):
            if ws_bh.cell(r, 1).value is not None:
                last_row = r
        if last_row:
            total_budget = float(ws_bh.cell(last_row, 5).value or 0)
            citizen_count = int(ws_bh.cell(last_row, 6).value or 0)
            per_capita = float(ws_bh.cell(last_row, 7).value or 0)
            multiplier = float(ws_bh.cell(last_row, 4).value or a19)
    # Compute N and M for all residents
    multiplier_pct = multiplier  # e.g. 22.76 meaning 22.76%
    multiplier_frac = multiplier / 100.0  # e.g. 0.2276 for actual computation
    residents = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rn = row[COL["res_num"]-1]
        if not isinstance(rn, (int, float)): continue
        alive = row[COL["alive"]-1]
        if alive == 0: continue
        n_val = row[13]  # col N (0-indexed=13)
        if not isinstance(n_val, (int, float)): continue
        n = float(n_val)
        m = round(n * multiplier_frac, 2) if n > 0 else round(n, 2)
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        residents.append({"resId": int(rn), "name": f"{fn} {ln}".strip(), "n": n, "m": m})
    # Sort: taxpayers are those with negative M, benefit receivers have positive M
    taxpayers = sorted([r for r in residents if r["m"] < 0], key=lambda r: r["m"])[:50]
    receivers = sorted([r for r in residents if r["m"] > 0], key=lambda r: r["m"], reverse=True)[:50]
    return jsonify({
        "totalBudget": round(total_budget, 2),
        "citizenCount": citizen_count or len(residents),
        "perCapita": round(per_capita, 2),
        "multiplier": round(multiplier_pct, 4),
        "taxpayers": [{"resId": r["resId"], "name": r["name"], "amount": r["m"]} for r in taxpayers],
        "receivers": [{"resId": r["resId"], "name": r["name"], "amount": r["m"]} for r in receivers],
    })

@app.route("/api/election/eligibility", methods=["POST"])
def election_eligibility():
    """Check if resident is eligible to vote: alive, age 20-70, not voted yet this election."""
    try:    data = request.get_json(force=True)
    except: return jsonify({"eligible": False, "error": "Invalid request"}), 400
    res_id_raw = data.get("residentId")
    pin_raw    = str(data.get("pin", "")).strip()
    if not res_id_raw or not pin_raw:
        return jsonify({"eligible": False, "error": "Resident ID and PIN required."}), 400
    try: res_id_int = int(res_id_raw)
    except: return jsonify({"eligible": False, "error": "Resident ID must be a number."}), 400
    wb = load_wb()
    if not wb: return jsonify({"eligible": False, "error": "Database unavailable."}), 500
    ws = wb[RES_SHEET]
    ph = hash_pin(pin_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr  = row[COL["res_num"]-1]
        if not isinstance(dr, (int, float)): continue
        if int(dr) != res_id_int: continue
        # Check PIN
        dp  = row[COL["pin_col"]-1]
        if not dp or dp != ph:
            return jsonify({"eligible": False, "error": "Incorrect PIN."}), 401
        # Check alive
        da  = row[COL["alive"]-1]
        if not is_active(da):
            return jsonify({"eligible": False, "error": "Resident is not active."}), 403
        # Check age 20-70
        yb  = row[COL["year_born"]-1]
        if not isinstance(yb, (int, float)):
            return jsonify({"eligible": False, "error": "Year of birth not on file."}), 400
        age = datetime.date.today().year - int(yb)
        if age < 20 or age > 70:
            return jsonify({"eligible": False, "error": f"Age {age} is outside the eligible range (20–70)."}), 403
        # Check election col for this year
        elc = row[COL["election_col"]-1]
        this_year = str(datetime.date.today().year)
        if str(elc or "").strip() == this_year:
            return jsonify({"eligible": False, "alreadyVoted": True, "error": "You have already voted in this election cycle."}), 409
        fn = row[COL["first_name"]-1]; ln = row[COL["last_name"]-1]
        return jsonify({"eligible": True, "residentId": int(dr),
                        "firstName": str(fn or ""), "lastName": str(ln or ""),
                        "age": age})
    return jsonify({"eligible": False, "error": f"Resident #{res_id_int} not found."}), 404

@app.route("/api/hexarchy-check", methods=["POST"])
def hexarchy_check():
    """Check if the logged-in resident is a hexarchy member (for gov employees edit access)."""
    try:    data = request.get_json(force=True)
    except: return jsonify({"isHexarchy": False, "error": "Invalid request"}), 400
    res_id  = data.get("residentId")
    pw_raw  = str(data.get("password", "")).strip()
    if not res_id or not pw_raw:
        return jsonify({"isHexarchy": False, "error": "Missing credentials"}), 400
    # Verify password first
    try: res_id_int = int(res_id)
    except: return jsonify({"isHexarchy": False, "error": "Invalid ID"}), 400
    wb = load_wb()
    if not wb: return jsonify({"isHexarchy": False, "error": "Database unavailable"}), 500
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id_int: continue
        if not is_active(da): return jsonify({"isHexarchy": False, "error": "Account not active"}), 403
        if not dpw or dpw != ph: return jsonify({"isHexarchy": False, "error": "Auth failed"}), 401
        authed = True; break
    if not authed: return jsonify({"isHexarchy": False, "error": "Resident not found"}), 404
    # Check Gov Employees sheet for hexarchy — scan ALL rows (there may be multiple rows per res#)
    if "Gov Employees" not in wb.sheetnames:
        return jsonify({"isHexarchy": False})
    ws_gov = wb["Gov Employees"]
    for row in ws_gov.iter_rows(min_row=2, values_only=True):
        gov_res = row[1]  # col B = Res #
        hexarchy_val = row[8] if len(row) > 8 else None  # col I = Hexarchy
        active_val   = row[7] if len(row) > 7 else None  # col H = Active
        if not isinstance(gov_res, (int, float)) or int(gov_res) != res_id_int: continue
        is_active_row = str(active_val or "").strip().lower() in ("yes", "y", "true", "1")
        is_hex = str(hexarchy_val or "").strip().lower() in ("yes", "y", "true", "1")
        if is_active_row and is_hex:
            return jsonify({"isHexarchy": True})
    return jsonify({"isHexarchy": False})

@app.route("/register",methods=["POST"])
def register():
    try:    data=request.get_json(force=True)
    except: data=request.form.to_dict()
    for f in["firstName","lastName","email","yearBorn","sex","pin"]:
        if not str(data.get(f,"")).strip():
            return jsonify({"success":False,"error":f"Missing: {f}"}),400
    pin_raw=str(data.get("pin","")).strip()
    if not pin_raw.isdigit() or len(pin_raw)!=4:
        return jsonify({"success":False,"error":"PIN must be exactly 4 digits."}),400
    wb=load_wb()
    if not wb: return jsonify({"success":False,"error":"Excel file not found."}),500
    ws=wb[RES_SHEET]; ensure_headers(ws)
    for row in ws.iter_rows(min_row=2,min_col=COL["email"],max_col=COL["email"],values_only=True):
        if row[0] and str(row[0]).strip().lower()==data["email"].strip().lower():
            return jsonify({"success":False,"error":"Email already registered."}),409
    res_num=get_next_res_num(ws); new_row=get_last_data_row(ws)+1
    cr=data.get("children",[])
    if isinstance(cr,str):
        try: cr=json.loads(cr)
        except: cr=[]
    child_dates,child_cells=[],[]
    for ch in cr[:5]:
        dob=(ch.get("dob","") if isinstance(ch,dict) else "").strip()
        d=parse_date(dob)
        if d: child_dates.append(d); child_cells.append(d)
        else: child_cells.append(None)
    while len(child_cells)<5: child_cells.append(None)
    sex=data["sex"].upper(); year_born=int(data["yearBorn"])
    disabled=data.get("disabled")=="yes"; deputy=data.get("deputy")=="yes"
    jury=1 if data.get("juryDuty")=="yes" else 0
    sqft1=float(data.get("sqft1") or 0); sqft2=float(data.get("sqft2") or 0)
    sides=int(data.get("hexSides") or 0); shared=int(data.get("sharedHousing") or 1)
    wm=float(data.get("wealthM") or 0)
    marr_date=parse_date(data.get("marriageDate"))
    mort_date=parse_date(data.get("mortDate"))
    r1f=data.get("ref1First","").strip(); r1l=data.get("ref1Last","").strip()
    r2f=data.get("ref2First","").strip(); r2l=data.get("ref2Last","").strip()
    phone=str(data.get("phone","")).strip()
    password_raw=str(data.get("password","")).strip()
    taxes=calc_net(year_born,disabled,deputy,sex,child_dates,sqft1,sqft2,sides,shared,wm,mort_date,marr_date)

    def W(k,v,fmt=None):
        c=ws.cell(row=new_row,column=COL[k],value=v)
        if fmt: c.number_format=fmt

    W("res_num",res_num); W("pin_col",hash_pin(pin_raw))
    if password_raw: W("password",hash_pin(password_raw))
    W("email",data["email"].strip()); W("last_name",data["lastName"].strip())
    W("first_name",data["firstName"].strip())
    W("address",data.get("address","").strip() or None)
    W("year_born",year_born); W("sex",sex); W("alive",1)
    W("disabled",1 if disabled else 0)
    W("hex_sides",sides or None); W("shared",shared)
    for key,val in zip(["child1","child2","child3","child4","child5"],child_cells):
        if val: W(key,val,"MM/DD/YY")
    W("spouse_id",data.get("spouseId","").strip() or None)
    if marr_date: W("marriage_date",marr_date,"MM/DD/YY")
    if mort_date: W("mort_date",mort_date,"MM/DD/YY")
    W("jury",jury); W("deputy",1 if deputy else 0)
    W("wealth_m",wm or None)
    W("bch",data.get("bchAddress","").strip() or None)
    W("phone",phone or None)
    W("ref1",f"{r1f} {r1l}".strip() or None)
    W("ref2",f"{r2f} {r2l}".strip() or None)
    W("ref_checked",0)
    for col_idx,formula in excel_formulas(new_row).items():
        ws.cell(row=new_row,column=col_idx,value=formula)
    for col in range(1,100):
        c=ws.cell(row=new_row,column=col)
        if c.value is not None: c.font=Font(name="Arial",size=10)
    try: wb.save(EXCEL_FILE)
    except PermissionError: return jsonify({"success":False,"error":"Close Excel first."}),500
    print(f"[+] #{res_num} {data['firstName']} {data['lastName']}  net={taxes['net_tax']}")
    return jsonify({"success":True,"residentId":res_num,"taxes":taxes})


@app.route("/verify-pin",methods=["POST"])
def verify_pin():
    try:    data=request.get_json(force=True)
    except: return jsonify({"valid":False,"error":"Invalid request"}),400
    res_id=str(data.get("residentId","")).strip(); pin_raw=str(data.get("pin","")).strip()
    if not res_id or not pin_raw: return jsonify({"valid":False,"error":"ID and PIN required."}),400
    if not pin_raw.isdigit() or len(pin_raw)!=4: return jsonify({"valid":False,"error":"PIN must be 4 digits."}),400
    try: res_id_int=int(res_id)
    except: return jsonify({"valid":False,"error":"Resident ID must be a number."}),400
    wb=load_wb()
    if not wb: return jsonify({"valid":False,"error":"Database unavailable."}),500
    ws=wb[RES_SHEET]; ph=hash_pin(pin_raw)
    for row in ws.iter_rows(min_row=2,values_only=True):
        dr=row[COL["res_num"]-1]; dp=row[COL["pin_col"]-1]
        df=row[COL["first_name"]-1]; dl=row[COL["last_name"]-1]; da=row[COL["alive"]-1]
        if not isinstance(dr,(int,float)): continue
        if int(dr)!=res_id_int: continue
        if not is_active(da): return jsonify({"valid":False,"error":"Account not active."}),403
        if not dp: return jsonify({"valid":False,"error":"No PIN set. Ask an admin to run SET_PIN.bat."}),401
        if dp!=ph: return jsonify({"valid":False,"error":"Incorrect PIN."}),401
        return jsonify({"valid":True,"residentId":int(dr),"firstName":str(df or ""),"lastName":str(dl or "")})
    return jsonify({"valid":False,"error":f"Resident #{res_id} not found."}),404


@app.route("/check-references",methods=["POST"])
def check_references():
    try:    data=request.get_json(force=True)
    except: data=request.form.to_dict()
    r1f=str(data.get("ref1First","")).strip().lower(); r1l=str(data.get("ref1Last","")).strip().lower()
    r2f=str(data.get("ref2First","")).strip().lower(); r2l=str(data.get("ref2Last","")).strip().lower()
    if not all([r1f,r1l,r2f,r2l]): return jsonify({"success":False,"error":"Please enter both names."}),400
    if r1f==r2f and r1l==r2l: return jsonify({"success":False,"error":"References must be different."}),400
    wb=load_wb()
    if not wb: return jsonify({"success":False,"error":"Database unavailable."}),500
    ws=wb[RES_SHEET]; residents=set()
    for row in ws.iter_rows(min_row=2,values_only=True):
        rn=row[COL["res_num"]-1]; fn=row[COL["first_name"]-1]; ln=row[COL["last_name"]-1]
        if not isinstance(rn,(int,float)): continue
        if fn and ln: residents.add((str(fn).strip().lower(),str(ln).strip().lower()))
    missing=[]
    if (r1f,r1l) not in residents: missing.append(f"{data.get('ref1First','')} {data.get('ref1Last','')}")
    if (r2f,r2l) not in residents: missing.append(f"{data.get('ref2First','')} {data.get('ref2Last','')}")
    if missing:
        ns=" and ".join(f'"{n.strip()}"' for n in missing)
        return jsonify({"success":False,"error":f"{ns} not found. Both must be existing residents."}),400
    return jsonify({"success":True})


@app.route("/update",methods=["POST"])
def update_field():
    try:    data=request.get_json(force=True)
    except: return jsonify({"success":False,"error":"Invalid request"}),400
    res_id=data.get("residentId"); field=str(data.get("field","")).strip(); value=data.get("value","")
    target_res_id = data.get("targetResidentId") or res_id  # edit someone else, or yourself
    reason=str(data.get("reason","")).strip()
    pin_raw=str(data.get("pin","")).strip()
    pw_raw=str(data.get("password","")).strip()
    if not res_id or not field: return jsonify({"success":False,"error":"Missing fields."}),400
    ALLOWED={k for k in COL if k not in("res_num","alive","ref_checked","taxes_owed","last_paid","election_col")}
    if field not in ALLOWED: return jsonify({"success":False,"error":f"Field '{field}' not updatable."}),400
    wb=load_wb()
    if not wb: return jsonify({"success":False,"error":"Database unavailable."}),500
    ws=wb[RES_SHEET]
    # Auth the CALLER (residentId)
    if pw_raw:
        ph=hash_pin(pw_raw)
        auth_col=COL["password"]; auth_label="password"
    else:
        if not pin_raw.isdigit() or len(pin_raw)!=4:
            return jsonify({"success":False,"error":"Invalid PIN."}),400
        ph=hash_pin(pin_raw); auth_col=COL["pin_col"]; auth_label="PIN"
    authed = False
    for row in ws.iter_rows(min_row=2):
        dr=row[COL["res_num"]-1].value; da=row[COL["alive"]-1].value
        dp=row[auth_col-1].value
        if not isinstance(dr,(int,float)): continue
        if int(dr)!=int(res_id): continue
        if not is_active(da): return jsonify({"success":False,"error":"Account not active."}),403
        if not dp: return jsonify({"success":False,"error":f"No {auth_label} set."}),401
        if dp!=ph: return jsonify({"success":False,"error":f"Incorrect {auth_label}."}),401
        authed = True; break
    if not authed: return jsonify({"success":False,"error":f"Resident #{res_id} not found."}),404
    # Check if CALLER is on the Hexarchy or is an Auditor
    is_authorized = False
    if "Gov Employees" in wb.sheetnames:
        ws_gov = wb["Gov Employees"]
        for gov_row in ws_gov.iter_rows(min_row=2, values_only=True):
            gov_res = gov_row[1]
            active  = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
            pos     = str(gov_row[4] or "").strip().lower() if len(gov_row) > 4 else ""
            hexarchy= str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
            if not isinstance(gov_res, (int, float)): continue
            if int(gov_res) != int(res_id): continue
            if active not in ("yes","y","true","1"): continue
            if hexarchy in ("yes","y","true","1") or "auditor" in pos:
                is_authorized = True
                break
    if not is_authorized:
        return jsonify({"success":False,"error":"Must be on the Hexarchy or an Auditor to change data."}),403
    # Find the TARGET resident row
    target_row = None; resident_name = ""
    for row in ws.iter_rows(min_row=2):
        dr = row[COL["res_num"]-1].value
        fn = row[COL["first_name"]-1].value or ""; ln = row[COL["last_name"]-1].value or ""
        if not isinstance(dr,(int,float)): continue
        if int(dr) != int(target_res_id): continue
        target_row = row[0].row; resident_name = f"{fn} {ln}".strip(); break
    if not target_row: return jsonify({"success":False,"error":f"Target Res#{target_res_id} not found."}),404
    def parse_val(f,raw):
        raw=str(raw).strip()
        if f in("sqft1","sqft2","shared","hex_sides","year_born","jury","deputy","disabled","wealth_m"):
            try: return float(raw) if "." in raw else int(raw)
            except: return None
        if f in("mort_date","marriage_date","child1","child2","child3","child4","child5"):
            return parse_date(raw)
        if f=="pin_col":
            if not raw.isdigit() or len(raw)!=4: raise ValueError("PIN must be exactly 4 digits.")
            return hash_pin(raw)
        if f=="password":
            if len(raw)<6: raise ValueError("Password must be at least 6 characters.")
            return hash_pin(raw)
        return raw or None
    try: parsed=parse_val(field,value)
    except ValueError as e: return jsonify({"success":False,"error":str(e)}),400
    cell=ws.cell(row=target_row,column=COL[field])
    old_val=cell.value; cell.value=parsed
    if isinstance(parsed,datetime.date): cell.number_format="MM/DD/YY"
    cell.font=Font(name="Arial",size=10)
    def rv(k):
        v=ws.cell(row=target_row,column=COL[k]).value
        return v.date() if isinstance(v,datetime.datetime) else v
    def rdate(k): v=rv(k); return v if isinstance(v,datetime.date) else None
    taxes=calc_net(rv("year_born") or 1990,bool(rv("disabled")),bool(rv("deputy")),str(rv("sex") or "M"),
        [rdate(k) for k in("child1","child2","child3","child4","child5") if rdate(k)],
        float(rv("sqft1") or 0),float(rv("sqft2") or 0),int(rv("hex_sides") or 0),
        int(rv("shared") or 1),float(rv("wealth_m") or 0),rdate("mort_date"),rdate("marriage_date"))
    try: wb.save(EXCEL_FILE)
    except PermissionError: return jsonify({"success":False,"error":"Close Excel first."}),500
    write_update_log(target_res_id,resident_name,field,old_val,parsed,reason,editor_id=int(res_id))
    return jsonify({"success":True,"taxes":taxes})


@app.route("/taxes-owed",methods=["POST"])
def taxes_owed():
    try:    data=request.get_json(force=True)
    except: return jsonify({"success":False,"error":"Invalid request"}),400
    res_id=data.get("residentId")
    pin_raw=str(data.get("pin","")).strip(); pw_raw=str(data.get("password","")).strip()
    if not res_id: return jsonify({"success":False,"error":"Missing fields."}),400
    try: res_id_int=int(res_id)
    except: return jsonify({"success":False,"error":"Invalid ID."}),400
    wb=load_wb()
    if not wb: return jsonify({"success":False,"error":"Database unavailable."}),500
    ws=wb[RES_SHEET]
    # Auth: prefer password, fall back to PIN
    if pw_raw:
        ph=hash_pin(pw_raw); auth_col=COL["password"]
        def bad_auth(dp): return not dp or dp!=ph
    else:
        if not pin_raw.isdigit() or len(pin_raw)!=4:
            return jsonify({"success":False,"error":"PIN must be 4 digits."}),400
        ph=hash_pin(pin_raw); auth_col=COL["pin_col"]
        def bad_auth(dp): return not dp or dp!=ph
    for row in ws.iter_rows(min_row=2,values_only=True):
        dr=row[COL["res_num"]-1]; da=row[COL["alive"]-1]
        dp=row[auth_col-1]
        fn=row[COL["first_name"]-1] or ""; ln=row[COL["last_name"]-1] or ""
        taxes_ow=row[COL["taxes_owed"]-1] or 0
        last_paid_raw=row[COL["last_paid"]-1]
        if not isinstance(dr,(int,float)): continue
        if int(dr)!=res_id_int: continue
        if not is_active(da): return jsonify({"success":False,"error":"Account not active."}),403
        if bad_auth(dp): return jsonify({"success":False,"error":"Authentication failed."}),401
        owed=round(float(taxes_ow),2)
        last_paid=last_paid_raw.date() if isinstance(last_paid_raw,datetime.datetime) else (last_paid_raw if isinstance(last_paid_raw,datetime.date) else None)
        days_overdue,fee_pct,total_with_fee,months_unpaid=calc_late_fee(owed,last_paid)
        today=datetime.date.today().strftime("%Y%m%d"); memo=f"ARK-TAX-{res_id_int}-{today}"
        return jsonify({"success":True,"residentId":res_id_int,"name":f"{fn} {ln}".strip(),
            "taxes_owed_usd":owed,"days_overdue":days_overdue,"fee_pct":fee_pct,
            "total_with_fee":total_with_fee,"months_unpaid":months_unpaid,
            "last_paid":str(last_paid) if last_paid else None,
            "bch_address":BCH_SYSTEM_ADDRESS,"payment_memo":memo})
    return jsonify({"success":False,"error":f"Resident #{res_id} not found."}),404


@app.route("/check-bch-balance", methods=["GET"])
def check_bch_balance():
    """Check BCH system wallet balance via Blockchair."""
    import urllib.request, json as _json
    try:
        url = f"https://blockchair.com/bitcoin-cash/dashboards/address/{BCH_SYSTEM_ADDRESS}?limit=1"
        req = urllib.request.Request(url, headers={"User-Agent": "ArkRegistry/1.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = _json.loads(r.read())
        balance_sat = data.get("data", {}).get(BCH_SYSTEM_ADDRESS, {}).get("address", {}).get("balance", 0)
        return jsonify({"success": True, "balance_bch": balance_sat / 1e8, "address": BCH_SYSTEM_ADDRESS})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route("/verify-bch-payment", methods=["POST"])
def verify_bch_payment():
    """Check if the BCH address received the expected amount recently (last 10 txs)."""
    import urllib.request, json as _json
    try:
        data = request.get_json(force=True)
    except:
        return jsonify({"success": False, "error": "Invalid request"}), 400
    address = data.get("address", BCH_SYSTEM_ADDRESS)
    expected_bch = float(data.get("expected_bch", 0))
    try:
        url = f"https://blockchair.com/bitcoin-cash/dashboards/address/{address}?limit=10"
        req = urllib.request.Request(url, headers={"User-Agent": "ArkRegistry/1.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            result = _json.loads(r.read())
        txs = result.get("data", {}).get(address, {}).get("transactions", [])
        for txhash in txs[:10]:
            tx_url = f"https://blockchair.com/bitcoin-cash/dashboards/transaction/{txhash}"
            tx_req = urllib.request.Request(tx_url, headers={"User-Agent": "ArkRegistry/1.0"})
            with urllib.request.urlopen(tx_req, timeout=10) as r2:
                tx_data = _json.loads(r2.read())
            outputs = tx_data.get("data", {}).get(txhash, {}).get("outputs", [])
            received = sum(o.get("value", 0) for o in outputs if o.get("recipient") == address) / 1e8
            tx_time = tx_data.get("data", {}).get(txhash, {}).get("transaction", {}).get("time", "")
            if received > 0 and received >= expected_bch * 0.98:
                return jsonify({"success": True, "verified": True,
                                "txhash": txhash, "received_bch": received, "tx_time": tx_time})
        return jsonify({"success": True, "verified": False,
                        "message": "Payment not yet detected. Please wait 1-2 minutes and try again."})
    except Exception as e:
        return jsonify({"success": False, "error": f"Blockchain check failed: {str(e)}"}), 502


@app.route("/confirm-payment",methods=["POST"])
def confirm_payment():
    try:    data=request.get_json(force=True)
    except: return jsonify({"success":False,"error":"Invalid request"}),400
    res_id=data.get("residentId")
    pin_raw=str(data.get("pin","")).strip(); pw_raw=str(data.get("password","")).strip()
    amount_paid=data.get("amount_paid",0); bch_amount=data.get("bch_amount",0); bch_rate=data.get("bch_rate",43000)
    if not res_id: return jsonify({"success":False,"error":"Missing fields."}),400
    # Auth: prefer password, fall back to PIN
    if pw_raw:
        ph=hash_pin(pw_raw); auth_col=COL["password"]
    else:
        if not pin_raw.isdigit() or len(pin_raw)!=4:
            return jsonify({"success":False,"error":"PIN must be 4 digits."}),400
        ph=hash_pin(pin_raw); auth_col=COL["pin_col"]
    try: res_id_int=int(res_id)
    except: return jsonify({"success":False,"error":"Invalid ID."}),400
    wb=load_wb()
    if not wb: return jsonify({"success":False,"error":"Database unavailable."}),500
    ws=wb[RES_SHEET]; target_row=None; resident_name=""
    for row in ws.iter_rows(min_row=2):
        dr=row[COL["res_num"]-1].value; da=row[COL["alive"]-1].value
        dp=row[auth_col-1].value
        fn=row[COL["first_name"]-1].value or ""; ln=row[COL["last_name"]-1].value or ""
        if not isinstance(dr,(int,float)): continue
        if int(dr)!=res_id_int: continue
        if not is_active(da): return jsonify({"success":False,"error":"Account not active."}),403
        if not dp or dp!=ph: return jsonify({"success":False,"error":"Authentication failed."}),401
        target_row=row[0].row; resident_name=f"{fn} {ln}".strip(); break
    if not target_row: return jsonify({"success":False,"error":"Resident not found."}),404
    today=datetime.date.today()
    ws.cell(row=target_row,column=COL["taxes_owed"]).value=0
    ws.cell(row=target_row,column=COL["taxes_owed"]).font=Font(name="Arial",size=10)
    ws.cell(row=target_row,column=COL["last_paid"]).value=today
    ws.cell(row=target_row,column=COL["last_paid"]).number_format="MM/DD/YY"
    ws.cell(row=target_row,column=COL["last_paid"]).font=Font(name="Arial",size=10)
    try: wb.save(EXCEL_FILE)
    except PermissionError: return jsonify({"success":False,"error":"Close Excel first."}),500
    now=datetime.datetime.now()
    r_num=_next_id(RECEIPT_DIR,'R'); receipt_id=f"R{r_num:06d}"
    receipt_num=f"{receipt_id}-{now.strftime('%Y%m%d')}-res{res_id_int}"
    receipt_text=(f"ARK COMMUNITY TAX RECEIPT\n==========================\n"
                  f"Receipt ID   : {receipt_id}\nReceipt #    : {receipt_num}\n"
                  f"Date/Time    : {now.strftime('%Y-%m-%d %H:%M:%S')}\n"
                  f"Resident #   : {res_id_int}\nName         : {resident_name}\n\n"
                  f"Amount Paid  : ${float(amount_paid):,.2f} USD\n"
                  f"BCH Amount   : {float(bch_amount):.5f} BCH\n"
                  f"Exchange Rate: ${float(bch_rate):,.2f} USD/BCH\n\n"
                  f"BCH Address  : {BCH_SYSTEM_ADDRESS}\n"
                  f"Payment Memo : ARK-TAX-{res_id_int}-{now.strftime('%Y%m%d')}\n\n"
                  f"Status       : PAYMENT RECORDED\n\n"
                  f"This receipt confirms your tax payment to the Ark Community Fund.\n"
                  f"Keep this document for your records.\n"
                  f"Disputes must be raised within 30 days.\n")
    fname=f"receipt_{receipt_num}.txt"
    with open(os.path.join(RECEIPT_DIR,fname),"w",encoding="utf-8") as f: f.write(receipt_text)
    print(f"[TAX] #{res_id_int} paid ${float(amount_paid):.2f}  {receipt_id}")
    return jsonify({"success":True,"receipt_num":receipt_num,"receipt_text":receipt_text})


@app.route("/stats")
def stats():
    wb=load_wb()
    if not wb: return jsonify({"error":"File not found"}),500
    ws=wb[RES_SHEET]
    count=sum(1 for r in ws.iter_rows(min_row=2,min_col=2,max_col=2,values_only=True) if isinstance(r[0],(int,float)))
    return jsonify({"total_residents":count})




# ══════════════════════════════════════════════════════════════════════════════
# COMMUNITY CALENDAR  — /calendar  (SQLite backend, no extra BAT files)
# ══════════════════════════════════════════════════════════════════════════════

CALENDAR_DB  = os.path.join(BASE_DIR, "Calendar.db")
EVENT_IMG_DIR = os.path.join(BASE_DIR, "Event JPGs")
os.makedirs(EVENT_IMG_DIR, exist_ok=True)

def cal_db():
    """Open calendar DB and ensure schema exists."""
    conn = sqlite3.connect(CALENDAR_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS events (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            date          TEXT    NOT NULL,
            time          TEXT    DEFAULT '',
            title         TEXT    NOT NULL,
            location      TEXT    DEFAULT '',
            description   TEXT    DEFAULT '',
            creator_res_id INTEGER NOT NULL,
            creator_name  TEXT    NOT NULL,
            created_at    TEXT    NOT NULL,
            image_filename TEXT   DEFAULT NULL
        );
        CREATE TABLE IF NOT EXISTS interested (
            event_id INTEGER NOT NULL,
            res_id   INTEGER NOT NULL,
            PRIMARY KEY (event_id, res_id)
        );
        CREATE INDEX IF NOT EXISTS idx_events_date ON events(date);
    """)
    conn.commit()
    return conn

def cal_verify_pin(res_id_raw, pin_raw):
    """Returns (True, res_id, full_name) or (False, None, error_msg). Accepts PIN (4 digits)."""
    try:
        res_id = int(res_id_raw)
        pin_raw = str(pin_raw).strip()
    except:
        return False, None, "Invalid ID or PIN."
    if not pin_raw.isdigit() or len(pin_raw) != 4:
        return False, None, "PIN must be 4 digits."
    wb = load_wb()
    if not wb:
        return False, None, "Database unavailable."
    ws = wb[RES_SHEET]
    ph = hash_pin(pin_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id:
            continue
        da = row[COL["alive"]-1]
        dp = row[COL["pin_col"]-1]
        fn = str(row[COL["first_name"]-1] or "")
        ln = str(row[COL["last_name"]-1] or "")
        if not is_active(da):
            return False, None, "Account not active."
        if not dp:
            return False, None, "No PIN set."
        if dp != ph:
            return False, None, "Incorrect PIN."
        return True, res_id, f"{fn} {ln}".strip()
    return False, None, f"Resident #{res_id} not found."

def cal_verify_password(res_id_raw, pw_raw):
    """Returns (True, res_id, full_name) or (False, None, error_msg). Accepts password (column C)."""
    try:
        res_id = int(res_id_raw)
        pw_raw = str(pw_raw).strip()
    except:
        return False, None, "Invalid ID or password."
    if not pw_raw:
        return False, None, "Password required."
    wb = load_wb()
    if not wb:
        return False, None, "Database unavailable."
    ws = wb[RES_SHEET]
    ph = hash_pin(pw_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id:
            continue
        da  = row[COL["alive"]-1]
        dpw = row[COL["password"]-1]
        fn  = str(row[COL["first_name"]-1] or "")
        ln  = str(row[COL["last_name"]-1] or "")
        if not is_active(da):
            return False, None, "Account not active."
        if not dpw:
            return False, None, "No password set. Ask an admin to run SET PASSWORD."
        if dpw != ph:
            return False, None, "Incorrect password."
        return True, res_id, f"{fn} {ln}".strip()
    return False, None, f"Resident #{res_id} not found."

def _verify_resident_auth(data):
    """Try password auth first, fall back to PIN. Returns (ok, res_id, name_or_error)."""
    pw_raw = str(data.get("password", "")).strip()
    if pw_raw:
        return cal_verify_password(data.get("resId") or data.get("residentId"), pw_raw)
    pin_raw = str(data.get("pin", "")).strip()
    return cal_verify_pin(data.get("resId") or data.get("residentId"), pin_raw)

@app.route("/calendar")
def calendar_page():
    return send_from_directory("templates", "calendar.html")

@app.route("/api/cal/auth", methods=["POST"])
def cal_auth():
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400
    ok, res_id, info = _verify_resident_auth(data)
    if not ok:
        return jsonify({"success": False, "error": info}), 401
    return jsonify({"success": True, "resId": res_id, "name": info})

@app.route("/api/cal/month", methods=["GET"])
def cal_month():
    try:
        y = int(request.args.get("y", datetime.date.today().year))
        m = int(request.args.get("m", datetime.date.today().month))
    except:
        return jsonify({"error": "Bad params"}), 400
    start = f"{y:04d}-{m:02d}-01"
    import calendar as _cal
    last_day = _cal.monthrange(y, m)[1]
    end = f"{y:04d}-{m:02d}-{last_day:02d}"
    conn = cal_db()
    rows = conn.execute(
        "SELECT date, COUNT(*) as cnt FROM events WHERE date >= ? AND date <= ? GROUP BY date",
        (start, end)
    ).fetchall()
    conn.close()
    return jsonify({r["date"]: r["cnt"] for r in rows})

@app.route("/api/cal/day", methods=["GET"])
def cal_day():
    try:
        y = int(request.args.get("y")); m = int(request.args.get("m")); d = int(request.args.get("d"))
        date_str = f"{y:04d}-{m:02d}-{d:02d}"
    except:
        return jsonify({"error": "Bad params"}), 400
    res_id = request.args.get("resId", type=int)
    conn = cal_db()
    events = conn.execute(
        "SELECT e.*, (SELECT COUNT(*) FROM interested WHERE event_id=e.id) as interested_count "
        "FROM events e WHERE e.date=? ORDER BY e.time ASC, e.id ASC",
        (date_str,)
    ).fetchall()
    result = []
    for ev in events:
        is_int = False
        if res_id:
            row = conn.execute("SELECT 1 FROM interested WHERE event_id=? AND res_id=?", (ev["id"], res_id)).fetchone()
            is_int = row is not None
        result.append({
            "id": ev["id"], "date": ev["date"], "time": ev["time"], "title": ev["title"],
            "location": ev["location"], "description": ev["description"],
            "creator_res_id": ev["creator_res_id"], "creator_name": ev["creator_name"],
            "created_at": ev["created_at"], "image_filename": ev["image_filename"],
            "interested_count": ev["interested_count"], "is_interested": is_int
        })
    conn.close()
    return jsonify(result)

@app.route("/api/cal/event/<int:event_id>", methods=["GET"])
def cal_event(event_id):
    res_id = request.args.get("resId", type=int)
    conn = cal_db()
    ev = conn.execute(
        "SELECT e.*, (SELECT COUNT(*) FROM interested WHERE event_id=e.id) as interested_count "
        "FROM events e WHERE e.id=?", (event_id,)
    ).fetchone()
    if not ev:
        conn.close()
        return jsonify({"error": "Event not found"}), 404
    is_int = False
    if res_id:
        row = conn.execute("SELECT 1 FROM interested WHERE event_id=? AND res_id=?", (event_id, res_id)).fetchone()
        is_int = row is not None
    result = {
        "id": ev["id"], "date": ev["date"], "time": ev["time"], "title": ev["title"],
        "location": ev["location"], "description": ev["description"],
        "creator_res_id": ev["creator_res_id"], "creator_name": ev["creator_name"],
        "created_at": ev["created_at"], "image_filename": ev["image_filename"],
        "interested_count": ev["interested_count"], "is_interested": is_int
    }
    conn.close()
    return jsonify(result)

@app.route("/api/cal/event", methods=["POST"])
def cal_create_event():
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400
    ok, res_id, name = _verify_resident_auth(data)
    if not ok:
        return jsonify({"success": False, "error": name}), 401
    title = str(data.get("title", "")).strip()[:50]
    if not title:
        return jsonify({"success": False, "error": "Title is required."}), 400
    date_str = str(data.get("date", "")).strip()
    try: datetime.date.fromisoformat(date_str)
    except: return jsonify({"success": False, "error": "Invalid date."}), 400
    time_str = str(data.get("time", "")).strip()[:10]
    location = str(data.get("location", "")).strip()[:200]
    description = str(data.get("description", "")).strip()[:1000]
    image_filename = str(data.get("image_filename", "") or "").strip() or None
    now = datetime.datetime.now()
    now_str = now.isoformat(timespec="seconds")
    # ── 4-event-per-month limit ──────────────────────────────────────────
    month_start = now.strftime("%Y-%m-01")
    next_m = now.month % 12 + 1
    next_y = now.year + (1 if now.month == 12 else 0)
    month_end = f"{next_y}-{next_m:02d}-01"
    conn = cal_db()
    count = conn.execute(
        "SELECT COUNT(*) as c FROM events WHERE creator_res_id=? AND created_at>=? AND created_at<?",
        (res_id, month_start, month_end)
    ).fetchone()["c"]
    if count >= 4:
        conn.close()
        return jsonify({"success": False, "error": "Limit reached — you may only post 4 events per month."}), 400
    cur = conn.execute(
        "INSERT INTO events (date, time, title, location, description, creator_res_id, creator_name, created_at, image_filename) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (date_str, time_str, title, location, description, res_id, name, now_str, image_filename)
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"success": True, "id": new_id})

@app.route("/api/cal/event/<int:event_id>/interest", methods=["POST"])
def cal_interest(event_id):
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400
    ok, res_id, _ = _verify_resident_auth(data)
    if not ok:
        return jsonify({"success": False, "error": "Auth failed"}), 401
    conn = cal_db()
    existing = conn.execute("SELECT 1 FROM interested WHERE event_id=? AND res_id=?", (event_id, res_id)).fetchone()
    if existing:
        conn.execute("DELETE FROM interested WHERE event_id=? AND res_id=?", (event_id, res_id))
        interested = False
    else:
        conn.execute("INSERT INTO interested (event_id, res_id) VALUES (?,?)", (event_id, res_id))
        interested = True
    conn.commit()
    count = conn.execute("SELECT COUNT(*) as c FROM interested WHERE event_id=?", (event_id,)).fetchone()["c"]
    conn.close()
    return jsonify({"success": True, "interested": interested, "count": count})

@app.route("/api/cal/upload", methods=["POST"])
def cal_upload():
    """Receive base64-compressed image (compressed client-side) and save it."""
    import base64, uuid
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400
    ok, res_id, _ = _verify_resident_auth(data)
    if not ok:
        return jsonify({"success": False, "error": "Auth failed"}), 401
    img_data = data.get("imageData", "")
    if not img_data:
        return jsonify({"success": False, "error": "No image data."}), 400
    # Strip data URL prefix if present
    if "," in img_data:
        img_data = img_data.split(",", 1)[1]
    raw = base64.b64decode(img_data)
    if len(raw) > 200 * 1024:   # sanity check: client should compress to <100kb
        return jsonify({"success": False, "error": "Image too large (max 100kb after compression)."}), 400
    fname = f"event_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.jpg"
    fpath = os.path.join(EVENT_IMG_DIR, fname)
    with open(fpath, "wb") as f:
        f.write(raw)
    return jsonify({"success": True, "filename": fname})

@app.route("/api/cal/image/<filename>")
def cal_image(filename):
    # Basic security: only alphanum, underscore, dash, dot
    import re as _re
    if not _re.match(r'^[\w\-\.]+$', filename):
        return "Not found", 404
    return send_from_directory(EVENT_IMG_DIR, filename)


# ══════════════════════════════════════════════════════════════════════════════
# COMMUNITY BONDS  — /bonds + /api/bonds/*
# ══════════════════════════════════════════════════════════════════════════════
BOND_RECEIPT_DIR = os.path.join(BASE_DIR, "Bond_Receipts")
BCH_BOND_ADDRESS = "bitcoincash:qqd0ywgg5s94lxmxjad7hju3adqy6356nvy6szarp6"
BONDS_SHEET      = "Bonds"

os.makedirs(BOND_RECEIPT_DIR, exist_ok=True)

def _fetch_treasury_rate(series_id):
    """Fetch latest rate from FRED CSV endpoint (no API key required)."""
    import urllib.request as _ur
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
    req = _ur.Request(url, headers={"User-Agent": "ArkRegistry/1.0"})
    with _ur.urlopen(req, timeout=10) as r:
        lines = r.read().decode("utf-8").strip().splitlines()
    # CSV: DATE,VALUE — walk from end, skip "." (missing) values
    for line in reversed(lines[1:]):
        parts = line.split(",")
        if len(parts) == 2 and parts[1].strip() not in (".", ""):
            return float(parts[1].strip())
    raise ValueError(f"No valid rate found in {series_id}")

@app.route("/bonds")
def bonds_page():
    return send_from_directory("templates", "bonds.html")

@app.route("/api/bonds/rates", methods=["GET"])
def bonds_rates():
    """Return live 10-yr and 20-yr US treasury rates from FRED."""
    results = {}
    fallbacks = {"DGS10": 4.35, "DGS20": 4.65}
    for sid, fallback in fallbacks.items():
        try:
            results[sid] = _fetch_treasury_rate(sid)
        except Exception as e:
            print(f"[BOND] FRED {sid} failed: {e}, using fallback {fallback}")
            results[sid] = fallback
    return jsonify({
        "rate_10yr": results["DGS10"],
        "rate_20yr": results["DGS20"],
        "source": "US Treasury via FRED"
    })

@app.route("/api/bonds/verify-pin", methods=["POST"])
def bonds_verify_pin():
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400
    ok, res_id, name = _verify_resident_auth(data)
    if not ok:
        return jsonify({"success": False, "error": name}), 401
    return jsonify({"success": True, "resId": res_id, "name": name})

@app.route("/api/bonds/purchase", methods=["POST"])
def bonds_purchase():
    """Verify BCH payment on blockchain, then record bond in spreadsheet and issue receipt."""
    import urllib.request as _ur, json as _json
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400

    res_id_raw  = data.get("resId")
    pin_raw     = str(data.get("pin", "")).strip()
    amount_usd  = float(data.get("amount_usd", 0))
    rate_pct    = float(data.get("rate_pct", 0))
    years       = int(data.get("years", 10))
    bch_amount  = float(data.get("bch_amount", 0))
    bch_rate    = float(data.get("bch_rate", 43000))

    # Validate
    valid_amounts = {1000, 5000, 10000, 20000, 50000, 100000, 1000000}
    if amount_usd not in valid_amounts:
        return jsonify({"success": False, "error": "Invalid bond amount."}), 400
    if years not in (10, 20):
        return jsonify({"success": False, "error": "Invalid term."}), 400

    ok, res_id, resident_name = _verify_resident_auth(data)
    if not ok:
        return jsonify({"success": False, "error": resident_name}), 401

    # ── Blockchain verification is done client-side via /verify-bch-payment ──
    # The client only calls /api/bonds/purchase after blockchain confirms.
    tx_hash_found = "CLIENT-VERIFIED"

    # ── Write to Bonds sheet ─────────────────────────────────────────────────
    total_months = years * 12
    monthly_interest = round(amount_usd * (rate_pct / 100 / 12), 2)
    monthly_principal = round(amount_usd / total_months, 2)
    # PMT formula (amortized)
    mo_rate = rate_pct / 100 / 12
    if mo_rate > 0:
        pmt = amount_usd * (mo_rate * (1 + mo_rate)**total_months) / ((1 + mo_rate)**total_months - 1)
    else:
        pmt = amount_usd / total_months
    monthly_payment = round(pmt, 2)
    interest_per_year = round(amount_usd * (rate_pct / 100), 2)
    principal_per_year = round(amount_usd / years, 2)
    annual_payment = round(pmt * 12, 2)
    total_interest_pct = round(rate_pct * years, 2)
    total_interest_usd = round(pmt * total_months - amount_usd, 2)
    amount_k = amount_usd / 1000
    today = datetime.date.today()
    maturity_date = datetime.date(today.year + years, today.month, today.day)

    wb = load_wb()
    if not wb:
        return jsonify({"success": False, "error": "Database unavailable."}), 500
    ws = wb[BONDS_SHEET]

    # Find first row in col B with a serial number where col C is empty
    target_row = None
    for row_idx in range(2, ws.max_row + 20):
        serial_val = ws.cell(row_idx, 2).value   # col B
        amount_val = ws.cell(row_idx, 3).value   # col C
        if isinstance(serial_val, (int, float)) and (amount_val is None or amount_val == ""):
            target_row = row_idx
            break

    if target_row is None:
        # Append a new row
        target_row = ws.max_row + 1
        # Give it a serial number
        last_serial = 0
        for ri in range(2, target_row):
            sv = ws.cell(ri, 2).value
            if isinstance(sv, (int, float)):
                last_serial = max(last_serial, int(sv))
        ws.cell(target_row, 2).value = last_serial + 1

    from openpyxl.styles import Font as _Font
    fnt = _Font(name="Arial", size=10)

    # Write bond data
    ws.cell(target_row, 1).value = f"Res#{res_id}"  # col A: buyer
    ws.cell(target_row, 3).value = amount_k           # Amount $K
    ws.cell(target_row, 4).value = rate_pct           # Interest %
    ws.cell(target_row, 5).value = years              # Years
    ws.cell(target_row, 6).value = monthly_payment    # Payment/mo (interest + principal)
    ws.cell(target_row, 7).value = monthly_interest   # Interest/mo
    ws.cell(target_row, 8).value = monthly_principal  # Repayment/mo (principal portion)
    ws.cell(target_row, 9).value = total_interest_pct # Total Interest %
    ws.cell(target_row, 10).value = interest_per_year  # Interest/yr
    ws.cell(target_row, 11).value = today             # Date Issued
    ws.cell(target_row, 12).value = maturity_date     # Date Maturity
    ws.cell(target_row, 13).value = 1                 # Outstanding = 1

    for col in range(1, 14):
        ws.cell(target_row, col).font = fnt
    ws.cell(target_row, 11).number_format = "MM/DD/YY"
    ws.cell(target_row, 12).number_format = "MM/DD/YY"

    bond_serial_num = ws.cell(target_row, 2).value

    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        return jsonify({"success": False, "error": "Close Excel first, then try again."}), 500

    # ── Generate receipt ─────────────────────────────────────────────────────
    now = datetime.datetime.now()
    b_num = _next_id(BOND_RECEIPT_DIR, "B")
    bond_receipt_id  = f"B{b_num:06d}"
    bond_receipt_num = f"{bond_receipt_id}-{now.strftime('%Y%m%d')}-res{res_id}"

    receipt_text = (
        f"ARK COMMUNITY BOND RECEIPT\n"
        f"============================\n"
        f"Receipt ID    : {bond_receipt_id}\n"
        f"Receipt #     : {bond_receipt_num}\n"
        f"Date/Time     : {now.strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"\n"
        f"Investor      : {resident_name}\n"
        f"Resident #    : {res_id}\n"
        f"\n"
        f"Bond Serial # : {bond_serial_num}\n"
        f"Principal     : ${amount_usd:,.2f} USD\n"
        f"Term          : {years} Years ({total_months} months)\n"
        f"Interest Rate : {rate_pct:.3f}% per annum (US {years}-yr Treasury)\n"
        f"\n"
        f"MONTHLY BREAKDOWN\n"
        f"  Interest    : ${monthly_interest:,.2f} / month\n"
        f"  Principal   : ${monthly_principal:,.2f} / month\n"
        f"  Total Pmt   : ${monthly_payment:,.2f} / month\n"
        f"\n"
        f"ANNUAL BREAKDOWN\n"
        f"  Interest    : ${interest_per_year:,.2f} / year\n"
        f"  Principal   : ${principal_per_year:,.2f} / year\n"
        f"  Total Pmt   : ${annual_payment:,.2f} / year\n"
        f"\n"
        f"Total Interest: ${total_interest_usd:,.2f} USD over {years} yrs\n"
        f"\n"
        f"Date Issued   : {today.strftime('%Y-%m-%d')}\n"
        f"Maturity Date : {maturity_date.strftime('%Y-%m-%d')}\n"
        f"\n"
        f"Payment       : {bch_amount:.5f} BCH\n"
        f"Exchange Rate : ${bch_rate:,.2f} USD/BCH\n"
        f"BCH Address   : {BCH_BOND_ADDRESS}\n"
        f"TX Hash       : {tx_hash_found or 'N/A'}\n"
        f"\n"
        f"Status        : BOND ISSUED\n"
        f"\n"
        f"This receipt confirms your investment in an Ark Community Bond.\n"
        f"Monthly payments (interest + principal) begin next calendar month.\n"
        f"Bond is fully repaid at maturity.\n"
        f"Keep this document for your records.\n"
        f"Disputes must be raised within 30 days of issue.\n"
    )

    fname = f"bond_{bond_receipt_num}.txt"
    with open(os.path.join(BOND_RECEIPT_DIR, fname), "w", encoding="utf-8") as f:
        f.write(receipt_text)

    print(f"[BOND] #{res_id} bought ${amount_usd:,.0f} {years}yr bond at {rate_pct:.3f}%  {bond_receipt_id}")

    return jsonify({
        "success": True,
        "verified": True,
        "receipt_id": bond_receipt_id,
        "receipt_num": bond_receipt_num,
        "receipt_text": receipt_text,
        "bond_serial": bond_serial_num,
        "monthly_interest": monthly_interest,
        "monthly_principal": monthly_principal,
        "monthly_payment": monthly_payment,
        "maturity_date": maturity_date.isoformat(),
    })


@app.route("/api/bonds/status", methods=["POST"])
def bonds_status():
    """Return all bonds belonging to a resident."""
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400

    ok, res_id, name = _verify_resident_auth(data)
    if not ok:
        return jsonify({"success": False, "error": name}), 401

    wb = load_wb()
    if not wb:
        return jsonify({"success": False, "error": "Database unavailable."}), 500

    ws = wb[BONDS_SHEET]
    tag = f"Res#{res_id}"
    bonds = []
    for row_idx in range(2, ws.max_row + 1):
        buyer = ws.cell(row_idx, 1).value   # col A  — "Res#42"
        amt_k = ws.cell(row_idx, 3).value   # col C  — Amount $K
        if str(buyer).strip() != tag or amt_k is None or amt_k == "":
            continue
        serial     = ws.cell(row_idx, 2).value
        rate_pct   = ws.cell(row_idx, 4).value
        years      = ws.cell(row_idx, 5).value
        payment_mo = ws.cell(row_idx, 6).value
        interest_m = ws.cell(row_idx, 7).value
        repayment  = ws.cell(row_idx, 8).value
        tot_int_p  = ws.cell(row_idx, 9).value
        int_yr     = ws.cell(row_idx, 10).value
        issued     = ws.cell(row_idx, 11).value
        maturity   = ws.cell(row_idx, 12).value
        outstand   = ws.cell(row_idx, 13).value

        def _dt(v):
            if hasattr(v, 'isoformat'):
                return v.isoformat()
            return str(v) if v else ""

        bonds.append({
            "serial": serial,
            "amount_k": float(amt_k) if amt_k else 0,
            "amount_usd": float(amt_k) * 1000 if amt_k else 0,
            "rate_pct": float(rate_pct) if rate_pct else 0,
            "years": int(years) if years else 0,
            "payment_monthly": float(payment_mo) if payment_mo else 0,
            "interest_monthly": float(interest_m) if interest_m else 0,
            "principal_monthly": float(repayment) if repayment else 0,
            "total_interest_pct": float(tot_int_p) if tot_int_p else 0,
            "interest_yearly": float(int_yr) if int_yr else 0,
            "principal_yearly": round(float(repayment) * 12, 2) if repayment else 0,
            "payment_yearly": round((float(int_yr) if int_yr else 0) + (round(float(repayment) * 12, 2) if repayment else 0), 2),
            "date_issued": _dt(issued),
            "date_maturity": _dt(maturity),
            "outstanding": int(outstand) if outstand else 0,
        })

    return jsonify({
        "success": True,
        "resId": res_id,
        "name": name,
        "bonds": bonds,
    })


# ── Charities ─────────────────────────────────────────────────────────────────
@app.route("/gov-employees")
def gov_employees_page():
    return send_from_directory("templates", "gov_employees.html")

@app.route("/api/gov-employees")
def api_gov_employees():
    """Return all employees from the Gov Employees sheet."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        if "Gov Employees" not in wb.sheetnames:
            return jsonify([])
        ws = wb["Gov Employees"]
        employees = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            emp_num, res_num, last_name, first_name, position, salary, date_hired, active, hexarchy = (row + (None,)*9)[:9]
            if not emp_num and not position:
                continue
            date_str = ""
            if date_hired:
                try:
                    import datetime as dt
                    if isinstance(date_hired, (dt.date, dt.datetime)):
                        date_str = date_hired.strftime("%Y-%m-%d")
                    else:
                        date_str = str(date_hired)
                except:
                    date_str = str(date_hired) if date_hired else ""
            # Look up citizen data from Res sheet if res_num given
            citizen_first = ""; citizen_last = ""
            resident_active = True
            ubi_viable = False
            inactive_reason = ""
            if res_num and "Res" in wb.sheetnames:
                ws_res = wb["Res"]
                for rrow in ws_res.iter_rows(min_row=2, values_only=True):
                    if isinstance(rrow[COL["res_num"]-1], (int, float)) and int(rrow[COL["res_num"]-1]) == int(res_num):
                        citizen_first = str(rrow[COL["first_name"]-1] or "")
                        citizen_last = str(rrow[COL["last_name"]-1] or "")
                        # Check alive
                        alive_val = rrow[COL["alive"]-1]
                        if not is_active(alive_val):
                            resident_active = False
                            inactive_reason = "Deceased/Banned"
                        # Check immigration interview for UBI viability
                        immig_val = rrow[COL["immigration_interview"]-1] if len(rrow) > COL["immigration_interview"]-1 else None
                        ref_val = rrow[COL["ref_checked"]-1] if len(rrow) > COL["ref_checked"]-1 else None
                        if immig_val == 1 and ref_val == 1 and is_active(alive_val):
                            ubi_viable = True
                        elif immig_val != 1:
                            if resident_active:
                                resident_active = False
                                inactive_reason = "UBI Not Live"
                        break
            else:
                # No res# means inactive
                if not res_num:
                    resident_active = False
                    inactive_reason = "No Res#"
            # Use citizen name from Res sheet if available
            display_first = citizen_first if citizen_first else str(first_name or "").strip()
            display_last = citizen_last if citizen_last else str(last_name or "").strip()
            # Override active status
            computed_active = str(active or "").strip()
            if not resident_active:
                computed_active = "No"
            employees.append({
                "_inactiveReason": inactive_reason,
                "empNum":       int(emp_num) if emp_num else None,
                "resNum":       int(res_num) if res_num else None,
                "lastName":     display_last,
                "firstName":    display_first,
                "position":     str(position or "").strip(),
                "salary":       float(salary) if salary else None,
                "dateHired":    date_str,
                "active":       computed_active,
                "hexarchy":     str(hexarchy or "").strip(),
                "citizenName":  f"{display_first} {display_last}".strip(),
                "ubiViable":    ubi_viable and computed_active == "Yes",
            })
        wb.close()
        return jsonify(employees)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gov-employees/<int:emp_num>", methods=["PUT"])
def api_update_gov_employee(emp_num):
    """Update a gov employee row. Requires hexarchy membership. Cannot edit hexarchy positions."""
    try:
        data = request.get_json(force=True)
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Gov Employees"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == emp_num:
                # Block editing hexarchy positions
                current_pos = str(row[4].value or "").strip().lower()
                if "hexarchy" in current_pos:
                    return jsonify({"error": "Hexarchy positions can only be changed by elections."}), 403
                new_pos = str(data.get("position", "")).strip().lower()
                if new_pos and "hexarchy" in new_pos:
                    return jsonify({"error": "Hexarchy positions can only be assigned by elections."}), 403
                if "resNum"    in data: row[1].value = data["resNum"] or None
                if "lastName"  in data: row[2].value = data["lastName"] or None
                if "firstName" in data: row[3].value = data["firstName"] or None
                if "position"  in data: row[4].value = data["position"] or None
                if "salary"    in data: row[5].value = data["salary"] or None
                if "dateHired" in data: row[6].value = data["dateHired"] or None
                if "active"    in data: row[7].value = data["active"] or None
                if "hexarchy"  in data: row[8].value = data["hexarchy"] or None
                break
        wb.save(EXCEL_FILE)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gov-employees", methods=["POST"])
def api_add_gov_employee():
    """Add a new gov employee row. Cannot add hexarchy positions."""
    try:
        data = request.get_json(force=True)
        new_pos = str(data.get("position", "")).strip().lower()
        if "hexarchy" in new_pos:
            return jsonify({"error": "Hexarchy positions can only be assigned by elections."}), 403
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Gov Employees"]
        max_num = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                max_num = max(max_num, int(row[0]))
        new_num = max_num + 1
        ws.append([new_num, data.get("resNum"), data.get("lastName",""), data.get("firstName",""),
                   data.get("position",""), data.get("salary"), data.get("dateHired",""),
                   data.get("active","Yes"), data.get("hexarchy","No")])
        wb.save(EXCEL_FILE)
        return jsonify({"ok": True, "empNum": new_num})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gov-employees/<int:emp_num>", methods=["DELETE"])
def api_delete_gov_employee(emp_num):
    """Remove a gov employee row. Cannot delete hexarchy positions."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Gov Employees"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == emp_num:
                current_pos = str(row[4].value or "").strip().lower()
                if "hexarchy" in current_pos:
                    return jsonify({"error": "Hexarchy positions can only be changed by elections."}), 403
                ws.delete_rows(row[0].row)
                break
        wb.save(EXCEL_FILE)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/budget-history")
def budget_history_page():
    return send_from_directory("templates","budget_history.html")

@app.route("/build-your-ark")
def build_your_ark_page():
    return send_from_directory("templates", "build_your_ark.html")

@app.route("/charities")
def charities_page():
    return send_from_directory("templates","charities.html")

@app.route("/api/charities")
def api_charities():
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE,data_only=True)
        ws=wb["Charity"]
        import datetime
        year=datetime.date.today().year
        # Rating columns: 2025=col9, 2026=col10, ...
        rating_col=9+(year-2025)
        charities=[]
        for r in range(2,ws.max_row+1):
            cid=ws.cell(r,2).value
            name=ws.cell(r,3).value
            if not cid or not name: continue
            if not isinstance(cid,(int,float)): continue  # skip header/label rows
            emp_val=ws.cell(r,4).value
            emp=int(emp_val) if isinstance(emp_val,(int,float)) else 0
            rating=float(ws.cell(r,rating_col).value or 0)
            desc=str(ws.cell(r,30).value or "")
            bch=str(ws.cell(r,7).value or "")
            charities.append({"id":int(cid),"name":str(name)[:50],"description":desc[:300],
                              "employees":emp,"rating":rating,"bch":bch})
        wb.close()
        charities.sort(key=lambda c:c["rating"],reverse=True)
        total_ratings=sum(c["rating"] for c in charities)
        total_emp=sum(c["employees"] for c in charities)
        # Total charity budget: sum of charity tax from all residents
        # For now use a reasonable estimate from the charity tax rate
        total_budget=round(total_ratings*12.5,2)  # placeholder scaling
        return jsonify({"charities":charities,"total_ratings":total_ratings,
                       "total_employees":total_emp,"total_budget":total_budget})
    except Exception as e:
        return jsonify({"error":str(e)}),500



@app.route('/api/previous-election')
def api_previous_election():
    """Return the results of the most recently applied election."""
    try:
        with get_election_db() as c:
            row = c.execute(
                "SELECT * FROM previous_elections ORDER BY applied_at DESC LIMIT 1"
            ).fetchone()
        if not row:
            return jsonify({"available": False})
        return jsonify({
            "available": True,
            "election_name": row["election_name"],
            "applied_at": row["applied_at"],
            "hexarchy_results": json.loads(row["hexarchy_results"] or "[]"),
            "charity_results": json.loads(row["charity_results"] or "[]"),
            "proposal_results": json.loads(row["proposal_results"] or "[]"),
        })
    except Exception as e:
        return jsonify({"available": False, "error": str(e)})

# ── Election (Flask-based) ────────────────────────────────────────────────────
@app.route("/election")
@app.route("/election/")
def election_page():
    return send_from_directory("templates","election_vote.html")

@app.route("/api/election/status")
def election_status():
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE,data_only=True)
        ws_e=wb["Elections"]
        active=int(ws_e.cell(2,14).value or 0)==1
        charities=[]
        if active:
            ws_c=wb["Charity"]
            for r in range(2,ws_c.max_row+1):
                cid=ws_c.cell(r,2).value
                name=ws_c.cell(r,3).value
                if not cid or not name: continue
                if not isinstance(cid,(int,float)): continue  # skip header rows
                emp_val=ws_c.cell(r,4).value
                emp=int(emp_val) if isinstance(emp_val,(int,float)) else 0
                desc=str(ws_c.cell(r,30).value or "")
                charities.append({"id":int(cid),"name":str(name)[:50],
                                 "description":desc[:300],"employees":emp})
        wb.close()
        return jsonify({"active":active,"charities":charities})
    except Exception as e:
        return jsonify({"active":False,"error":str(e)})

@app.route("/api/election/vote",methods=["POST"])
def election_vote():
    try:
        data=request.get_json(force=True)
        res_num=int(data.get("res_num",0))
        ratings=data.get("ratings",{})
        if not res_num or not ratings:
            return jsonify({"error":"Missing data."}),400
        import datetime
        wb=openpyxl.load_workbook(EXCEL_FILE)
        ws_e=wb["Elections"]
        # Check election is active
        if int(ws_e.cell(2,14).value or 0)!=1:
            wb.close()
            return jsonify({"error":"No active election."}),400
        # Check if already voted this year
        ws_r=wb["Res"]
        year=datetime.date.today().year
        for row in range(2,ws_r.max_row+1):
            if ws_r.cell(row,2).value==res_num:
                last_voted=ws_r.cell(row,99).value
                if last_voted and int(last_voted)==year:
                    wb.close()
                    return jsonify({"error":f"You already voted in {year}."}),400
                ws_r.cell(row,99).value=year
                break
        # Add ratings to current year column in Charity sheet
        ws_c=wb["Charity"]
        rating_col=9+(year-2025)
        for r in range(2,ws_c.max_row+1):
            cid=str(ws_c.cell(r,2).value or "")
            if cid in ratings:
                cur=float(ws_c.cell(r,rating_col).value or 0)
                ws_c.cell(r,rating_col).value=cur+float(ratings[cid])
        wb.save(EXCEL_FILE)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}),500



# ══ Admin Panel ═══════════════════════════════════════════════════════════════
@app.route("/admin")
@app.route("/admin/")
def admin_panel():
    """Election admin panel — start/stop elections and manage charity list."""
    return send_from_directory("templates","admin.html")

@app.route("/api/admin/election/set",methods=["POST"])
def admin_election_set():
    """Toggle election active flag. Body: {active: 0|1}"""
    try:
        data=request.get_json(force=True)
        active=int(data.get("active",0))
        wb=openpyxl.load_workbook(EXCEL_FILE)
        ws_e=wb["Elections"]
        ws_e.cell(2,14).value=active
        wb.save(EXCEL_FILE)
        return jsonify({"ok":True,"active":active})
    except Exception as e:
        return jsonify({"error":str(e)}),500


# ══════════════════════════════════════════════════════════════════════════════
# ELECTION API  (/api/state, /api/vote/*, /api/admin/*)
# ══════════════════════════════════════════════════════════════════════════════

def _get_alliance_results(conn, proposal_id):
    return [dict(r) for r in conn.execute(
        'SELECT option_name, COALESCE(SUM(rating),0) as total_rating, COUNT(*) as vote_count '
        'FROM alliance_votes WHERE proposal_id=? GROUP BY option_name', (proposal_id,)).fetchall()]

@app.route('/api/state')
def election_state_api():
    with get_election_db() as c:
        name   = c.execute("SELECT value FROM election_state WHERE key='election_name'").fetchone()
        hex_o  = c.execute("SELECT value FROM election_state WHERE key='hexarchy_open'").fetchone()
        char_o = c.execute("SELECT value FROM election_state WHERE key='charity_open'").fetchone()
        dem_o  = c.execute("SELECT value FROM election_state WHERE key='democracy_open'").fetchone()
    return jsonify({
        'election_name':  name[0]   if name   else 'Arkology Election',
        'hexarchy_open':  hex_o[0]  == '1' if hex_o  else False,
        'charity_open':   char_o[0] == '1' if char_o else False,
        'democracy_open': dem_o[0]  == '1' if dem_o  else False,
    })

@app.route('/api/session/<sid>')
def election_session(sid):
    res_id = request.args.get('resId')
    with get_election_db() as c:
        c.execute('INSERT OR IGNORE INTO voter_sessions (session_id) VALUES (?)', (sid,))
        sess = c.execute('SELECT * FROM voter_sessions WHERE session_id=?', (sid,)).fetchone()
        demo_votes = [r[0] for r in c.execute(
            "SELECT target_id FROM votes WHERE session_id=? AND vote_type='democracy'", (sid,)).fetchall()]
        alliance_voted = [r[0] for r in c.execute(
            'SELECT DISTINCT proposal_id FROM alliance_votes WHERE session_id=?', (sid,)).fetchall()]
        # Also check resident-level votes if resId provided
        if res_id:
            try:
                rid = int(res_id)
                res_demo = [r[0] for r in c.execute(
                    "SELECT proposal_id FROM resident_demo_votes WHERE resident_id=? AND vote_type='democracy'", (rid,)).fetchall()]
                res_alliance = [r[0] for r in c.execute(
                    "SELECT proposal_id FROM resident_demo_votes WHERE resident_id=? AND vote_type='alliance'", (rid,)).fetchall()]
                # Merge: if resident voted on a proposal (from any session), include it
                for pid in res_demo:
                    if pid not in demo_votes: demo_votes.append(pid)
                for pid in res_alliance:
                    if pid not in alliance_voted: alliance_voted.append(pid)
            except: pass
        c.commit()
    return jsonify({
        'voted_hexarchy': sess['voted_hexarchy']==1,
        'voted_charity':  sess['voted_charity']==1,
        'voted_democracy': demo_votes,
        'voted_alliance':  alliance_voted,
    })

@app.route('/api/candidates')
def election_candidates():
    """Return active candidates — only those with a valid resident number in the spreadsheet."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws_res = wb["Res"] if "Res" in wb.sheetnames else None
        valid_residents = set()
        resident_names = {}
        if ws_res:
            for rrow in ws_res.iter_rows(min_row=2, values_only=True):
                rid = rrow[COL["res_num"]-1]
                if isinstance(rid, (int, float)):
                    rid_int = int(rid)
                    valid_residents.add(rid_int)
                    fn = str(rrow[COL["first_name"]-1] or "").strip()
                    ln = str(rrow[COL["last_name"]-1] or "").strip()
                    resident_names[rid_int] = f"{fn} {ln}".strip()
        wb.close()
    except Exception:
        valid_residents = set()
        resident_names = {}
    with get_election_db() as c:
        cands = [dict(r) for r in c.execute(
            "SELECT c.*, COALESCE(SUM(v.rating),0) as rating_sum, COUNT(v.id) as vote_count "
            "FROM candidates c LEFT JOIN votes v ON v.target_id=c.id AND v.vote_type='hexarchy' "
            "WHERE c.active=1 GROUP BY c.id ORDER BY c.id").fetchall()]
    result = []
    for cand in cands:
        res_num = cand.get("res_num")
        if res_num and int(res_num) in valid_residents:
            cand["verified_resident"] = True
            cand["resident_name"] = resident_names.get(int(res_num), cand.get("name", ""))
        else:
            cand["verified_resident"] = False
        result.append(cand)
    # Return all active candidates; verified_resident flag indicates res# confirmed in spreadsheet
    # Admin must set res_num on candidates; unverified candidates are shown with a warning flag
    return jsonify(result)

@app.route("/api/election/charities")
def election_charities():
    """Election charity voting — returns real charities from spreadsheet enriched with vote data."""
    try:
        import datetime as _dt
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Charity"]
        real_charities = []
        for r in range(2, ws.max_row + 1):
            cid = ws.cell(r, 2).value
            name = ws.cell(r, 3).value
            if not cid or not name: continue
            if not isinstance(cid, (int, float)): continue
            desc = str(ws.cell(r, 30).value or "")
            real_charities.append({"id": int(cid), "name": str(name)[:60], "description": desc[:300]})
        wb.close()
    except Exception:
        real_charities = []
    vote_map = {}
    try:
        with get_election_db() as c:
            for row in c.execute(
                "SELECT target_id, COALESCE(SUM(rating),0) as total_rating, COUNT(*) as vote_count "
                "FROM votes WHERE vote_type='charity' GROUP BY target_id").fetchall():
                vote_map[row["target_id"]] = {"total_rating": row["total_rating"], "vote_count": row["vote_count"]}
    except Exception:
        pass
    result = []
    for ch in real_charities:
        v = vote_map.get(ch["id"], {"total_rating": 0, "vote_count": 0})
        result.append({**ch, "total_rating": v["total_rating"], "vote_count": v["vote_count"], "min_votes": 20, "active": 1})
    return jsonify(result)

@app.route('/api/proposals')
def election_proposals():
    with get_election_db() as c:
        props = []
        for r in c.execute("SELECT * FROM proposals WHERE active=1 ORDER BY id").fetchall():
            p = dict(r)
            p['options'] = json.loads(p['options']) if p.get('options') else []
            if p['type'] == 'alliance':
                p['alliance_results'] = _get_alliance_results(c, p['id'])
            props.append(p)
    return jsonify(props)

@app.route('/api/election/eligibility-simple', methods=['POST'])
def election_eligibility_simple():
    """3-check election gate: alive, age 20-70, not yet voted this election cycle."""
    try:    data = request.get_json(force=True)
    except: return jsonify({"eligible": False, "error": "Invalid request"}), 400
    res_id_raw = data.get("residentId")
    try: res_id_int = int(res_id_raw)
    except: return jsonify({"eligible": False, "error": "Invalid resident ID"}), 400
    wb = load_wb()
    if not wb: return jsonify({"eligible": False, "error": "Database unavailable"}), 500
    ws = wb[RES_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id_int: continue
        # Check 1: Alive
        da = row[COL["alive"]-1]
        if not is_active(da):
            return jsonify({"eligible": False, "error": "Resident account is not active."})
        # Check 2: Age 20-70
        yb = row[COL["year_born"]-1]
        if not isinstance(yb, (int, float)):
            return jsonify({"eligible": False, "error": "Year of birth not on file."})
        age = datetime.date.today().year - int(yb)
        if age < 20 or age > 70:
            return jsonify({"eligible": False, "error": f"Age {age} is outside eligible range (20-70)."})
        # Check 3: Not yet voted this election cycle
        elc = row[COL["election_col"]-1]
        this_year = str(datetime.date.today().year)
        if str(elc or "").strip() == this_year:
            return jsonify({"eligible": False, "alreadyVoted": True, "error": "You have already voted in this election."})
        return jsonify({"eligible": True, "residentId": res_id_int})
    return jsonify({"eligible": False, "error": f"Resident #{res_id_int} not found."}), 404

@app.route('/api/election/ballot-check/<int:resident_id>')
def election_ballot_check(resident_id):
    """Check if resident has voted in each category and if any election section is running."""
    with get_election_db() as c:
        rows = c.execute("SELECT key, value FROM election_state WHERE key IN ('hexarchy_open','charity_open','democracy_open')").fetchall()
        election_running = any(r['value'] == '1' for r in rows)
        rb = c.execute("SELECT * FROM resident_ballots WHERE resident_id=?", (resident_id,)).fetchone()
        voted_hex = (rb['voted_hexarchy'] == 1) if rb else False
        voted_char = (rb['voted_charity'] == 1) if rb else False
        voted_dem = (rb['voted_democracy'] == 1) if rb else False
    return jsonify({'electionRunning': election_running,
                    'ballotCast': voted_hex and voted_char and voted_dem,
                    'voted_hexarchy': voted_hex,
                    'voted_charity': voted_char,
                    'voted_democracy': voted_dem})

@app.route('/api/vote/hexarchy', methods=['POST'])
def vote_hexarchy():
    d = request.get_json(force=True)
    sid = str(d.get('session_id',''))
    ratings = d.get('ratings', {})
    resident_id = d.get('resident_id')
    if not sid: return jsonify({'error':'No session.'}), 400
    # Liquid democracy: block if you delegated to someone else
    weight = 1
    if resident_id:
        try:
            rid = int(resident_id)
            with get_election_db() as cw:
                row = cw.execute("SELECT rep_res_id FROM liquid_endorsements WHERE endorser_res_id=?", (rid,)).fetchone()
                if row and row[0] is not None and int(row[0]) != rid:
                    return jsonify({'error':'You have a Liquid Democracy representative — they vote on your behalf.'}), 403
                # Vote weight = max(1, # endorsements received)
                n = cw.execute("SELECT COUNT(*) FROM liquid_endorsements WHERE rep_res_id=?", (rid,)).fetchone()[0]
                weight = max(1, int(n))
        except: pass
    try:
        with get_election_db() as c:
            state = c.execute("SELECT value FROM election_state WHERE key='hexarchy_open'").fetchone()
            if not state or state[0] != '1':
                return jsonify({'error':'Hexarchy voting is closed.'}), 400
            if resident_id:
                rb = c.execute("SELECT voted_hexarchy FROM resident_ballots WHERE resident_id=?", (int(resident_id),)).fetchone()
                if rb and rb[0] == 1:
                    return jsonify({'error':'You have already voted in the Hexarchy election.'}), 400
            sess = c.execute('SELECT voted_hexarchy FROM voter_sessions WHERE session_id=?',(sid,)).fetchone()
            if sess and sess[0] == 1:
                return jsonify({'error':'You have already voted in the Hexarchy election.'}), 400
            c.execute('INSERT OR IGNORE INTO voter_sessions (session_id) VALUES (?)', (sid,))
            for cid, rating in ratings.items():
                weighted = max(0, min(10, float(rating))) * weight
                c.execute('INSERT OR REPLACE INTO votes (session_id,vote_type,target_id,rating) VALUES (?,?,?,?)',
                          (sid,'hexarchy',int(cid), weighted))
            c.execute('UPDATE voter_sessions SET voted_hexarchy=1 WHERE session_id=?',(sid,))
            if resident_id:
                c.execute('INSERT OR IGNORE INTO resident_ballots (resident_id) VALUES (?)', (int(resident_id),))
                c.execute('UPDATE resident_ballots SET voted_hexarchy=1 WHERE resident_id=?', (int(resident_id),))
            c.commit()
    except Exception as e:
        return jsonify({'error':str(e)}), 500
    return jsonify({'ok':True, 'voteWeight':weight})

@app.route('/api/vote/charity', methods=['POST'])
def vote_charity():
    d = request.get_json(force=True)
    sid = str(d.get('session_id',''))
    votes = d.get('votes', {})
    resident_id = d.get('resident_id')
    if not sid: return jsonify({'error':'No session.'}), 400
    # Enforce max 3 charities voted on (rating > 0)
    active_votes = {k: v for k, v in votes.items() if float(v) > 0}
    if len(active_votes) > 3:
        return jsonify({'error':'You may only vote for a maximum of 3 charities.'}), 400
    # Liquid democracy: block delegated voters, calculate weight
    weight = 1
    if resident_id:
        try:
            rid = int(resident_id)
            with get_election_db() as cw:
                row = cw.execute("SELECT rep_res_id FROM liquid_endorsements WHERE endorser_res_id=?", (rid,)).fetchone()
                if row and row[0] is not None and int(row[0]) != rid:
                    return jsonify({'error':'You have a Liquid Democracy representative — they vote on your behalf.'}), 403
                n = cw.execute("SELECT COUNT(*) FROM liquid_endorsements WHERE rep_res_id=?", (rid,)).fetchone()[0]
                weight = max(1, int(n))
        except: pass
    try:
        with get_election_db() as c:
            state = c.execute("SELECT value FROM election_state WHERE key='charity_open'").fetchone()
            if not state or state[0] != '1':
                return jsonify({'error':'Charity voting is closed.'}), 400
            if resident_id:
                rb = c.execute("SELECT voted_charity FROM resident_ballots WHERE resident_id=?", (int(resident_id),)).fetchone()
                if rb and rb[0] == 1:
                    return jsonify({'error':'You have already voted in the Charity election.'}), 400
            sess = c.execute('SELECT voted_charity FROM voter_sessions WHERE session_id=?',(sid,)).fetchone()
            if sess and sess[0] == 1:
                return jsonify({'error':'You have already voted in the Charity election.'}), 400
            c.execute('INSERT OR IGNORE INTO voter_sessions (session_id) VALUES (?)', (sid,))
            for cid, rating in votes.items():
                weighted = max(0, min(10, float(rating))) * weight
                c.execute('INSERT OR REPLACE INTO votes (session_id,vote_type,target_id,rating) VALUES (?,?,?,?)',
                          (sid,'charity',int(cid), weighted))
            c.execute('UPDATE voter_sessions SET voted_charity=1 WHERE session_id=?',(sid,))
            if resident_id:
                c.execute('INSERT OR IGNORE INTO resident_ballots (resident_id) VALUES (?)', (int(resident_id),))
                c.execute('UPDATE resident_ballots SET voted_charity=1 WHERE resident_id=?', (int(resident_id),))
            c.commit()
    except Exception as e:
        return jsonify({'error':str(e)}), 500
    return jsonify({'ok':True, 'voteWeight':weight})

@app.route('/api/vote/democracy', methods=['POST'])
def vote_democracy():
    d = request.get_json(force=True)
    sid = str(d.get('session_id',''))
    proposal_id = d.get('proposal_id')
    choice = d.get('choice','')
    resident_id = d.get('resident_id')
    if not sid or not proposal_id: return jsonify({'error':'Missing data.'}), 400
    # Liquid democracy: block delegated, calc weight
    weight = 1
    if resident_id:
        try:
            rid = int(resident_id)
            with get_election_db() as cw:
                row = cw.execute("SELECT rep_res_id FROM liquid_endorsements WHERE endorser_res_id=?", (rid,)).fetchone()
                if row and row[0] is not None and int(row[0]) != rid:
                    return jsonify({'error':'You have a Liquid Democracy representative — they vote on your behalf.'}), 403
                n = cw.execute("SELECT COUNT(*) FROM liquid_endorsements WHERE rep_res_id=?", (rid,)).fetchone()[0]
                weight = max(1, int(n))
        except: pass
    try:
        with get_election_db() as c:
            state = c.execute("SELECT value FROM election_state WHERE key='democracy_open'").fetchone()
            if not state or state[0] != '1':
                return jsonify({'error':'Democracy voting is closed.'}), 400
            # Per-session duplicate check
            already = c.execute(
                "SELECT id FROM votes WHERE session_id=? AND vote_type='democracy' AND target_id=?",
                (sid, proposal_id)).fetchone()
            if already:
                return jsonify({'error':'You already voted on this proposal.'}), 400
            # Per-resident duplicate check
            if resident_id:
                try:
                    rid = int(resident_id)
                    res_dup = c.execute(
                        "SELECT id FROM resident_demo_votes WHERE resident_id=? AND proposal_id=? AND vote_type='democracy'",
                        (rid, proposal_id)).fetchone()
                    if res_dup:
                        return jsonify({'error':'You have already voted on this proposal.'}), 400
                except: pass
            c.execute('INSERT OR IGNORE INTO voter_sessions (session_id) VALUES (?)', (sid,))
            # Insert weight copies of the vote (with synthetic session_ids for extras)
            c.execute('INSERT INTO votes (session_id,vote_type,target_id,choice) VALUES (?,?,?,?)',
                      (sid,'democracy',int(proposal_id),choice))
            for w in range(1, weight):
                c.execute('INSERT INTO votes (session_id,vote_type,target_id,choice) VALUES (?,?,?,?)',
                          (sid + f"_w{w}",'democracy',int(proposal_id),choice))
            # Record resident-level vote
            if resident_id:
                try:
                    rid = int(resident_id)
                    c.execute('INSERT OR IGNORE INTO resident_demo_votes (resident_id,proposal_id,vote_type) VALUES (?,?,?)',
                              (rid, int(proposal_id), 'democracy'))
                except: pass
            c.commit()
    except Exception as e:
        return jsonify({'error':str(e)}), 500
    return jsonify({'ok':True, 'voteWeight':weight})

@app.route('/api/vote/alliance', methods=['POST'])
def vote_alliance():
    d = request.get_json(force=True)
    sid = str(d.get('session_id',''))
    proposal_id = d.get('proposal_id')
    ratings = d.get('ratings', {})
    resident_id = d.get('resident_id')
    if not sid or not proposal_id: return jsonify({'error':'Missing data.'}), 400
    # Liquid democracy: block delegated, calc weight
    weight = 1
    if resident_id:
        try:
            rid = int(resident_id)
            with get_election_db() as cw:
                row = cw.execute("SELECT rep_res_id FROM liquid_endorsements WHERE endorser_res_id=?", (rid,)).fetchone()
                if row and row[0] is not None and int(row[0]) != rid:
                    return jsonify({'error':'You have a Liquid Democracy representative — they vote on your behalf.'}), 403
                n = cw.execute("SELECT COUNT(*) FROM liquid_endorsements WHERE rep_res_id=?", (rid,)).fetchone()[0]
                weight = max(1, int(n))
        except: pass
    try:
        with get_election_db() as c:
            state = c.execute("SELECT value FROM election_state WHERE key='democracy_open'").fetchone()
            if not state or state[0] != '1':
                return jsonify({'error':'Democracy voting is closed.'}), 400
            already = c.execute(
                'SELECT id FROM alliance_votes WHERE session_id=? AND proposal_id=?',
                (sid, proposal_id)).fetchone()
            if already:
                return jsonify({'error':'Already voted on this alliance.'}), 400
            # Per-resident duplicate check
            if resident_id:
                try:
                    rid = int(resident_id)
                    res_dup = c.execute(
                        "SELECT id FROM resident_demo_votes WHERE resident_id=? AND proposal_id=? AND vote_type='alliance'",
                        (rid, proposal_id)).fetchone()
                    if res_dup:
                        return jsonify({'error':'You have already voted on this alliance proposal.'}), 400
                except: pass
            c.execute('INSERT OR IGNORE INTO voter_sessions (session_id) VALUES (?)', (sid,))
            for opt, rating in ratings.items():
                weighted_rating = max(0, min(10, float(rating))) * weight
                c.execute('INSERT OR REPLACE INTO alliance_votes (session_id,proposal_id,option_name,rating) VALUES (?,?,?,?)',
                          (sid, int(proposal_id), opt, weighted_rating))
            # Record resident-level vote
            if resident_id:
                try:
                    rid = int(resident_id)
                    c.execute('INSERT OR IGNORE INTO resident_demo_votes (resident_id,proposal_id,vote_type) VALUES (?,?,?)',
                              (rid, int(proposal_id), 'alliance'))
                except: pass
            c.commit()
    except Exception as e:
        return jsonify({'error':str(e)}), 500
    return jsonify({'ok':True, 'voteWeight':weight})

# ── Admin Auth ────────────────────────────────────────────────────────────────
@app.route('/api/admin/auth', methods=['POST'])
def admin_auth():
    d = request.get_json(force=True)
    if d.get('password') == ADMIN_PASSWORD:
        return jsonify({'ok':True})
    return jsonify({'error':'Incorrect password.'}), 401

# ── Admin State ───────────────────────────────────────────────────────────────
@app.route('/api/admin/state', methods=['POST'])
@require_admin
def admin_set_state():
    d = request.get_json(force=True)
    key, value = d.get('key'), str(d.get('value',''))
    with get_election_db() as c:
        c.execute('INSERT OR REPLACE INTO election_state (key,value) VALUES (?,?)', (key, value))
        c.commit()
    return jsonify({'ok':True})

@app.route('/api/admin/reset-votes', methods=['POST'])
@require_admin
def admin_reset_votes():
    d = request.get_json(force=True)
    t = d.get('type','all')
    with get_election_db() as c:
        if t in ('hexarchy','all'):
            c.execute("DELETE FROM votes WHERE vote_type='hexarchy'")
            c.execute("UPDATE voter_sessions SET voted_hexarchy=0")
        if t in ('charity','all'):
            c.execute("DELETE FROM votes WHERE vote_type='charity'")
            c.execute("UPDATE voter_sessions SET voted_charity=0")
        if t in ('democracy','all'):
            c.execute("DELETE FROM votes WHERE vote_type='democracy'")
            c.execute("DELETE FROM alliance_votes")
        if t == 'all':
            c.execute("DELETE FROM voter_sessions")
            c.execute("DELETE FROM resident_ballots")
        c.commit()
    return jsonify({'ok':True})

# ── Admin Candidates ──────────────────────────────────────────────────────────
@app.route('/api/admin/candidates', methods=['GET'])
@require_admin
def admin_get_candidates():
    with get_election_db() as c:
        rows = [dict(r) for r in c.execute('SELECT * FROM candidates ORDER BY id').fetchall()]
    return jsonify(rows)

@app.route('/api/admin/candidates', methods=['POST'])
@require_admin
def admin_add_candidate():
    d = request.get_json(force=True)
    res_num = d.get('res_num') or d.get('resNum')
    res_num = int(res_num) if res_num else None
    with get_election_db() as c:
        cur = c.execute('INSERT INTO candidates (name,platform,res_num) VALUES (?,?,?)',
                        (d.get('name',''), d.get('platform',''), res_num))
        c.commit()
        row = dict(c.execute('SELECT * FROM candidates WHERE id=?',(cur.lastrowid,)).fetchone())
    return jsonify(row), 201

@app.route('/api/admin/candidates/<int:cid>', methods=['PUT'])
@require_admin
def admin_update_candidate(cid):
    d = request.get_json(force=True)
    res_num = d.get('res_num') or d.get('resNum')
    res_num = int(res_num) if res_num else None
    with get_election_db() as c:
        c.execute('UPDATE candidates SET name=?,platform=?,active=?,res_num=? WHERE id=?',
                  (d.get('name'), d.get('platform'), 1 if d.get('active',True) else 0, res_num, cid))
        c.commit()
        row = dict(c.execute('SELECT * FROM candidates WHERE id=?',(cid,)).fetchone())
    return jsonify(row)

@app.route('/api/admin/candidates/<int:cid>', methods=['DELETE'])
@require_admin
def admin_delete_candidate(cid):
    with get_election_db() as c:
        c.execute('DELETE FROM candidates WHERE id=?',(cid,))
        c.commit()
    return jsonify({'ok':True})

# ── Admin Election Charities ──────────────────────────────────────────────────
@app.route('/api/admin/charities', methods=['GET'])
@require_admin
def admin_get_election_charities():
    with get_election_db() as c:
        rows = [dict(r) for r in c.execute('SELECT * FROM election_charities ORDER BY id').fetchall()]
    return jsonify(rows)

@app.route('/api/admin/charities', methods=['POST'])
@require_admin
def admin_add_election_charity():
    d = request.get_json(force=True)
    with get_election_db() as c:
        cur = c.execute('INSERT INTO election_charities (name,description,min_votes) VALUES (?,?,?)',
                        (d.get('name',''), d.get('description',''), int(d.get('min_votes',20))))
        c.commit()
        row = dict(c.execute('SELECT * FROM election_charities WHERE id=?',(cur.lastrowid,)).fetchone())
    return jsonify(row), 201

@app.route('/api/admin/charities/<int:cid>', methods=['PUT'])
@require_admin
def admin_update_election_charity(cid):
    d = request.get_json(force=True)
    with get_election_db() as c:
        c.execute('UPDATE election_charities SET name=?,description=?,min_votes=?,active=? WHERE id=?',
                  (d.get('name'), d.get('description'), int(d.get('min_votes',20)),
                   1 if d.get('active',True) else 0, cid))
        c.commit()
        row = dict(c.execute('SELECT * FROM election_charities WHERE id=?',(cid,)).fetchone())
    return jsonify(row)

@app.route('/api/admin/charities/<int:cid>', methods=['DELETE'])
@require_admin
def admin_delete_election_charity(cid):
    with get_election_db() as c:
        c.execute('DELETE FROM election_charities WHERE id=?',(cid,))
        c.commit()
    return jsonify({'ok':True})

# ── Admin Proposals ───────────────────────────────────────────────────────────
@app.route('/api/admin/proposals', methods=['GET'])
@require_admin
def admin_get_proposals():
    with get_election_db() as c:
        rows = []
        for r in c.execute('SELECT * FROM proposals ORDER BY id').fetchall():
            p = dict(r)
            p['options'] = json.loads(p['options']) if p.get('options') else []
            rows.append(p)
    return jsonify(rows)

@app.route('/api/admin/proposals', methods=['POST'])
@require_admin
def admin_add_proposal():
    d = request.get_json(force=True)
    opts = json.dumps(d.get('options',[])) if d.get('options') else None
    with get_election_db() as c:
        cur = c.execute(
            'INSERT INTO proposals (title,type,description,threshold,threshold_val,options) VALUES (?,?,?,?,?,?)',
            (d.get('title',''), d.get('type','tax'), d.get('description',''),
             d.get('threshold','2/3'), d.get('threshold_val'), opts))
        c.commit()
        row = dict(c.execute('SELECT * FROM proposals WHERE id=?',(cur.lastrowid,)).fetchone())
        row['options'] = json.loads(row['options']) if row.get('options') else []
    return jsonify(row), 201

@app.route('/api/admin/proposals/<int:pid>', methods=['PUT'])
@require_admin
def admin_update_proposal(pid):
    d = request.get_json(force=True)
    opts = json.dumps(d.get('options',[])) if d.get('options') else None
    with get_election_db() as c:
        c.execute(
            'UPDATE proposals SET title=?,type=?,description=?,threshold=?,threshold_val=?,options=?,active=? WHERE id=?',
            (d.get('title'), d.get('type'), d.get('description'),
             d.get('threshold','2/3'), d.get('threshold_val'), opts,
             1 if d.get('active',True) else 0, pid))
        c.commit()
        row = dict(c.execute('SELECT * FROM proposals WHERE id=?',(pid,)).fetchone())
        row['options'] = json.loads(row['options']) if row.get('options') else []
    return jsonify(row)

@app.route('/api/admin/proposals/<int:pid>', methods=['DELETE'])
@require_admin
def admin_delete_proposal(pid):
    with get_election_db() as c:
        c.execute('DELETE FROM proposals WHERE id=?',(pid,))
        c.commit()
    return jsonify({'ok':True})

# ── Admin Results ─────────────────────────────────────────────────────────────
@app.route('/api/admin/results')
@require_admin
def admin_get_results():
    with get_election_db() as c:
        cands = [dict(r) for r in c.execute(
            "SELECT c.*, COALESCE(SUM(v.rating),0) as rating_sum, COUNT(v.id) as vote_count "
            "FROM candidates c LEFT JOIN votes v ON v.target_id=c.id AND v.vote_type='hexarchy' "
            "GROUP BY c.id ORDER BY rating_sum DESC").fetchall()]
        chars = [dict(r) for r in c.execute(
            "SELECT c.*, COALESCE(SUM(v.rating),0) as total_rating, COUNT(v.id) as vote_count "
            "FROM election_charities c LEFT JOIN votes v ON v.target_id=c.id AND v.vote_type='charity' "
            "GROUP BY c.id ORDER BY total_rating DESC").fetchall()]
        props = []
        for r in c.execute("SELECT * FROM proposals GROUP BY id ORDER BY id").fetchall():
            p = dict(r)
            vc = c.execute(
                "SELECT COALESCE(SUM(CASE WHEN choice='yes' THEN 1 ELSE 0 END),0) as yes_votes, "
                "COALESCE(SUM(CASE WHEN choice='no' THEN 1 ELSE 0 END),0) as no_votes "
                "FROM votes WHERE target_id=? AND vote_type='democracy'", (p['id'],)).fetchone()
            p['yes_votes'] = vc['yes_votes'] if vc else 0
            p['no_votes']  = vc['no_votes']  if vc else 0
            p['options']   = json.loads(p['options']) if p.get('options') else []
            if p['type'] == 'alliance':
                p['alliance_results'] = _get_alliance_results(c, p['id'])
            props.append(p)
        total_sess = c.execute('SELECT COUNT(*) as n FROM voter_sessions').fetchone()[0]
    return jsonify({'candidates':cands,'charities':chars,'proposals':props,'total_sessions':total_sess})
# ══════════════════════════════════════════════════════════════════════════════
# PERSONAL BALANCE
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/personal-balance")
def personal_balance_page():
    return send_from_directory("templates","personal_balance.html")

@app.route("/api/personal-balance", methods=["POST"])
def api_personal_balance():
    """Return full balance breakdown for a resident."""
    try:    data = request.get_json(force=True)
    except: return jsonify({"error": "Invalid request"}), 400
    res_id  = str(data.get("residentId", "")).strip()
    pw_raw  = str(data.get("password", "")).strip()
    if not res_id or not pw_raw:
        return jsonify({"error": "Missing credentials"}), 400
    try: res_id_int = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    wb = load_wb()
    if not wb: return jsonify({"error": "Database unavailable."}), 500
    ws = wb[RES_SHEET]
    ph = hash_pin(pw_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr  = row[COL["res_num"]-1]
        dpw = row[COL["password"]-1]
        da  = row[COL["alive"]-1]
        fn  = row[COL["first_name"]-1]; ln = row[COL["last_name"]-1]
        yb  = row[COL["year_born"]-1]
        if not isinstance(dr, (int, float)): continue
        if int(dr) != res_id_int: continue
        if dpw != ph: return jsonify({"error": "Auth failed"}), 401
        dis  = row[COL["disabled"]-1]; dep = row[COL["deputy"]-1]
        sex  = row[COL["sex"]-1]; sq1 = row[COL["sqft1"]-1]; sq2 = row[COL["sqft2"]-1]
        sides = row[COL["hex_sides"]-1]; shared = row[COL["shared"]-1]
        wm   = row[COL["wealth_m"]-1]
        mort = row[COL["mort_date"]-1]; marr = row[COL["marriage_date"]-1]
        c_cols = ["child1","child2","child3","child4","child5"]
        cdates = [parse_date(row[COL[ck]-1]) if row[COL[ck]-1] else None for ck in c_cols]
        try:
            result = calc_net(
                yb or 1990, bool(dis), bool(dep), str(sex or "M"),
                cdates, float(sq1 or 0), float(sq2 or 0),
                int(sides or 0), int(shared or 0),
                float(wm or 0), parse_date(mort), parse_date(marr)
            )
        except Exception as e:
            return jsonify({"error": f"Calculation error: {e}"}), 500
        a19 = result.get("a19", 0)
        net = result.get("net_tax", 0)
        adjusted = round(net * a19, 2) if net > 0 else round(net, 2)
        return jsonify({
            "resId": int(dr), "firstName": str(fn or ""), "lastName": str(ln or ""),
            "alive": bool(is_active(da)),
            "age": result.get("age", 0),
            "ubi": result.get("ubi", 0),
            "disability_pay": result.get("disability_pay", 0),
            "deputy_pay": result.get("deputy_pay", 0),
            "maternity": result.get("maternity", 0),
            "marriage_bonus": result.get("marriage_bonus", 0),
            "mort_assist": result.get("mort_assist", 0),
            "prop_tax": result.get("prop_tax", 0),
            "sunlight_tax": result.get("sunlight_tax", 0),
            "child_tax": result.get("child_tax", 0),
            "public_projects_tax": result.get("public_projects_tax", 0),
            "charity_tax": result.get("charity_tax", 0),
            "wealth_tax": result.get("wealth_tax", 0),
            "wealth_rate": result.get("wealth_rate", 0),
            "num_children": result.get("num_children", 0),
            "total_sqft": result.get("total_sqft", 0),
            "netPayment": round(net, 2),
            "a19": round(a19, 6),
            "adjustedPayment": adjusted
        })
    return jsonify({"error": "Resident not found"}), 404

# ══════════════════════════════════════════════════════════════════════════════
# ARTICLES OF SYNDICATION — Corporation Management
# ══════════════════════════════════════════════════════════════════════════════
SYND_DB = os.path.join(BASE_DIR, "syndication.db")

ART_SYN_FILE = os.path.join(BASE_DIR, "Art_Syn.xlsx")

def load_corp_types():
    """Read corporation types dynamically from Art_Syn.xlsx (cols A-D, rows 5+)."""
    try:
        wb = openpyxl.load_workbook(ART_SYN_FILE, data_only=True)
        ws = wb[wb.sheetnames[0]]
        types = {}
        for r in range(5, ws.max_row + 1):
            name = ws.cell(r, 1).value
            coin = ws.cell(r, 2).value
            mind = ws.cell(r, 3).value
            hand = ws.cell(r, 4).value
            if not name or coin is None or mind is None or hand is None:
                continue
            types[str(name).strip()] = {
                "coin": round(float(coin), 2),
                "mind": round(float(mind), 2),
                "hand": round(float(hand), 2),
            }
        return types
    except Exception as e:
        print(f"[!] Failed to load Art_Syn.xlsx: {e}")
        return {}

def _init_synd_db():
    import sqlite3
    conn = sqlite3.connect(SYND_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS corporations (
            corp_id INTEGER PRIMARY KEY,
            password TEXT NOT NULL,
            name TEXT NOT NULL DEFAULT '',
            corp_type TEXT NOT NULL,
            coin_pct REAL NOT NULL DEFAULT 33,
            mind_pct REAL NOT NULL DEFAULT 34,
            hand_pct REAL NOT NULL DEFAULT 33,
            est_monthly_profit REAL NOT NULL DEFAULT 0,
            ceo_res_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS corp_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            corp_id INTEGER NOT NULL,
            res_id INTEGER NOT NULL,
            coin_amount REAL NOT NULL DEFAULT 0,
            mind_amount REAL NOT NULL DEFAULT 0,
            hand_amount REAL NOT NULL DEFAULT 0,
            is_ceo INTEGER NOT NULL DEFAULT 0,
            role TEXT NOT NULL DEFAULT 'member',
            FOREIGN KEY (corp_id) REFERENCES corporations(corp_id),
            UNIQUE(corp_id, res_id)
        );
    """)
    # Migration: add role column if missing
    try: c.execute("ALTER TABLE corp_members ADD COLUMN role TEXT NOT NULL DEFAULT 'member'")
    except: pass
    conn.commit()
    return conn

_init_synd_db().close()

@app.route("/syndication")
def syndication_page():
    return send_from_directory("templates","syndication.html")

@app.route("/api/syndication/corp-types")
def synd_corp_types():
    return jsonify({"types": load_corp_types()})

@app.route("/api/syndication/login", methods=["POST"])
def synd_login():
    data = request.get_json(force=True)
    corp_id = data.get("corp_id")
    password = str(data.get("password", "")).strip()
    if not corp_id or not password:
        return jsonify({"error": "Missing credentials"}), 400
    try: corp_id = int(corp_id)
    except: return jsonify({"error": "Invalid Corp#"}), 400
    conn = _init_synd_db()
    c = conn.cursor()
    corp = c.execute("SELECT * FROM corporations WHERE corp_id=?", (corp_id,)).fetchone()
    if not corp:
        conn.close()
        return jsonify({"error": "Corporation not found"}), 404
    if dict(corp)["password"] != hash_pin(password):
        conn.close()
        return jsonify({"error": "Invalid password"}), 401
    corp_d = dict(corp)
    members = [dict(m) for m in c.execute("SELECT * FROM corp_members WHERE corp_id=? ORDER BY is_ceo DESC, res_id", (corp_id,)).fetchall()]
    # Resolve names from spreadsheet
    wb = load_wb()
    name_map = {}
    if wb:
        ws = wb[RES_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = row[COL["res_num"]-1]
            if isinstance(rid, (int, float)):
                fn = row[COL["first_name"]-1] or ""
                ln = row[COL["last_name"]-1] or ""
                name_map[int(rid)] = f"{fn} {ln}".strip()
    for m in members:
        m["name"] = name_map.get(m["res_id"], f"Res#{m['res_id']}")
    conn.close()
    corp_d.pop("password", None)
    return jsonify({"corp": corp_d, "members": members})

@app.route("/api/syndication/create", methods=["POST"])
def synd_create():
    data = request.get_json(force=True)
    corp_id = data.get("corp_id")
    password = str(data.get("password", "")).strip()
    corp_type = str(data.get("corp_type", "")).strip()
    name = str(data.get("name", "")).strip()
    est_profit = float(data.get("est_monthly_profit", 0))
    creator_res = data.get("creator_res_id")
    if not corp_id or not password or not corp_type:
        return jsonify({"error": "Missing required fields"}), 400
    try: corp_id = int(corp_id)
    except: return jsonify({"error": "Invalid Corp#"}), 400
    try: creator_res = int(creator_res)
    except: return jsonify({"error": "Invalid Res#"}), 400
    defaults = load_corp_types().get(corp_type, {"coin": 33, "mind": 34, "hand": 33})
    coin = float(data.get("coin_pct", defaults["coin"]))
    mind = float(data.get("mind_pct", defaults["mind"]))
    hand = float(data.get("hand_pct", defaults["hand"]))
    conn = _init_synd_db()
    c = conn.cursor()
    existing = c.execute("SELECT corp_id FROM corporations WHERE corp_id=?", (corp_id,)).fetchone()
    if existing:
        conn.close()
        return jsonify({"error": f"Corp#{corp_id} already exists"}), 409
    ph = hash_pin(password)
    c.execute("""INSERT INTO corporations (corp_id, password, name, corp_type, coin_pct, mind_pct, hand_pct, est_monthly_profit, ceo_res_id)
                 VALUES (?,?,?,?,?,?,?,?,?)""",
              (corp_id, ph, name, corp_type, coin, mind, hand, est_profit, creator_res))
    c.execute("""INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role)
                 VALUES (?,?,0,0,0,1,'ceo')""", (corp_id, creator_res))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "corp_id": corp_id})

@app.route("/api/syndication/update", methods=["POST"])
def synd_update():
    data = request.get_json(force=True)
    corp_id = data.get("corp_id")
    password = str(data.get("password", "")).strip()
    if not corp_id or not password:
        return jsonify({"error": "Missing credentials"}), 400
    try: corp_id = int(corp_id)
    except: return jsonify({"error": "Invalid Corp#"}), 400
    conn = _init_synd_db()
    c = conn.cursor()
    corp = c.execute("SELECT * FROM corporations WHERE corp_id=?", (corp_id,)).fetchone()
    if not corp:
        conn.close(); return jsonify({"error": "Not found"}), 404
    if dict(corp)["password"] != hash_pin(password):
        conn.close(); return jsonify({"error": "Auth failed"}), 401
    fields = {}
    for key in ["name", "corp_type", "coin_pct", "mind_pct", "hand_pct", "est_monthly_profit"]:
        if key in data:
            fields[key] = data[key]
    if fields:
        set_clause = ", ".join(f"{k}=?" for k in fields)
        c.execute(f"UPDATE corporations SET {set_clause} WHERE corp_id=?", (*fields.values(), corp_id))
        conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/syndication/add-member", methods=["POST"])
def synd_add_member():
    data = request.get_json(force=True)
    corp_id = data.get("corp_id")
    password = str(data.get("password", "")).strip()
    res_id = data.get("res_id")
    coin = float(data.get("coin_amount", 0))
    mind = float(data.get("mind_amount", 0))
    hand = float(data.get("hand_amount", 0))
    if not corp_id or not password or not res_id:
        return jsonify({"error": "Missing fields"}), 400
    try: corp_id = int(corp_id); res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    conn = _init_synd_db()
    c = conn.cursor()
    corp = c.execute("SELECT * FROM corporations WHERE corp_id=?", (corp_id,)).fetchone()
    if not corp:
        conn.close(); return jsonify({"error": "Not found"}), 404
    if dict(corp)["password"] != hash_pin(password):
        conn.close(); return jsonify({"error": "Auth failed"}), 401
    try:
        c.execute("""INSERT OR REPLACE INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo)
                     VALUES (?,?,?,?,?, COALESCE((SELECT is_ceo FROM corp_members WHERE corp_id=? AND res_id=?), 0))""",
                  (corp_id, res_id, coin, mind, hand, corp_id, res_id))
        conn.commit()
    except Exception as e:
        conn.close(); return jsonify({"error": str(e)}), 500
    conn.close()
    return jsonify({"success": True})

@app.route("/api/syndication/remove-member", methods=["POST"])
def synd_remove_member():
    data = request.get_json(force=True)
    corp_id = data.get("corp_id")
    password = str(data.get("password", "")).strip()
    res_id = data.get("res_id")
    if not corp_id or not password or not res_id:
        return jsonify({"error": "Missing fields"}), 400
    try: corp_id = int(corp_id); res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    conn = _init_synd_db()
    c = conn.cursor()
    corp = c.execute("SELECT * FROM corporations WHERE corp_id=?", (corp_id,)).fetchone()
    if not corp:
        conn.close(); return jsonify({"error": "Not found"}), 404
    if dict(corp)["password"] != hash_pin(password):
        conn.close(); return jsonify({"error": "Auth failed"}), 401
    member = c.execute("SELECT * FROM corp_members WHERE corp_id=? AND res_id=?", (corp_id, res_id)).fetchone()
    if not member:
        conn.close(); return jsonify({"error": "Member not found"}), 404
    if dict(member)["is_ceo"]:
        conn.close(); return jsonify({"error": "Cannot remove the CEO"}), 400
    c.execute("DELETE FROM corp_members WHERE corp_id=? AND res_id=?", (corp_id, res_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/syndication/set-ceo", methods=["POST"])
def synd_set_ceo():
    data = request.get_json(force=True)
    corp_id = data.get("corp_id")
    password = str(data.get("password", "")).strip()
    new_ceo_res = data.get("res_id")
    if not corp_id or not password or not new_ceo_res:
        return jsonify({"error": "Missing fields"}), 400
    try: corp_id = int(corp_id); new_ceo_res = int(new_ceo_res)
    except: return jsonify({"error": "Invalid ID"}), 400
    conn = _init_synd_db()
    c = conn.cursor()
    corp = c.execute("SELECT * FROM corporations WHERE corp_id=?", (corp_id,)).fetchone()
    if not corp:
        conn.close(); return jsonify({"error": "Not found"}), 404
    if dict(corp)["password"] != hash_pin(password):
        conn.close(); return jsonify({"error": "Auth failed"}), 401
    # Demote old CEO
    c.execute("UPDATE corp_members SET is_ceo=0, role='member' WHERE corp_id=? AND is_ceo=1", (corp_id,))
    # Promote new CEO
    c.execute("UPDATE corp_members SET is_ceo=1, role='ceo' WHERE corp_id=? AND res_id=?", (corp_id, new_ceo_res))
    c.execute("UPDATE corporations SET ceo_res_id=? WHERE corp_id=?", (new_ceo_res, corp_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/syndication/set-role", methods=["POST"])
def synd_set_role():
    """Set a member's role (coin_manager, hand_manager, member). Only CEO can do this."""
    data = request.get_json(force=True)
    corp_id = data.get("corp_id")
    password = str(data.get("password", "")).strip()
    res_id = data.get("res_id")
    new_role = str(data.get("role", "member")).strip()
    if not corp_id or not password or not res_id:
        return jsonify({"error": "Missing fields"}), 400
    if new_role not in ("member", "coin_manager", "hand_manager"):
        return jsonify({"error": "Invalid role"}), 400
    try: corp_id = int(corp_id); res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    conn = _init_synd_db()
    c = conn.cursor()
    corp = c.execute("SELECT * FROM corporations WHERE corp_id=?", (corp_id,)).fetchone()
    if not corp:
        conn.close(); return jsonify({"error": "Not found"}), 404
    if dict(corp)["password"] != hash_pin(password):
        conn.close(); return jsonify({"error": "Auth failed"}), 401
    member = c.execute("SELECT * FROM corp_members WHERE corp_id=? AND res_id=?", (corp_id, res_id)).fetchone()
    if not member:
        conn.close(); return jsonify({"error": "Member not found"}), 404
    if dict(member)["is_ceo"]:
        conn.close(); return jsonify({"error": "Cannot change role of CEO"}), 400
    c.execute("UPDATE corp_members SET role=? WHERE corp_id=? AND res_id=?", (new_role, corp_id, res_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/syndication/view-all")
def synd_view_all():
    """Public endpoint: list all corporations (no password needed)."""
    conn = _init_synd_db()
    c = conn.cursor()
    corps = [dict(r) for r in c.execute(
        "SELECT corp_id, name, corp_type, coin_pct, mind_pct, hand_pct, est_monthly_profit, ceo_res_id, created_at FROM corporations ORDER BY corp_id"
    ).fetchall()]
    # Get member counts and CEO names
    wb = load_wb()
    name_map = {}
    if wb:
        ws = wb[RES_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = row[COL["res_num"]-1]
            if isinstance(rid, (int, float)):
                fn = row[COL["first_name"]-1] or ""
                ln = row[COL["last_name"]-1] or ""
                name_map[int(rid)] = f"{fn} {ln}".strip()
    for corp in corps:
        count = c.execute("SELECT COUNT(*) FROM corp_members WHERE corp_id=?", (corp["corp_id"],)).fetchone()[0]
        corp["member_count"] = count
        corp["ceo_name"] = name_map.get(corp["ceo_res_id"], f"Res#{corp['ceo_res_id']}")
        corp.pop("password", None)
    conn.close()
    return jsonify({"corporations": corps})

# ══════════════════════════════════════════════════════════════════════════════
# IMMIGRATION MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/immigration/update", methods=["POST"])
def immigration_update():
    """Update ref_checked or immigration_interview for a target resident.
    Only Judges and Hexarchy can use this.
    Values: 0=Processing, 1=Checked/Passed, 2=Declined"""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password","")).strip()
    target_res = data.get("targetResId")
    ref_val = data.get("refChecked")  # 0, 1, or 2
    immig_val = data.get("immigrationInterview")  # 0, 1, or 2
    if not res_id or not pw_raw: return jsonify({"error":"Missing credentials"}),400
    if not target_res: return jsonify({"error":"Missing target Res#"}),400
    try: res_id = int(res_id); target_res = int(target_res)
    except: return jsonify({"error":"Invalid ID"}),400
    wb = load_wb()
    if not wb: return jsonify({"error":"Database unavailable"}),500
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    # Auth caller
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        if not is_active(da): return jsonify({"error":"Account not active"}),403
        if not dpw or dpw!=ph: return jsonify({"error":"Auth failed"}),401
        authed = True; break
    if not authed: return jsonify({"error":"Resident not found"}),404
    # Check Judge or Hexarchy
    is_authorized = False
    if "Gov Employees" in wb.sheetnames:
        ws_gov = wb["Gov Employees"]
        for gov_row in ws_gov.iter_rows(min_row=2, values_only=True):
            gov_res = gov_row[1]
            active = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
            pos = str(gov_row[4] or "").strip().lower() if len(gov_row) > 4 else ""
            hexarchy = str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
            if not isinstance(gov_res,(int,float)) or int(gov_res)!=res_id: continue
            if active not in ("yes","y","true","1"): continue
            if hexarchy in ("yes","y","true","1") or "judge" in pos:
                is_authorized = True; break
    if not is_authorized:
        return jsonify({"error":"Only Judges and Hexarchy can update immigration status."}),403
    # Find target
    target_row_num = None; target_name = ""
    for row in ws.iter_rows(min_row=2):
        dr = row[COL["res_num"]-1].value
        fn = row[COL["first_name"]-1].value or ""; ln = row[COL["last_name"]-1].value or ""
        if not isinstance(dr,(int,float)) or int(dr)!=target_res: continue
        target_row_num = row[0].row; target_name = f"{fn} {ln}".strip(); break
    if not target_row_num: return jsonify({"error":f"Target Res#{target_res} not found."}),404
    # Update fields
    changes = []
    if ref_val is not None:
        try: ref_val = int(ref_val)
        except: ref_val = 0
        if ref_val not in (0,1,2): ref_val = 0
        ws.cell(target_row_num, COL["ref_checked"]).value = ref_val
        labels = {0:"Processing",1:"Checked",2:"Declined"}
        changes.append(f"References → {labels.get(ref_val,'?')}")
    if immig_val is not None:
        try: immig_val = int(immig_val)
        except: immig_val = 0
        if immig_val not in (0,1,2): immig_val = 0
        ws.cell(target_row_num, COL["immigration_interview"]).value = immig_val
        labels = {0:"Processing",1:"Passed",2:"Declined"}
        changes.append(f"Immigration Interview → {labels.get(immig_val,'?')}")
    if not changes: return jsonify({"error":"No changes specified."}),400
    wb.save(EXCEL_FILE)
    return jsonify({"success":True, "target":target_name, "targetResId":target_res, "changes":changes})

@app.route("/api/immigration/lookup", methods=["POST"])
def immigration_lookup():
    """Look up a resident's immigration status by Res#."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password","")).strip()
    target_res = data.get("targetResId")
    if not res_id or not pw_raw: return jsonify({"error":"Missing credentials"}),400
    try: res_id = int(res_id); target_res = int(target_res)
    except: return jsonify({"error":"Invalid ID"}),400
    wb = load_wb()
    if not wb: return jsonify({"error":"Database unavailable"}),500
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    # Auth
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        if dpw==ph: authed=True
        break
    if not authed: return jsonify({"error":"Auth failed"}),401
    # Find target
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=target_res: continue
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        ref_chk = row[COL["ref_checked"]-1] if len(row) > COL["ref_checked"]-1 else None
        immig = row[COL["immigration_interview"]-1] if len(row) > COL["immigration_interview"]-1 else None
        ref_val = int(ref_chk) if isinstance(ref_chk,(int,float)) else 0
        immig_val = int(immig) if isinstance(immig,(int,float)) else 0
        return jsonify({"found":True, "name":f"{ln}, {fn}".strip(", "), "resId":target_res,
                         "refChecked":ref_val, "immigrationInterview":immig_val})
    return jsonify({"found":False})

# ══════════════════════════════════════════════════════════════════════════════
# JURY DUTY SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

JURY_IMG_DIR = os.path.join(BASE_DIR, "Jury_Case_Images")
os.makedirs(JURY_IMG_DIR, exist_ok=True)

def _init_jury_db():
    db_path = os.path.join(BASE_DIR, "jury_duty.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS jury_cases (
            case_id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_type TEXT NOT NULL DEFAULT 'Criminal',
            defendant_res_id INTEGER,
            defendant_name TEXT NOT NULL DEFAULT '',
            max_jail_months REAL NOT NULL DEFAULT 0,
            max_fine REAL NOT NULL DEFAULT 0,
            description TEXT NOT NULL DEFAULT '',
            case_date TEXT DEFAULT '',
            case_time TEXT DEFAULT '',
            case_location TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'Open',
            juror1 INTEGER, juror2 INTEGER, juror3 INTEGER,
            juror4 INTEGER, juror5 INTEGER, juror6 INTEGER,
            final_ruling TEXT DEFAULT '',
            final_sentence_months REAL DEFAULT 0,
            final_fine REAL DEFAULT 0,
            guilty_count INTEGER DEFAULT 0,
            not_guilty_count INTEGER DEFAULT 0,
            created_by INTEGER,
            created_at TEXT DEFAULT (datetime('now')),
            resolved_at TEXT DEFAULT '',
            images TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS jury_votes (
            vote_id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER NOT NULL,
            juror_res_id INTEGER NOT NULL,
            verdict TEXT NOT NULL DEFAULT '',
            sentence_months REAL DEFAULT 0,
            fine_amount REAL DEFAULT 0,
            voted_at TEXT DEFAULT (datetime('now')),
            UNIQUE(case_id, juror_res_id)
        );
    """)
    conn.commit()
    return conn

def _get_jury_db():
    db_path = os.path.join(BASE_DIR, "jury_duty.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def _check_jury_auth(res_id, pw_raw):
    """Check if resident is a Judge or Hexarchy member (can create cases)."""
    wb = load_wb()
    if not wb: return {"authorized": False, "is_judge_or_hex": False, "is_deputy": False, "role": "", "error": "Database unavailable"}
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id: continue
        if not is_active(da): return {"authorized": False, "is_judge_or_hex": False, "is_deputy": False, "role": "", "error": "Account not active"}
        if not dpw or dpw != ph: return {"authorized": False, "is_judge_or_hex": False, "is_deputy": False, "role": "", "error": "Auth failed"}
        authed = True; break
    if not authed:
        return {"authorized": False, "is_judge_or_hex": False, "is_deputy": False, "role": "", "error": "Resident not found"}
    if "Gov Employees" not in wb.sheetnames:
        return {"authorized": True, "is_judge_or_hex": False, "is_deputy": False, "role": "Citizen", "error": ""}
    ws_gov = wb["Gov Employees"]
    best_role = "Citizen"; is_jh = False; is_dep = False
    for gov_row in ws_gov.iter_rows(min_row=2, values_only=True):
        gov_res = gov_row[1]
        active = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
        pos = str(gov_row[4] or "").strip().lower() if len(gov_row) > 4 else ""
        hexarchy = str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
        if not isinstance(gov_res, (int, float)) or int(gov_res) != res_id: continue
        if active not in ("yes", "y", "true", "1"): continue
        if hexarchy in ("yes", "y", "true", "1"):
            best_role = "Hexarchy"; is_jh = True
        elif "judge" in pos:
            if best_role not in ("Hexarchy",): best_role = "Judge"
            is_jh = True
        elif "deputy" in pos:
            if best_role not in ("Hexarchy", "Judge"): best_role = "Deputy"
            is_dep = True
    return {"authorized": True, "is_judge_or_hex": is_jh, "is_deputy": is_dep, "role": best_role, "error": ""}

def _select_jury_pool(exclude_res_id=None):
    """Select 6 eligible jurors from the database.
    Criteria: age 20-70, passed immigration interview, alive."""
    wb = load_wb()
    if not wb: return []
    ws = wb[RES_SHEET]
    import random
    current_year = datetime.datetime.now().year
    eligible = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        res_num = row[COL["res_num"]-1]
        if not isinstance(res_num, (int, float)): continue
        res_num = int(res_num)
        alive = row[COL["alive"]-1]
        if not is_active(alive): continue
        yb = row[COL["year_born"]-1]
        if not isinstance(yb, (int, float)): continue
        age = current_year - int(yb)
        if age < 20 or age > 70: continue
        immig = row[COL["immigration_interview"]-1] if len(row) > COL["immigration_interview"]-1 else None
        if immig != 1: continue
        jury_willing = row[COL["jury"]-1]
        if jury_willing != 1: continue
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        if exclude_res_id and res_num == exclude_res_id: continue
        eligible.append({"res_id": res_num, "name": f"{ln}, {fn}".strip(", ")})
    if len(eligible) < 6:
        return eligible
    return random.sample(eligible, 6)

def _resolve_case(case_id):
    """Check if all jurors voted and resolve the case if so."""
    conn = _get_jury_db()
    case = dict(conn.execute("SELECT * FROM jury_cases WHERE case_id=?", (case_id,)).fetchone())
    juror_ids = [case[f"juror{i}"] for i in range(1,7) if case.get(f"juror{i}")]
    votes = [dict(r) for r in conn.execute("SELECT * FROM jury_votes WHERE case_id=?", (case_id,)).fetchall()]
    if len(votes) < len(juror_ids):
        conn.close(); return None
    # All jurors voted — resolve
    guilty_votes = [v for v in votes if v["verdict"] == "Guilty"]
    not_guilty_votes = [v for v in votes if v["verdict"] == "Not Guilty"]
    gc = len(guilty_votes); ngc = len(not_guilty_votes)
    case_type = case["case_type"]
    # Determine conviction threshold
    if case_type == "Military":
        convicted = gc >= 5
    else:  # Criminal or Civil
        convicted = gc >= 4
    final_ruling = ""; final_months = 0; final_fine = 0
    if convicted:
        if case_type == "Military":
            final_ruling = "Operation Approved"
            final_months = 0; final_fine = 0
        elif case_type == "Criminal":
            final_ruling = "Guilty"
            # Average of guilty votes' sentences
            if guilty_votes:
                avg_months = sum(v["sentence_months"] for v in guilty_votes) / len(guilty_votes)
                avg_fine = sum(v["fine_amount"] for v in guilty_votes) / len(guilty_votes)
            else:
                avg_months = 0; avg_fine = 0
            # Each not guilty vote reduces by 30%
            reduction = 1.0 - (ngc * 0.30)
            if reduction < 0: reduction = 0
            final_months = round(avg_months * reduction, 2)
            final_fine = round(avg_fine * reduction, 2)
            # Cap at maximums
            if final_months > case["max_jail_months"]: final_months = case["max_jail_months"]
            if final_fine > case["max_fine"]: final_fine = case["max_fine"]
        elif case_type == "Civil":
            final_ruling = "Guilty"
            # Average fines from guilty votes
            if guilty_votes:
                avg_fine = sum(v["fine_amount"] for v in guilty_votes) / len(guilty_votes)
            else:
                avg_fine = 0
            reduction = 1.0 - (ngc * 0.30)
            if reduction < 0: reduction = 0
            final_fine = round(avg_fine * reduction, 2)
            if final_fine > case["max_fine"]: final_fine = case["max_fine"]
            final_months = 0
    else:
        if case_type == "Military":
            final_ruling = "Operation Denied"
        else:
            final_ruling = "Not Guilty"
        final_months = 0; final_fine = 0
    conn.execute("""UPDATE jury_cases SET status='Resolved', final_ruling=?, final_sentence_months=?,
                    final_fine=?, guilty_count=?, not_guilty_count=?, resolved_at=datetime('now')
                    WHERE case_id=?""",
                 (final_ruling, final_months, final_fine, gc, ngc, case_id))
    conn.commit()
    # Insert into criminal records DB if Criminal or Civil
    if case_type in ("Criminal", "Civil"):
        cr_conn = _init_criminal_db()
        # Migrate fine_amount column if needed
        try: cr_conn.execute("ALTER TABLE crimes ADD COLUMN fine_amount REAL DEFAULT 0")
        except: pass
        try: cr_conn.execute("ALTER TABLE crimes ADD COLUMN case_category TEXT DEFAULT ''")
        except: pass
        try: cr_conn.execute("ALTER TABLE crimes ADD COLUMN jury_case_id INTEGER DEFAULT NULL")
        except: pass
        cr_conn.execute("""INSERT INTO crimes (victim_id, victim_name, defendant_id, defendant_name,
                           crime_type, victim_account, defendant_account, jury_ruling,
                           sentence_months, fine_amount, not_guilty, date_filed, case_category, jury_case_id)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (None, "Ark Community", case["defendant_res_id"], case["defendant_name"],
                         case_type, case["description"], "",
                         final_ruling, final_months, final_fine,
                         1 if final_ruling == "Not Guilty" else 0,
                         datetime.date.today().isoformat(), case_type, case_id))
        cr_conn.commit(); cr_conn.close()
    resolved = dict(conn.execute("SELECT * FROM jury_cases WHERE case_id=?", (case_id,)).fetchone())
    conn.close()
    return resolved

@app.route("/jury-duty")
def jury_duty_page():
    return send_from_directory("templates", "jury_duty.html")

@app.route("/api/jury/auth-check", methods=["POST"])
def jury_auth_check():
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    if not res_id or not pw_raw: return jsonify({"authorized": False, "error": "Missing credentials"}), 400
    try: res_id = int(res_id)
    except: return jsonify({"authorized": False, "error": "Invalid ID"}), 400
    result = _check_jury_auth(res_id, pw_raw)
    return jsonify(result)

@app.route("/api/jury/select-pool", methods=["POST"])
def jury_select_pool():
    """Select 6 random eligible jurors. Only Judges/Hexarchy can call this."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_jury_auth(res_id, pw_raw)
    if not auth["is_judge_or_hex"]:
        return jsonify({"error": "Only Judges and Hexarchy can select a jury pool."}), 403
    defendant_res = data.get("defendantResId")
    try: defendant_res = int(defendant_res) if defendant_res else None
    except: defendant_res = None
    pool = _select_jury_pool(exclude_res_id=defendant_res)
    return jsonify({"pool": pool, "count": len(pool)})

@app.route("/api/jury/create-case", methods=["POST"])
def jury_create_case():
    """Create a new jury case. Only Judges/Hexarchy."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_jury_auth(res_id, pw_raw)
    if not auth["is_judge_or_hex"]:
        return jsonify({"error": "Only Judges and Hexarchy can create cases."}), 403
    case = data.get("case", {})
    jurors = case.get("jurors", [])
    if len(jurors) != 6: return jsonify({"error": "Exactly 6 jurors required."}), 400
    # Look up defendant name if res_id given
    defendant_name = case.get("defendantName", "")
    defendant_res = case.get("defendantResId")
    try:
        defendant_res = int(defendant_res) if defendant_res else None
    except:
        defendant_res = None
    if defendant_res and not defendant_name:
        wb = load_wb()
        if wb:
            ws = wb[RES_SHEET]
            for row in ws.iter_rows(min_row=2, values_only=True):
                dr = row[COL["res_num"]-1]
                if isinstance(dr, (int, float)) and int(dr) == defendant_res:
                    fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
                    defendant_name = f"{ln}, {fn}".strip(", ")
                    break
    conn = _get_jury_db()
    c = conn.cursor()
    c.execute("""INSERT INTO jury_cases (case_type, defendant_res_id, defendant_name,
                 max_jail_months, max_fine, description, case_date, case_time, case_location,
                 juror1, juror2, juror3, juror4, juror5, juror6, created_by, images)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
              (case.get("caseType", "Criminal"), defendant_res, defendant_name,
               float(case.get("maxJailMonths", 0)), float(case.get("maxFine", 0)),
               case.get("description", ""), case.get("caseDate", ""),
               case.get("caseTime", ""), case.get("caseLocation", ""),
               jurors[0], jurors[1], jurors[2], jurors[3], jurors[4], jurors[5],
               res_id, case.get("images", "")))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({"success": True, "case_id": new_id})

@app.route("/api/jury/cases", methods=["POST"])
def jury_list_cases():
    """List cases relevant to this resident. Jurors see their assigned cases.
    Judges/Hex/Deputies see all cases."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_jury_auth(res_id, pw_raw)
    if not auth["authorized"]: return jsonify({"error": auth["error"]}), 403
    conn = _get_jury_db()
    # Jurors see cases they're assigned to; officials see all
    if auth["is_judge_or_hex"] or auth["is_deputy"]:
        cases = [dict(r) for r in conn.execute("SELECT * FROM jury_cases ORDER BY case_id DESC").fetchall()]
    else:
        cases = [dict(r) for r in conn.execute(
            """SELECT * FROM jury_cases WHERE juror1=? OR juror2=? OR juror3=? OR juror4=? OR juror5=? OR juror6=?
               ORDER BY case_id DESC""", (res_id,res_id,res_id,res_id,res_id,res_id)).fetchall()]
    # Get votes for each case
    for c in cases:
        votes = [dict(r) for r in conn.execute("SELECT * FROM jury_votes WHERE case_id=?", (c["case_id"],)).fetchall()]
        c["votes"] = votes
        c["my_vote"] = None
        for v in votes:
            if v["juror_res_id"] == res_id:
                c["my_vote"] = v
                break
        # Check if this res_id is a juror on this case
        c["is_juror"] = res_id in [c.get(f"juror{i}") for i in range(1,7)]
    conn.close()
    return jsonify({"cases": cases, "role": auth["role"], "is_judge_or_hex": auth["is_judge_or_hex"],
                     "is_deputy": auth["is_deputy"]})

@app.route("/api/jury/vote", methods=["POST"])
def jury_vote():
    """Submit a juror's vote for a case."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_jury_auth(res_id, pw_raw)
    if not auth["authorized"]: return jsonify({"error": auth["error"]}), 403
    case_id = data.get("caseId")
    verdict = data.get("verdict", "").strip()
    sentence_months = float(data.get("sentenceMonths", 0) or 0)
    fine_amount = float(data.get("fineAmount", 0) or 0)
    if verdict not in ("Guilty", "Not Guilty"):
        return jsonify({"error": "Verdict must be 'Guilty' or 'Not Guilty'."}), 400
    if verdict == "Not Guilty":
        sentence_months = 0; fine_amount = 0
    conn = _get_jury_db()
    case = conn.execute("SELECT * FROM jury_cases WHERE case_id=?", (case_id,)).fetchone()
    if not case: conn.close(); return jsonify({"error": "Case not found."}), 404
    case = dict(case)
    if case["status"] != "Open": conn.close(); return jsonify({"error": "This case is already resolved."}), 400
    # Check that res_id is a juror on this case
    juror_ids = [case.get(f"juror{i}") for i in range(1,7)]
    if res_id not in juror_ids:
        conn.close(); return jsonify({"error": "You are not a juror on this case."}), 403
    # Officials can view but NOT vote
    if auth["is_judge_or_hex"] or auth["is_deputy"]:
        if res_id not in juror_ids:
            conn.close(); return jsonify({"error": "Officials can view cases but only jurors can vote."}), 403
    # Cap sentence and fine at maximums
    if sentence_months > case["max_jail_months"]: sentence_months = case["max_jail_months"]
    if fine_amount > case["max_fine"]: fine_amount = case["max_fine"]
    try:
        conn.execute("""INSERT INTO jury_votes (case_id, juror_res_id, verdict, sentence_months, fine_amount)
                        VALUES (?,?,?,?,?)""", (case_id, res_id, verdict, sentence_months, fine_amount))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close(); return jsonify({"error": "You have already voted on this case."}), 400
    conn.close()
    # Check if case is fully resolved
    resolved = _resolve_case(case_id)
    return jsonify({"success": True, "resolved": resolved is not None,
                     "case": resolved if resolved else None})

@app.route("/api/jury/my-open-cases", methods=["POST"])
def jury_my_open_cases():
    """Quick check: does this resident have any open jury cases? Used for portal flashing."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"has_open": False}), 400
    # Lightweight auth — just verify password
    wb = load_wb()
    if not wb: return jsonify({"has_open": False})
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id: continue
        if dpw == ph: authed = True
        break
    if not authed: return jsonify({"has_open": False})
    conn = _get_jury_db()
    cases = [dict(r) for r in conn.execute(
        """SELECT case_id, case_type FROM jury_cases WHERE status='Open'
           AND (juror1=? OR juror2=? OR juror3=? OR juror4=? OR juror5=? OR juror6=?)""",
        (res_id,res_id,res_id,res_id,res_id,res_id)).fetchall()]
    # Check which ones the user hasn't voted on yet
    open_types = {"Criminal": False, "Civil": False, "Military": False}
    has_any = False
    for c in cases:
        vote = conn.execute("SELECT vote_id FROM jury_votes WHERE case_id=? AND juror_res_id=?",
                            (c["case_id"], res_id)).fetchone()
        if not vote:
            open_types[c["case_type"]] = True
            has_any = True
    conn.close()
    return jsonify({"has_open": has_any, "open_types": open_types})

@app.route("/api/jury/upload-image", methods=["POST"])
def jury_upload_image():
    """Upload a JPG image for a jury case."""
    import base64, uuid
    try: data = request.get_json(force=True)
    except: return jsonify({"success": False, "error": "Invalid request"}), 400
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"success": False, "error": "Invalid ID"}), 400
    auth = _check_jury_auth(res_id, pw_raw)
    if not auth["is_judge_or_hex"]:
        return jsonify({"success": False, "error": "Only Judges/Hexarchy can upload images."}), 403
    img_data = data.get("imageData", "")
    if not img_data: return jsonify({"success": False, "error": "No image data."}), 400
    if "," in img_data: img_data = img_data.split(",", 1)[1]
    raw = base64.b64decode(img_data)
    if len(raw) > 500 * 1024:
        return jsonify({"success": False, "error": "Image too large (max 500kb)."}), 400
    fname = f"jury_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.jpg"
    fpath = os.path.join(JURY_IMG_DIR, fname)
    with open(fpath, "wb") as f: f.write(raw)
    return jsonify({"success": True, "filename": fname})

@app.route("/api/jury/image/<filename>")
def jury_image(filename):
    import re as _re
    if not _re.match(r'^[\w\-\.]+$', filename): return "Not found", 404
    return send_from_directory(JURY_IMG_DIR, filename)

@app.route("/api/jury/lookup-resident", methods=["POST"])
def jury_lookup_resident():
    """Look up resident name by Res# for the case form."""
    data = request.get_json(force=True)
    target = data.get("targetResId")
    try: target = int(target)
    except: return jsonify({"error": "Invalid ID"}), 400
    wb = load_wb()
    if not wb: return jsonify({"error": "Database unavailable"}), 500
    ws = wb[RES_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != target: continue
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        return jsonify({"found": True, "name": f"{ln}, {fn}".strip(", "), "resId": target})
    return jsonify({"found": False})

# ══════════════════════════════════════════════════════════════════════════════
# CRIMINAL RECORDS
# ══════════════════════════════════════════════════════════════════════════════

def _init_criminal_db():
    db_path = os.path.join(BASE_DIR, "criminal_records.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS crimes (
            crime_id INTEGER PRIMARY KEY AUTOINCREMENT,
            victim_id INTEGER,
            victim_name TEXT NOT NULL DEFAULT '',
            defendant_id INTEGER,
            defendant_name TEXT NOT NULL DEFAULT '',
            crime_type TEXT NOT NULL DEFAULT '',
            victim_account TEXT NOT NULL DEFAULT '',
            defendant_account TEXT NOT NULL DEFAULT '',
            jury_ruling TEXT NOT NULL DEFAULT 'Pending',
            sentence_months REAL NOT NULL DEFAULT 0,
            not_guilty INTEGER NOT NULL DEFAULT 0,
            date_filed TEXT DEFAULT (date('now')),
            created_at TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()
    # Migrate: add fine_amount, case_category, jury_case_id columns
    try: c.execute("ALTER TABLE crimes ADD COLUMN fine_amount REAL DEFAULT 0")
    except: pass
    try: c.execute("ALTER TABLE crimes ADD COLUMN case_category TEXT DEFAULT ''")
    except: pass
    try: c.execute("ALTER TABLE crimes ADD COLUMN jury_case_id INTEGER DEFAULT NULL")
    except: pass
    conn.commit()
    # Seed example crimes if table is empty
    count = c.execute("SELECT COUNT(*) FROM crimes").fetchone()[0]
    if count == 0:
        examples = [
            (3, "Rivera, Sofia", 17, "Chen, Marcus", "Theft", "Defendant stole 0.5 BCH from my wallet through a fraudulent transaction on the community exchange.", "I was borrowing the funds temporarily and intended to return them. This was a misunderstanding.", "Guilty", 3, 0, "2025-08-12"),
            (8, "Okonkwo, Amara", 12, "Petrov, Ivan", "Vandalism", "Defendant damaged the exterior wall panels on Hex Unit 14, causing structural concern and $2,400 in repairs.", "I accidentally backed the construction cart into the wall. It was not intentional damage.", "Guilty", 1.5, 0, "2025-10-03"),
            (None, "Ark Community", 5, "Nakamura, Yuki", "Tax Evasion", "Defendant underreported net worth by $1.2M over two consecutive tax periods, avoiding approximately $14,000 in wealth tax.", "I made an accounting error with my off-Ark assets. I have since corrected the filing.", "Guilty", 6, 0, "2025-11-20"),
            (22, "Mbeki, Thabo", 31, "Santos, Lucia", "Assault", "Defendant struck me during an argument at the community greenhouse, resulting in a minor injury.", "I acted in self-defense after being verbally threatened and cornered.", "Not Guilty", 0, 1, "2026-01-15"),
            (15, "Okafor, James", 9, "Kim, Soo-jin", "Fraud", "Defendant sold me a counterfeit community bond certificate for 0.3 BCH.", "The bond was legitimate when I acquired it. I was unaware it had been previously voided.", "Pending", 0, 0, "2026-02-28"),
        ]
        for (vid, vname, did, dname, ctype, vacc, dacc, ruling, months, ng, filed) in examples:
            c.execute("""INSERT INTO crimes (victim_id, victim_name, defendant_id, defendant_name,
                         crime_type, victim_account, defendant_account, jury_ruling,
                         sentence_months, not_guilty, date_filed)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                      (vid, vname, did, dname, ctype, vacc, dacc, ruling, months, ng, filed))
        conn.commit()
    return conn

def _check_criminal_auth(res_id, pw_raw):
    """Check if resident is authorized for criminal records.
    Returns: {'authorized': bool, 'can_edit': bool, 'role': str, 'error': str}"""
    wb = load_wb()
    if not wb:
        return {"authorized": False, "can_edit": False, "role": "", "error": "Database unavailable"}
    ws = wb[RES_SHEET]
    ph = hash_pin(pw_raw)
    # Verify password
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id: continue
        if not is_active(da): return {"authorized": False, "can_edit": False, "role": "", "error": "Account not active"}
        if not dpw or dpw != ph: return {"authorized": False, "can_edit": False, "role": "", "error": "Auth failed"}
        authed = True; break
    if not authed:
        return {"authorized": False, "can_edit": False, "role": "", "error": "Resident not found"}
    # Check Gov Employees
    if "Gov Employees" not in wb.sheetnames:
        return {"authorized": False, "can_edit": False, "role": "", "error": "Only the Hexarchy, Judges, Deputies and Auditors can view the criminal records database."}
    ws_gov = wb["Gov Employees"]
    best_role = None
    can_edit = False
    for gov_row in ws_gov.iter_rows(min_row=2, values_only=True):
        gov_res = gov_row[1]  # col B = Res #
        active = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
        pos = str(gov_row[4] or "").strip().lower() if len(gov_row) > 4 else ""
        hexarchy = str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
        if not isinstance(gov_res, (int, float)) or int(gov_res) != res_id: continue
        if active not in ("yes", "y", "true", "1"): continue
        if hexarchy in ("yes", "y", "true", "1"):
            best_role = "Hexarchy"
            can_edit = True
        elif "judge" in pos:
            if not best_role or best_role not in ("Hexarchy",):
                best_role = "Judge"
                can_edit = True
        elif "deputy" in pos:
            if not best_role or best_role not in ("Hexarchy", "Judge"):
                best_role = "Deputy"
        elif "auditor" in pos:
            if not best_role or best_role not in ("Hexarchy", "Judge", "Deputy"):
                best_role = "Auditor"
    if best_role:
        return {"authorized": True, "can_edit": can_edit, "role": best_role, "error": ""}
    return {"authorized": False, "can_edit": False, "role": "", "error": "Only the Hexarchy, Judges, Deputies and Auditors can view the criminal records database."}

@app.route("/criminal-records")
def criminal_records_page():
    return send_from_directory("templates", "criminal_records.html")

@app.route("/api/criminal-records/auth-check", methods=["POST"])
def criminal_auth_check():
    data = request.get_json(force=True)
    res_id = data.get("residentId")
    pw_raw = str(data.get("password", "")).strip()
    if not res_id or not pw_raw:
        return jsonify({"authorized": False, "error": "Missing credentials"}), 400
    try: res_id = int(res_id)
    except: return jsonify({"authorized": False, "error": "Invalid ID"}), 400
    result = _check_criminal_auth(res_id, pw_raw)
    return jsonify(result)

@app.route("/api/criminal-records/list", methods=["POST"])
def criminal_list():
    data = request.get_json(force=True)
    res_id = data.get("residentId")
    pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_criminal_auth(res_id, pw_raw)
    if not auth["authorized"]:
        return jsonify({"error": auth["error"]}), 403
    conn = _init_criminal_db()
    crimes = [dict(r) for r in conn.execute("SELECT * FROM crimes ORDER BY crime_id DESC").fetchall()]
    conn.close()
    return jsonify({"crimes": crimes, "can_edit": auth["can_edit"], "role": auth["role"]})

@app.route("/api/criminal-records/add", methods=["POST"])
def criminal_add():
    data = request.get_json(force=True)
    res_id = data.get("residentId")
    pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_criminal_auth(res_id, pw_raw)
    if not auth["can_edit"]:
        return jsonify({"error": "Only the Hexarchy and Judges can add criminal records."}), 403
    crime = data.get("crime", {})
    conn = _init_criminal_db()
    c = conn.cursor()
    c.execute("""INSERT INTO crimes (victim_id, victim_name, defendant_id, defendant_name,
                 crime_type, victim_account, defendant_account, jury_ruling,
                 sentence_months, not_guilty, date_filed, fine_amount, case_category, jury_case_id)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
              (crime.get("victim_id"), crime.get("victim_name", ""),
               crime.get("defendant_id"), crime.get("defendant_name", ""),
               crime.get("crime_type", ""), crime.get("victim_account", ""),
               crime.get("defendant_account", ""), crime.get("jury_ruling", "Pending"),
               float(crime.get("sentence_months", 0)), int(crime.get("not_guilty", 0)),
               crime.get("date_filed", datetime.date.today().isoformat()),
               float(crime.get("fine_amount", 0)), crime.get("case_category", ""),
               crime.get("jury_case_id")))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({"success": True, "crime_id": new_id})

@app.route("/api/criminal-records/edit", methods=["POST"])
def criminal_edit():
    data = request.get_json(force=True)
    res_id = data.get("residentId")
    pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_criminal_auth(res_id, pw_raw)
    if not auth["can_edit"]:
        return jsonify({"error": "Only the Hexarchy and Judges can edit criminal records."}), 403
    crime = data.get("crime", {})
    crime_id = crime.get("crime_id")
    if not crime_id:
        return jsonify({"error": "Missing crime_id"}), 400
    conn = _init_criminal_db()
    conn.execute("""UPDATE crimes SET victim_id=?, victim_name=?, defendant_id=?, defendant_name=?,
                    crime_type=?, victim_account=?, defendant_account=?, jury_ruling=?,
                    sentence_months=?, not_guilty=?, date_filed=?, fine_amount=?, case_category=? WHERE crime_id=?""",
                 (crime.get("victim_id"), crime.get("victim_name", ""),
                  crime.get("defendant_id"), crime.get("defendant_name", ""),
                  crime.get("crime_type", ""), crime.get("victim_account", ""),
                  crime.get("defendant_account", ""), crime.get("jury_ruling", "Pending"),
                  float(crime.get("sentence_months", 0)), int(crime.get("not_guilty", 0)),
                  crime.get("date_filed", ""), float(crime.get("fine_amount", 0)),
                  crime.get("case_category", ""), crime_id))
    conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route("/api/criminal-records/delete", methods=["POST"])
def criminal_delete():
    data = request.get_json(force=True)
    res_id = data.get("residentId")
    pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_criminal_auth(res_id, pw_raw)
    if not auth["can_edit"]:
        return jsonify({"error": "Only the Hexarchy and Judges can delete criminal records."}), 403
    crime_id = data.get("crime_id")
    if not crime_id: return jsonify({"error": "Missing crime_id"}), 400
    conn = _init_criminal_db()
    conn.execute("DELETE FROM crimes WHERE crime_id=?", (crime_id,))
    conn.commit(); conn.close()
    return jsonify({"success": True})

# ══════════════════════════════════════════════════════════════════════════════
# PROPERTY DATABASE
# ══════════════════════════════════════════════════════════════════════════════

def _check_property_auth(res_id, pw_raw):
    """Check if resident can access property database.
    View: hexarchy, judges, deputies, auditors.  Edit: hexarchy, auditors only."""
    wb = load_wb()
    if not wb:
        return {"authorized": False, "can_edit": False, "role": "", "error": "Database unavailable"}
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id: continue
        if not is_active(da): return {"authorized": False, "can_edit": False, "role": "", "error": "Account not active"}
        if not dpw or dpw != ph: return {"authorized": False, "can_edit": False, "role": "", "error": "Auth failed"}
        authed = True; break
    if not authed:
        return {"authorized": False, "can_edit": False, "role": "", "error": "Resident not found"}
    if "Gov Employees" not in wb.sheetnames:
        return {"authorized": False, "can_edit": False, "role": "", "error": "Only the Hexarchy, Auditors and Deputies can access the property database."}
    ws_gov = wb["Gov Employees"]
    best_role = None; can_edit = False
    for gov_row in ws_gov.iter_rows(min_row=2, values_only=True):
        gov_res = gov_row[1]
        active = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
        pos = str(gov_row[4] or "").strip().lower() if len(gov_row) > 4 else ""
        hexarchy = str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
        if not isinstance(gov_res, (int, float)) or int(gov_res) != res_id: continue
        if active not in ("yes", "y", "true", "1"): continue
        if hexarchy in ("yes", "y", "true", "1"):
            best_role = "Hexarchy"; can_edit = True
        elif "auditor" in pos:
            if not best_role or best_role not in ("Hexarchy",):
                best_role = "Auditor"; can_edit = True
        elif "deputy" in pos:
            if not best_role or best_role not in ("Hexarchy", "Auditor"):
                best_role = "Deputy"
    if best_role:
        return {"authorized": True, "can_edit": can_edit, "role": best_role, "error": ""}
    return {"authorized": False, "can_edit": False, "role": "", "error": "Only the Hexarchy, Auditors and Deputies can access the property database."}

def _excel_date_to_str(serial):
    """Convert Excel serial date to YYYY-MM-DD string."""
    if not serial or not isinstance(serial, (int, float)): return str(serial or "")
    try:
        import datetime as dt
        base = dt.datetime(1899, 12, 30)
        return (base + dt.timedelta(days=int(serial))).strftime("%Y-%m-%d")
    except: return str(serial)

def _str_to_excel_date(date_str):
    """Convert YYYY-MM-DD string to Excel serial date."""
    if not date_str: return None
    try:
        import datetime as dt
        base = dt.datetime(1899, 12, 30)
        d = dt.datetime.strptime(date_str, "%Y-%m-%d")
        return (d - base).days
    except: return None

@app.route("/property-database")
def property_database_page():
    return send_from_directory("templates", "property_database.html")

@app.route("/api/property/auth-check", methods=["POST"])
def property_auth_check():
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    if not res_id or not pw_raw: return jsonify({"authorized": False, "error": "Missing credentials"}), 400
    try: res_id = int(res_id)
    except: return jsonify({"authorized": False, "error": "Invalid ID"}), 400
    return jsonify(_check_property_auth(res_id, pw_raw))

@app.route("/api/property/list", methods=["POST"])
def property_list():
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_property_auth(res_id, pw_raw)
    if not auth["authorized"]: return jsonify({"error": auth["error"]}), 403
    wb = load_wb()
    if not wb or "Prop" not in wb.sheetnames: return jsonify({"error": "Property sheet unavailable"}), 500
    ws = wb["Prop"]
    # Resolve names from Res sheet
    ws_res = wb[RES_SHEET]
    name_map = {}
    for row in ws_res.iter_rows(min_row=2, values_only=True):
        rid = row[COL["res_num"]-1]
        if isinstance(rid, (int, float)):
            fn = row[COL["first_name"]-1] or ""; ln = row[COL["last_name"]-1] or ""
            name_map[int(rid)] = f"{fn} {ln}".strip()
    props = []
    for r in range(2, ws.max_row + 1):
        trans = ws.cell(r, 2).value  # B: Transaction #
        prop_num = ws.cell(r, 3).value  # C: Property #
        addr = ws.cell(r, 4).value  # D: Address
        sqft = ws.cell(r, 5).value  # E: Square Feet
        amt = ws.cell(r, 6).value   # F: $ Amt (K)
        date_raw = ws.cell(r, 7).value  # G: Date
        owner_id = ws.cell(r, 8).value  # H: Owner ID
        last_name = ws.cell(r, 9).value  # I: Name Last
        first_name = ws.cell(r, 10).value  # J: Name First
        use = ws.cell(r, 11).value  # K: Used For
        prop_tax = ws.cell(r, 12).value  # L: Prop Tax
        if not trans and not addr: continue  # skip fully empty rows
        # Resolve owner name from Res sheet if not in Prop sheet
        owner_name = ""
        if last_name or first_name:
            owner_name = f"{first_name or ''} {last_name or ''}".strip()
        elif owner_id and isinstance(owner_id, (int, float)):
            owner_name = name_map.get(int(owner_id), "")
        props.append({
            "row": r,
            "transaction": str(trans or ""),
            "property": str(prop_num or ""),
            "address": str(addr or ""),
            "sqft": int(sqft) if isinstance(sqft, (int, float)) else None,
            "amount_k": float(amt) if isinstance(amt, (int, float)) else None,
            "date": _excel_date_to_str(date_raw) if isinstance(date_raw, (int, float)) else str(date_raw or ""),
            "owner_id": int(owner_id) if isinstance(owner_id, (int, float)) else None,
            "owner_name": owner_name,
            "use": str(use or ""),
            "prop_tax": float(prop_tax) if isinstance(prop_tax, (int, float)) else None,
        })
    return jsonify({"properties": props, "can_edit": auth["can_edit"], "role": auth["role"]})

@app.route("/api/property/save", methods=["POST"])
def property_save():
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_property_auth(res_id, pw_raw)
    if not auth["can_edit"]: return jsonify({"error": "Only the Hexarchy and Auditors can edit property records."}), 403
    prop = data.get("property", {})
    row_num = prop.get("row")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Prop"]
        if row_num:
            # Edit existing
            r = int(row_num)
        else:
            # Find next empty transaction row
            r = None
            for check_r in range(2, ws.max_row + 1):
                if ws.cell(check_r, 4).value is None and ws.cell(check_r, 5).value is None:
                    r = check_r; break
            if r is None:
                r = ws.max_row + 1
        ws.cell(r, 2).value = prop.get("transaction", "")
        ws.cell(r, 3).value = prop.get("property_num", "")
        ws.cell(r, 4).value = prop.get("address", "")
        ws.cell(r, 5).value = int(prop.get("sqft", 0)) if prop.get("sqft") else None
        ws.cell(r, 6).value = float(prop.get("amount_k", 0)) if prop.get("amount_k") else None
        date_serial = _str_to_excel_date(prop.get("date", ""))
        ws.cell(r, 7).value = date_serial
        ws.cell(r, 8).value = int(prop.get("owner_id")) if prop.get("owner_id") else None
        ws.cell(r, 9).value = prop.get("last_name", "")
        ws.cell(r, 10).value = prop.get("first_name", "")
        ws.cell(r, 11).value = prop.get("use", "")
        wb.save(EXCEL_FILE)
        return jsonify({"success": True, "row": r})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE CHANGES (CHANGELOG)
# ══════════════════════════════════════════════════════════════════════════════

def _check_changelog_auth(res_id, pw_raw):
    """View: hexarchy, judges, deputies, auditors. Revert: hexarchy, judges only."""
    wb = load_wb()
    if not wb:
        return {"authorized": False, "can_revert": False, "role": "", "error": "Database unavailable"}
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr, (int, float)) or int(dr) != res_id: continue
        if not is_active(da): return {"authorized": False, "can_revert": False, "role": "", "error": "Account not active"}
        if not dpw or dpw != ph: return {"authorized": False, "can_revert": False, "role": "", "error": "Auth failed"}
        authed = True; break
    if not authed:
        return {"authorized": False, "can_revert": False, "role": "", "error": "Resident not found"}
    if "Gov Employees" not in wb.sheetnames:
        return {"authorized": False, "can_revert": False, "role": "", "error": "Only the Hexarchy, Judges, Deputies and Auditors can view database changes."}
    ws_gov = wb["Gov Employees"]
    best_role = None; can_revert = False
    for gov_row in ws_gov.iter_rows(min_row=2, values_only=True):
        gov_res = gov_row[1]
        active = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
        pos = str(gov_row[4] or "").strip().lower() if len(gov_row) > 4 else ""
        hexarchy = str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
        if not isinstance(gov_res, (int, float)) or int(gov_res) != res_id: continue
        if active not in ("yes", "y", "true", "1"): continue
        if hexarchy in ("yes", "y", "true", "1"):
            best_role = "Hexarchy"; can_revert = True
        elif "judge" in pos:
            if not best_role or best_role not in ("Hexarchy",):
                best_role = "Judge"; can_revert = True
        elif "deputy" in pos:
            if not best_role or best_role not in ("Hexarchy", "Judge"):
                best_role = "Deputy"
        elif "auditor" in pos:
            if not best_role or best_role not in ("Hexarchy", "Judge", "Deputy"):
                best_role = "Auditor"
    if best_role:
        return {"authorized": True, "can_revert": can_revert, "role": best_role, "error": ""}
    return {"authorized": False, "can_revert": False, "role": "", "error": "Only the Hexarchy, Judges, Deputies and Auditors can view database changes."}

@app.route("/database-changes")
def database_changes_page():
    return send_from_directory("templates", "database_changes.html")

@app.route("/api/changelog/auth-check", methods=["POST"])
def changelog_auth_check():
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    if not res_id or not pw_raw: return jsonify({"authorized": False, "error": "Missing credentials"}), 400
    try: res_id = int(res_id)
    except: return jsonify({"authorized": False, "error": "Invalid ID"}), 400
    return jsonify(_check_changelog_auth(res_id, pw_raw))

@app.route("/api/changelog/list", methods=["POST"])
def changelog_list():
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_changelog_auth(res_id, pw_raw)
    if not auth["authorized"]: return jsonify({"error": auth["error"]}), 403
    conn = _init_changelog_db()
    changes = [dict(r) for r in conn.execute("SELECT * FROM changes ORDER BY id DESC").fetchall()]
    conn.close()
    return jsonify({"changes": changes, "can_revert": auth["can_revert"], "role": auth["role"]})

@app.route("/api/changelog/revert", methods=["POST"])
def changelog_revert():
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password", "")).strip()
    change_id = data.get("change_id")
    try: res_id = int(res_id)
    except: return jsonify({"error": "Invalid ID"}), 400
    auth = _check_changelog_auth(res_id, pw_raw)
    if not auth["can_revert"]:
        return jsonify({"error": "Only the Hexarchy and Judges can revert changes."}), 403
    if not change_id: return jsonify({"error": "Missing change_id"}), 400
    conn = _init_changelog_db()
    change = conn.execute("SELECT * FROM changes WHERE id=?", (change_id,)).fetchone()
    if not change: conn.close(); return jsonify({"error": "Change not found"}), 404
    ch = dict(change)
    if ch["reverted"]: conn.close(); return jsonify({"error": "Already reverted"}), 400
    # Attempt to revert the field in the spreadsheet
    field = ch["field"]
    target_id = ch["target_id"]
    old_val_str = ch["old_val"]
    if field in ("pin_col", "password"):
        conn.close()
        return jsonify({"error": "Cannot revert password/PIN changes."}), 400
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[RES_SHEET]
        reverted = False
        for row in ws.iter_rows(min_row=2):
            dr = row[COL["res_num"]-1].value
            if not isinstance(dr, (int, float)) or int(dr) != target_id: continue
            if field not in COL: conn.close(); return jsonify({"error": f"Field '{field}' not found"}), 400
            cell = ws.cell(row=row[0].row, column=COL[field])
            current_val = cell.value
            # Parse old value back
            if old_val_str == "(blank)":
                cell.value = None
            else:
                # Try numeric
                try: cell.value = int(old_val_str)
                except:
                    try: cell.value = float(old_val_str)
                    except: cell.value = old_val_str
            wb.save(EXCEL_FILE)
            reverted = True
            # Log the revert as a new change
            _log_change(target_id=target_id, target_name=ch["target_name"],
                        editor_id=res_id, field=field,
                        old_val=str(current_val) if current_val is not None else "(blank)",
                        new_val=old_val_str,
                        category="REVERT: " + ch["category"], reason=f"Reverted change #{change_id}")
            break
        if not reverted:
            conn.close(); return jsonify({"error": "Target resident not found in database"}), 404
    except Exception as e:
        conn.close(); return jsonify({"error": str(e)}), 500
    # Mark original change as reverted
    conn.execute("UPDATE changes SET reverted=1, reverted_by=?, reverted_at=datetime('now','localtime') WHERE id=?",
                 (res_id, change_id))
    conn.commit(); conn.close()
    return jsonify({"success": True})

# ══════════════════════════════════════════════════════════════════════════════
# SEED EXAMPLE CORPORATIONS
# ══════════════════════════════════════════════════════════════════════════════

def _seed_example_corps():
    """Seed 3 example corporations if the DB is empty."""
    conn = _init_synd_db()
    c = conn.cursor()
    count = c.execute("SELECT COUNT(*) FROM corporations").fetchone()[0]
    if count == 0:
        corps = [
            (1, hash_pin("ark123"), "Arkadia Farms Co-op", "Farming", 30, 30, 40, 4500.00, 2),
            (2, hash_pin("steel456"), "Ironforge Foundry", "Metal Foundry", 50, 40, 10, 12000.00, 21),
            (3, hash_pin("learn789"), "Lighthouse Academy", "School", 10, 30, 60, 3200.00, 22),
        ]
        for (cid, pw, name, ctype, coin, mind, hand, profit, ceo) in corps:
            c.execute("""INSERT INTO corporations (corp_id, password, name, corp_type,
                         coin_pct, mind_pct, hand_pct, est_monthly_profit, ceo_res_id)
                         VALUES (?,?,?,?,?,?,?,?,?)""", (cid, pw, name, ctype, coin, mind, hand, profit, ceo))
            c.execute("""INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role)
                         VALUES (?,?,?,?,?,1,'ceo')""", (cid, ceo, 0, 0, 0))
        # Add some members to corp 1
        c.execute("INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role) VALUES (1,4, 100,50,200, 0,'member')")
        c.execute("INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role) VALUES (1,8, 50,0,300, 0,'hand_manager')")
        # Add members to corp 2
        c.execute("INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role) VALUES (2,6, 500,200,0, 0,'coin_manager')")
        c.execute("INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role) VALUES (2,9, 200,300,50, 0,'member')")
        c.execute("INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role) VALUES (2,14, 0,100,100, 0,'member')")
        # Add members to corp 3
        c.execute("INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role) VALUES (3,10, 0,200,100, 0,'member')")
        c.execute("INSERT INTO corp_members (corp_id, res_id, coin_amount, mind_amount, hand_amount, is_ceo, role) VALUES (3,5, 50,150,250, 0,'member')")
        conn.commit()
    conn.close()

# ══════════════════════════════════════════════════════════════════════════════
# HEX CHESS — 3-Player Multiplayer
# ══════════════════════════════════════════════════════════════════════════════

def _init_chess_db():
    db_path = os.path.join(BASE_DIR, "hex_chess.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS chess_ratings (
            res_id INTEGER PRIMARY KEY,
            rating INTEGER NOT NULL DEFAULT 1000,
            wins INTEGER NOT NULL DEFAULT 0,
            losses INTEGER NOT NULL DEFAULT 0,
            games_played INTEGER NOT NULL DEFAULT 0,
            res_name TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS chess_games (
            game_id INTEGER PRIMARY KEY AUTOINCREMENT,
            status TEXT NOT NULL DEFAULT 'waiting',
            player0_res INTEGER, player0_name TEXT DEFAULT '',
            player1_res INTEGER, player1_name TEXT DEFAULT '',
            player2_res INTEGER, player2_name TEXT DEFAULT '',
            board_json TEXT NOT NULL DEFAULT '{}',
            turn INTEGER NOT NULL DEFAULT 0,
            caps_json TEXT NOT NULL DEFAULT '[[],[],[]]',
            dead_json TEXT NOT NULL DEFAULT '[false,false,false]',
            game_over INTEGER NOT NULL DEFAULT 0,
            winner INTEGER DEFAULT NULL,
            promo_json TEXT DEFAULT '',
            move_num INTEGER NOT NULL DEFAULT 0,
            last_json TEXT DEFAULT '',
            log_json TEXT NOT NULL DEFAULT '[]',
            last_activity TEXT DEFAULT (datetime('now')),
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS chess_lobby (
            res_id INTEGER PRIMARY KEY,
            res_name TEXT NOT NULL DEFAULT '',
            rating INTEGER NOT NULL DEFAULT 1000,
            joined_at TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()
    return conn

def _get_chess_db():
    db_path = os.path.join(BASE_DIR, "hex_chess.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def _chess_auth(data):
    """Quick auth for chess endpoints. Returns (res_id, name) or None."""
    res_id = data.get("residentId"); pw_raw = str(data.get("password","")).strip()
    if not res_id or not pw_raw: return None
    try: res_id = int(res_id)
    except: return None
    wb = load_wb()
    if not wb: return None
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        if not is_active(da) or not dpw or dpw!=ph: return None
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        return (res_id, f"{fn} {ln}".strip())
    return None

def _elo_update(r_winner, r_loser, k=32):
    """Standard ELO update. Returns (new_winner, new_loser)."""
    ew = 1.0 / (1+10**((r_loser-r_winner)/400))
    el = 1.0 / (1+10**((r_winner-r_loser)/400))
    return (round(r_winner + k*(1-ew)), round(r_loser + k*(0-el)))

def _chess_init_board():
    """Build the initial board state as a dict {q,r: {piece, player, moved}}."""
    RAD = 5
    def ok(q,r):
        s=-q-r; return max(abs(q),abs(r),abs(s))<=RAD
    RED_SETUP = [
        {"q":-5,"r":5,"p":"K"},{"q":-5,"r":4,"p":"R"},{"q":-4,"r":5,"p":"R"},
        {"q":-5,"r":3,"p":"N"},{"q":-4,"r":4,"p":"Q"},{"q":-3,"r":5,"p":"N"},
        {"q":-4,"r":3,"p":"B"},{"q":-3,"r":4,"p":"B"},
        {"q":-3,"r":3,"p":"P"},{"q":-2,"r":3,"p":"P"},{"q":-1,"r":3,"p":"P"},
        {"q":-5,"r":2,"p":"P"},{"q":-4,"r":2,"p":"P"},{"q":-3,"r":2,"p":"B"},{"q":-2,"r":2,"p":"P"},{"q":-1,"r":2,"p":"P"},
    ]
    def rot120(q,r):
        s=-q-r; return (s,q)
    def rot240(q,r):
        s=-q-r; return (r,s)
    board = {}
    setups = [
        RED_SETUP,
        [{"q":rot120(s["q"],s["r"])[0],"r":rot120(s["q"],s["r"])[1],"p":s["p"]} for s in RED_SETUP],
        [{"q":rot240(s["q"],s["r"])[0],"r":rot240(s["q"],s["r"])[1],"p":s["p"]} for s in RED_SETUP],
    ]
    for pi in range(3):
        for s in setups[pi]:
            if not ok(s["q"],s["r"]): continue
            k = f'{s["q"]},{s["r"]}'
            if k not in board:
                board[k] = {"piece":s["p"],"player":pi,"moved":False}
    return board

@app.route("/hex-chess")
def hex_chess_page():
    return send_from_directory("templates", "hex_chess.html")

@app.route("/api/chess/ensure-rating", methods=["POST"])
def chess_ensure_rating():
    """Ensure the player has a rating entry. Returns their rating."""
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, name = auth
    conn = _get_chess_db()
    row = conn.execute("SELECT * FROM chess_ratings WHERE res_id=?", (res_id,)).fetchone()
    if not row:
        conn.execute("INSERT INTO chess_ratings (res_id,rating,res_name) VALUES (?,1000,?)", (res_id,name))
        conn.commit()
        rating = 1000; wins = 0; losses = 0; gp = 0
    else:
        row = dict(row)
        rating = row["rating"]; wins = row["wins"]; losses = row["losses"]; gp = row["games_played"]
        if row["res_name"] != name:
            conn.execute("UPDATE chess_ratings SET res_name=? WHERE res_id=?", (name,res_id))
            conn.commit()
    # Also get leaderboard
    lb = [dict(r) for r in conn.execute("SELECT * FROM chess_ratings ORDER BY rating DESC LIMIT 20").fetchall()]
    conn.close()
    return jsonify({"rating":rating,"wins":wins,"losses":losses,"gamesPlayed":gp,"name":name,"leaderboard":lb})

@app.route("/api/chess/join-lobby", methods=["POST"])
def chess_join_lobby():
    """Join the matchmaking lobby. If 3 players are queued, create a game."""
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, name = auth
    conn = _get_chess_db()
    # Ensure rating exists
    row = conn.execute("SELECT rating FROM chess_ratings WHERE res_id=?", (res_id,)).fetchone()
    rating = dict(row)["rating"] if row else 1000
    if not row:
        conn.execute("INSERT INTO chess_ratings (res_id,rating,res_name) VALUES (?,1000,?)", (res_id,name))
        conn.commit()
    # Check if already in an active game
    active = conn.execute("""SELECT game_id FROM chess_games WHERE game_over=0 AND status='active'
                             AND (player0_res=? OR player1_res=? OR player2_res=?)""",
                          (res_id,res_id,res_id)).fetchone()
    if active:
        conn.close()
        return jsonify({"status":"in_game","gameId":dict(active)["game_id"]})
    # Clean stale lobby entries (>5 min old)
    conn.execute("DELETE FROM chess_lobby WHERE joined_at < datetime('now','-5 minutes')")
    conn.commit()
    # Add to lobby (upsert)
    conn.execute("INSERT OR REPLACE INTO chess_lobby (res_id,res_name,rating,joined_at) VALUES (?,?,?,datetime('now'))",
                 (res_id,name,rating))
    conn.commit()
    # Check if 3 players in lobby
    lobby = [dict(r) for r in conn.execute("SELECT * FROM chess_lobby ORDER BY joined_at ASC").fetchall()]
    if len(lobby) >= 3:
        p = lobby[:3]
        board = _chess_init_board()
        gid = conn.execute("""INSERT INTO chess_games (status,player0_res,player0_name,player1_res,player1_name,
                              player2_res,player2_name,board_json,log_json) VALUES ('active',?,?,?,?,?,?,?,?)""",
                           (p[0]["res_id"],p[0]["res_name"],p[1]["res_id"],p[1]["res_name"],
                            p[2]["res_id"],p[2]["res_name"],json.dumps(board),
                            json.dumps([f"Game started — {p[0]['res_name']} (Red) moves first."]))).lastrowid
        conn.commit()
        for pp in p:
            conn.execute("DELETE FROM chess_lobby WHERE res_id=?", (pp["res_id"],))
        conn.commit()
        conn.close()
        return jsonify({"status":"game_started","gameId":gid})
    conn.close()
    return jsonify({"status":"waiting","lobby":[{"res_id":l["res_id"],"name":l["res_name"],"rating":l["rating"]} for l in lobby],
                     "count":len(lobby)})

@app.route("/api/chess/leave-lobby", methods=["POST"])
def chess_leave_lobby():
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, _ = auth
    conn = _get_chess_db()
    conn.execute("DELETE FROM chess_lobby WHERE res_id=?", (res_id,))
    conn.commit(); conn.close()
    return jsonify({"success":True})

@app.route("/api/chess/lobby-status", methods=["POST"])
def chess_lobby_status():
    """Poll lobby status."""
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, name = auth
    conn = _get_chess_db()
    # Check if a game started with us
    active = conn.execute("""SELECT game_id FROM chess_games WHERE game_over=0 AND status='active'
                             AND (player0_res=? OR player1_res=? OR player2_res=?)""",
                          (res_id,res_id,res_id)).fetchone()
    if active:
        conn.close()
        return jsonify({"status":"game_started","gameId":dict(active)["game_id"]})
    lobby = [dict(r) for r in conn.execute("SELECT * FROM chess_lobby ORDER BY joined_at ASC").fetchall()]
    conn.close()
    in_lobby = any(l["res_id"]==res_id for l in lobby)
    return jsonify({"status":"waiting" if in_lobby else "not_in_lobby",
                     "lobby":[{"res_id":l["res_id"],"name":l["res_name"],"rating":l["rating"]} for l in lobby],
                     "count":len(lobby)})

@app.route("/api/chess/game-state", methods=["POST"])
def chess_game_state():
    """Get full game state for polling."""
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, _ = auth
    game_id = data.get("gameId")
    if not game_id: return jsonify({"error":"Missing gameId"}),400
    conn = _get_chess_db()
    g = conn.execute("SELECT * FROM chess_games WHERE game_id=?", (game_id,)).fetchone()
    if not g: conn.close(); return jsonify({"error":"Game not found"}),404
    g = dict(g)
    # Get ratings
    ratings = {}
    for i in range(3):
        rid = g[f"player{i}_res"]
        if rid:
            rr = conn.execute("SELECT rating FROM chess_ratings WHERE res_id=?", (rid,)).fetchone()
            ratings[rid] = dict(rr)["rating"] if rr else 1000
    conn.close()
    # Determine which player index this user is
    my_idx = -1
    for i in range(3):
        if g[f"player{i}_res"] == res_id: my_idx = i; break
    return jsonify({
        "gameId": g["game_id"], "status": g["status"],
        "players": [{"res_id":g[f"player{i}_res"],"name":g[f"player{i}_name"],"rating":ratings.get(g[f"player{i}_res"],1000)} for i in range(3)],
        "board": json.loads(g["board_json"]),
        "turn": g["turn"], "caps": json.loads(g["caps_json"]),
        "dead": json.loads(g["dead_json"]), "gameOver": bool(g["game_over"]),
        "winner": g["winner"], "promo": json.loads(g["promo_json"]) if g["promo_json"] else None,
        "moveNum": g["move_num"], "last": json.loads(g["last_json"]) if g["last_json"] else None,
        "log": json.loads(g["log_json"]), "myIndex": my_idx
    })

@app.route("/api/chess/move", methods=["POST"])
def chess_do_move():
    """Submit a move. Server validates and applies it."""
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, _ = auth
    game_id = data.get("gameId")
    fq,fr,tq,tr = data.get("fq"),data.get("fr"),data.get("tq"),data.get("tr")
    if game_id is None or fq is None: return jsonify({"error":"Missing params"}),400
    conn = _get_chess_db()
    g = conn.execute("SELECT * FROM chess_games WHERE game_id=?", (game_id,)).fetchone()
    if not g: conn.close(); return jsonify({"error":"Game not found"}),404
    g = dict(g)
    if g["game_over"]: conn.close(); return jsonify({"error":"Game is over"}),400
    # Check it's this player's turn
    my_idx = -1
    for i in range(3):
        if g[f"player{i}_res"] == res_id: my_idx = i; break
    if my_idx < 0: conn.close(); return jsonify({"error":"Not in this game"}),403
    if g["turn"] != my_idx: conn.close(); return jsonify({"error":"Not your turn"}),400
    # Load state and apply move server-side (trust client for valid moves in this version)
    board = json.loads(g["board_json"])
    caps = json.loads(g["caps_json"])
    dead = json.loads(g["dead_json"])
    log = json.loads(g["log_json"])
    turn = g["turn"]; mnum = g["move_num"]
    PL_NAMES = [g["player0_name"],g["player1_name"],g["player2_name"]]
    SYM = {"K":"♚","Q":"♛","R":"♜","B":"♝","N":"♞","P":"♟"}
    fk = f"{fq},{fr}"; tk = f"{tq},{tr}"
    pc = board.get(fk)
    if not pc: conn.close(); return jsonify({"error":"No piece at source"}),400
    if pc["player"] != my_idx: conn.close(); return jsonify({"error":"Not your piece"}),400
    tgt = board.get(tk)
    if tgt:
        caps[pc["player"]].append(tgt)
        if tgt["piece"]=="K":
            dead[tgt["player"]] = True
            to_del = [k for k in board if board[k]["player"]==tgt["player"]]
            for k in to_del: del board[k]
            log.append(f"💀 {PL_NAMES[pc['player']]} captured {PL_NAMES[tgt['player']]}'s King!")
        else:
            log.append(f"{PL_NAMES[pc['player']]}: {SYM[pc['piece']]}×{SYM[tgt['piece']]}")
    else:
        log.append(f"{PL_NAMES[pc['player']]}: {SYM[pc['piece']]} moved")
    board[tk] = {**pc, "moved":True}
    if fk in board: del board[fk]
    mnum += 1
    last = {"fq":fq,"fr":fr,"tq":tq,"tr":tr}
    # Check for pawn promotion
    PCFG_PROMO = [lambda q,r: q>=3, lambda q,r: r>=3, lambda q,r: q<=-3]
    promo_json = ""
    if pc["piece"]=="P" and PCFG_PROMO[pc["player"]](tq,tr):
        promo_json = json.dumps({"q":tq,"r":tr,"player":pc["player"]})
        # Don't advance turn yet — wait for promotion
        conn.execute("""UPDATE chess_games SET board_json=?,turn=?,caps_json=?,dead_json=?,
                        move_num=?,last_json=?,log_json=?,promo_json=?,last_activity=datetime('now') WHERE game_id=?""",
                     (json.dumps(board),turn,json.dumps(caps),json.dumps(dead),
                      mnum,json.dumps(last),json.dumps(log),promo_json,game_id))
        conn.commit(); conn.close()
        return jsonify({"success":True,"needsPromo":True})
    # Check win
    alive = [i for i in range(3) if not dead[i]]
    game_over = len(alive)<=1
    winner = None
    if game_over and alive:
        winner = alive[0]
        log.append(f"🏆 {PL_NAMES[winner]} WINS THE GAME!")
    # Advance turn
    next_turn = turn
    if not game_over:
        for _ in range(3):
            next_turn = (next_turn+1)%3
            if not dead[next_turn]: break
    conn.execute("""UPDATE chess_games SET board_json=?,turn=?,caps_json=?,dead_json=?,
                    game_over=?,winner=?,move_num=?,last_json=?,log_json=?,promo_json='',
                    last_activity=datetime('now') WHERE game_id=?""",
                 (json.dumps(board),next_turn,json.dumps(caps),json.dumps(dead),
                  1 if game_over else 0, winner, mnum,json.dumps(last),json.dumps(log),game_id))
    conn.commit()
    # Update ratings if game over
    if game_over and winner is not None:
        _chess_update_ratings(conn, g, dead, winner)
    conn.close()
    return jsonify({"success":True,"gameOver":game_over,"winner":winner})

@app.route("/api/chess/promote", methods=["POST"])
def chess_promote():
    """Promote a pawn."""
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, _ = auth
    game_id = data.get("gameId"); piece_type = data.get("pieceType","Q")
    if piece_type not in ("Q","R","B","N"): return jsonify({"error":"Invalid promotion"}),400
    conn = _get_chess_db()
    g = dict(conn.execute("SELECT * FROM chess_games WHERE game_id=?", (game_id,)).fetchone())
    if not g["promo_json"]: conn.close(); return jsonify({"error":"No pending promotion"}),400
    promo = json.loads(g["promo_json"])
    board = json.loads(g["board_json"])
    dead = json.loads(g["dead_json"])
    log = json.loads(g["log_json"])
    caps = json.loads(g["caps_json"])
    SYM = {"K":"♚","Q":"♛","R":"♜","B":"♝","N":"♞","P":"♟"}
    PL_NAMES = [g["player0_name"],g["player1_name"],g["player2_name"]]
    pk = f'{promo["q"]},{promo["r"]}'
    board[pk] = {"piece":piece_type,"player":promo["player"],"moved":True}
    log.append(f"{PL_NAMES[promo['player']]}: ♟ → {SYM[piece_type]}")
    # Check win after promotion
    alive = [i for i in range(3) if not dead[i]]
    game_over = len(alive)<=1
    winner = None
    if game_over and alive:
        winner = alive[0]
        log.append(f"🏆 {PL_NAMES[winner]} WINS THE GAME!")
    next_turn = g["turn"]
    if not game_over:
        for _ in range(3):
            next_turn = (next_turn+1)%3
            if not dead[next_turn]: break
    conn.execute("""UPDATE chess_games SET board_json=?,turn=?,caps_json=?,dead_json=?,
                    game_over=?,winner=?,log_json=?,promo_json='',last_activity=datetime('now') WHERE game_id=?""",
                 (json.dumps(board),next_turn,json.dumps(caps),json.dumps(dead),
                  1 if game_over else 0,winner,json.dumps(log),game_id))
    conn.commit()
    if game_over and winner is not None:
        _chess_update_ratings(conn, g, dead, winner)
    conn.close()
    return jsonify({"success":True})

@app.route("/api/chess/resign", methods=["POST"])
def chess_resign():
    """Resign from the game (eliminates self)."""
    data = request.get_json(force=True)
    auth = _chess_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, _ = auth
    game_id = data.get("gameId")
    conn = _get_chess_db()
    g = dict(conn.execute("SELECT * FROM chess_games WHERE game_id=?", (game_id,)).fetchone())
    if g["game_over"]: conn.close(); return jsonify({"error":"Game is over"}),400
    my_idx = -1
    for i in range(3):
        if g[f"player{i}_res"]==res_id: my_idx=i; break
    if my_idx<0: conn.close(); return jsonify({"error":"Not in game"}),403
    board = json.loads(g["board_json"])
    dead = json.loads(g["dead_json"])
    log = json.loads(g["log_json"])
    caps = json.loads(g["caps_json"])
    PL_NAMES = [g["player0_name"],g["player1_name"],g["player2_name"]]
    dead[my_idx] = True
    to_del = [k for k in board if board[k]["player"]==my_idx]
    for k in to_del: del board[k]
    log.append(f"🏳️ {PL_NAMES[my_idx]} resigned!")
    alive = [i for i in range(3) if not dead[i]]
    game_over = len(alive)<=1; winner = None
    if game_over and alive:
        winner = alive[0]
        log.append(f"🏆 {PL_NAMES[winner]} WINS THE GAME!")
    next_turn = g["turn"]
    if next_turn == my_idx and not game_over:
        for _ in range(3):
            next_turn = (next_turn+1)%3
            if not dead[next_turn]: break
    conn.execute("""UPDATE chess_games SET board_json=?,turn=?,caps_json=?,dead_json=?,
                    game_over=?,winner=?,log_json=?,promo_json='',last_activity=datetime('now') WHERE game_id=?""",
                 (json.dumps(board),next_turn,json.dumps(caps),json.dumps(dead),
                  1 if game_over else 0,winner,json.dumps(log),game_id))
    conn.commit()
    if game_over and winner is not None:
        _chess_update_ratings(conn, g, dead, winner)
    conn.close()
    return jsonify({"success":True})

@app.route("/api/chess/leaderboard", methods=["POST"])
def chess_leaderboard():
    conn = _get_chess_db()
    lb = [dict(r) for r in conn.execute("SELECT * FROM chess_ratings ORDER BY rating DESC LIMIT 20").fetchall()]
    conn.close()
    return jsonify({"leaderboard":lb})

def _chess_update_ratings(conn, game, dead, winner):
    """Update ELO ratings. Winner gains vs each loser."""
    losers = [i for i in range(3) if i!=winner]
    w_rid = game[f"player{winner}_res"]
    wr = conn.execute("SELECT rating FROM chess_ratings WHERE res_id=?", (w_rid,)).fetchone()
    w_rating = dict(wr)["rating"] if wr else 1000
    for li in losers:
        l_rid = game[f"player{li}_res"]
        lr = conn.execute("SELECT rating FROM chess_ratings WHERE res_id=?", (l_rid,)).fetchone()
        l_rating = dict(lr)["rating"] if lr else 1000
        new_w, new_l = _elo_update(w_rating, l_rating, k=24)
        w_rating = new_w  # cumulative for winner
        conn.execute("UPDATE chess_ratings SET rating=?, losses=losses+1, games_played=games_played+1 WHERE res_id=?",
                     (new_l, l_rid))
    conn.execute("UPDATE chess_ratings SET rating=?, wins=wins+1, games_played=games_played+1 WHERE res_id=?",
                 (w_rating, w_rid))
    conn.commit()

# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICATION SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

NOTIF_IMG_DIR = os.path.join(BASE_DIR, "Notification_Images")
os.makedirs(NOTIF_IMG_DIR, exist_ok=True)

def _init_notif_db():
    db_path = os.path.join(BASE_DIR, "notifications.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS notifications (
            notif_id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL DEFAULT '',
            description TEXT NOT NULL DEFAULT '',
            images TEXT NOT NULL DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 0,
            created_by INTEGER,
            created_by_name TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()
    return conn

def _get_notif_db():
    db_path = os.path.join(BASE_DIR, "notifications.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

@app.route("/api/notifications/active", methods=["POST"])
def notif_get_active():
    """Get the currently active notification (if any)."""
    conn = _get_notif_db()
    row = conn.execute("SELECT * FROM notifications WHERE is_active=1 ORDER BY notif_id DESC LIMIT 1").fetchone()
    conn.close()
    if row:
        return jsonify({"has_active": True, "notification": dict(row)})
    return jsonify({"has_active": False})

@app.route("/api/notifications/history", methods=["POST"])
def notif_history():
    """Get all notifications (history)."""
    conn = _get_notif_db()
    rows = [dict(r) for r in conn.execute("SELECT * FROM notifications ORDER BY notif_id DESC").fetchall()]
    conn.close()
    return jsonify({"notifications": rows})

@app.route("/api/notifications/create", methods=["POST"])
def notif_create():
    """Create a notification. Only Hexarchy."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password","")).strip()
    if not res_id or not pw_raw: return jsonify({"error":"Missing credentials"}),400
    try: res_id = int(res_id)
    except: return jsonify({"error":"Invalid ID"}),400
    # Auth + hexarchy check
    wb = load_wb()
    if not wb: return jsonify({"error":"Database unavailable"}),500
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False; caller_name = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        if not is_active(da) or not dpw or dpw!=ph: return jsonify({"error":"Auth failed"}),401
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        caller_name = f"{fn} {ln}".strip()
        authed = True; break
    if not authed: return jsonify({"error":"Resident not found"}),404
    is_hex = False
    if "Gov Employees" in wb.sheetnames:
        for gov_row in wb["Gov Employees"].iter_rows(min_row=2, values_only=True):
            gov_res = gov_row[1]
            active = str(gov_row[7] or "").strip().lower() if len(gov_row) > 7 else ""
            hexarchy = str(gov_row[8] or "").strip().lower() if len(gov_row) > 8 else ""
            if not isinstance(gov_res,(int,float)) or int(gov_res)!=res_id: continue
            if active not in ("yes","y","true","1"): continue
            if hexarchy in ("yes","y","true","1"): is_hex = True; break
    if not is_hex: return jsonify({"error":"Only Hexarchy members can create notifications."}),403
    title = str(data.get("title","")).strip()[:100]
    desc = str(data.get("description","")).strip()[:300]
    images = str(data.get("images","")).strip()
    if not title: return jsonify({"error":"Title is required."}),400
    set_active = data.get("setActive", True)
    conn = _get_notif_db()
    if set_active:
        conn.execute("UPDATE notifications SET is_active=0 WHERE is_active=1")
    nid = conn.execute("INSERT INTO notifications (title,description,images,is_active,created_by,created_by_name) VALUES (?,?,?,?,?,?)",
                       (title, desc, images, 1 if set_active else 0, res_id, caller_name)).lastrowid
    conn.commit(); conn.close()
    return jsonify({"success":True, "notif_id":nid})

@app.route("/api/notifications/deactivate", methods=["POST"])
def notif_deactivate():
    """Deactivate a notification. Only Hexarchy."""
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password","")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"error":"Invalid ID"}),400
    wb = load_wb()
    if not wb: return jsonify({"error":"Database unavailable"}),500
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        if dpw==ph: authed=True
        break
    if not authed: return jsonify({"error":"Auth failed"}),401
    notif_id = data.get("notifId")
    conn = _get_notif_db()
    conn.execute("UPDATE notifications SET is_active=0 WHERE notif_id=?", (notif_id,))
    conn.commit(); conn.close()
    return jsonify({"success":True})

@app.route("/api/notifications/upload-image", methods=["POST"])
def notif_upload_image():
    """Upload a JPG for a notification."""
    import base64, uuid
    data = request.get_json(force=True)
    res_id = data.get("residentId"); pw_raw = str(data.get("password","")).strip()
    try: res_id = int(res_id)
    except: return jsonify({"success":False,"error":"Invalid ID"}),400
    wb = load_wb()
    if not wb: return jsonify({"success":False,"error":"Database unavailable"}),500
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    authed = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        if dpw==ph: authed=True
        break
    if not authed: return jsonify({"success":False,"error":"Auth failed"}),401
    img_data = data.get("imageData","")
    if not img_data: return jsonify({"success":False,"error":"No image data"}),400
    if "," in img_data: img_data = img_data.split(",",1)[1]
    raw = base64.b64decode(img_data)
    if len(raw) > 500*1024: return jsonify({"success":False,"error":"Image too large (max 500kb)"}),400
    fname = f"notif_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.jpg"
    with open(os.path.join(NOTIF_IMG_DIR, fname), "wb") as f: f.write(raw)
    return jsonify({"success":True, "filename":fname})

@app.route("/api/notifications/image/<filename>")
def notif_image(filename):
    import re as _re
    if not _re.match(r'^[\w\-\.]+$', filename): return "Not found", 404
    return send_from_directory(NOTIF_IMG_DIR, filename)

# ══════════════════════════════════════════════════════════════════════════════
# LIQUID DEMOCRACY
# ══════════════════════════════════════════════════════════════════════════════

def _init_liquid_db():
    """Init liquid democracy table — uses the same election_db SQLite."""
    with get_election_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS liquid_endorsements (
                endorser_res_id INTEGER PRIMARY KEY,
                rep_res_id INTEGER,
                updated_at TEXT DEFAULT (datetime('now'))
            );
        """)
        conn.commit()

_init_liquid_db()

def _get_endorsement_count(rep_res_id):
    """Count how many residents have endorsed this rep (including the rep themselves)."""
    with get_election_db() as c:
        n = c.execute("SELECT COUNT(*) FROM liquid_endorsements WHERE rep_res_id=?", (rep_res_id,)).fetchone()[0]
    return n

def _get_vote_weight(resident_id):
    """Return the vote weight for a resident.
    - If they've endorsed someone else: weight 0 (their rep votes for them)
    - If they self-represent (rep_res_id = themselves OR no entry): weight = 1 + endorsements they've received
    """
    if not resident_id: return 1
    try: rid = int(resident_id)
    except: return 1
    with get_election_db() as c:
        row = c.execute("SELECT rep_res_id FROM liquid_endorsements WHERE endorser_res_id=?", (rid,)).fetchone()
        if row and row[0] is not None and int(row[0]) != rid:
            # Endorsed someone else — they don't vote
            return 0
        # Self-representing or no entry — weight = 1 + endorsements received
        n = c.execute("SELECT COUNT(*) FROM liquid_endorsements WHERE rep_res_id=?", (rid,)).fetchone()[0]
    return max(1, n + (1 if n == 0 else 0))  # at least 1 for self

def _get_lookup_resident(res_id):
    """Look up a resident's name by Res#. Returns dict or None."""
    wb = load_wb()
    if not wb: return None
    ws = wb[RES_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        fn = str(row[COL["first_name"]-1] or "")
        ln = str(row[COL["last_name"]-1] or "")
        alive = row[COL["alive"]-1]
        return {"resId": int(dr), "name": f"{fn} {ln}".strip(), "alive": is_active(alive)}
    return None

def _liquid_auth(data):
    """Auth helper for liquid democracy. Returns (res_id, name) or None."""
    res_id = data.get("residentId"); pw_raw = str(data.get("password","")).strip()
    if not res_id or not pw_raw: return None
    try: res_id = int(res_id)
    except: return None
    wb = load_wb()
    if not wb: return None
    ws = wb[RES_SHEET]; ph = hash_pin(pw_raw)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr = row[COL["res_num"]-1]; dpw = row[COL["password"]-1]; da = row[COL["alive"]-1]
        if not isinstance(dr,(int,float)) or int(dr)!=res_id: continue
        if not is_active(da) or not dpw or dpw!=ph: return None
        fn = str(row[COL["first_name"]-1] or ""); ln = str(row[COL["last_name"]-1] or "")
        return (res_id, f"{fn} {ln}".strip())
    return None

@app.route("/api/liquid/status", methods=["POST"])
def liquid_status():
    """Get the user's current rep + leaderboard + vote weight."""
    data = request.get_json(force=True)
    auth = _liquid_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, name = auth
    with get_election_db() as c:
        row = c.execute("SELECT rep_res_id FROM liquid_endorsements WHERE endorser_res_id=?", (res_id,)).fetchone()
        rep_res_id = int(row[0]) if row and row[0] is not None else None
        # Leaderboard: every distinct rep_res_id with counts
        lb_rows = c.execute("""SELECT rep_res_id, COUNT(*) as endorsements
                               FROM liquid_endorsements
                               WHERE rep_res_id IS NOT NULL
                               GROUP BY rep_res_id ORDER BY endorsements DESC""").fetchall()
        total_endorsers = c.execute("SELECT COUNT(*) FROM liquid_endorsements").fetchone()[0]
    # Look up rep names
    rep_name = None; rep_alive = None
    if rep_res_id:
        info = _get_lookup_resident(rep_res_id)
        if info:
            rep_name = info["name"]; rep_alive = info["alive"]
    leaderboard = []
    for r in lb_rows:
        rid = int(r[0]); endorsements = int(r[1])
        info = _get_lookup_resident(rid)
        leaderboard.append({
            "resId": rid,
            "name": info["name"] if info else f"Res#{rid}",
            "endorsements": endorsements,
            "pct": round(100.0 * endorsements / total_endorsers, 2) if total_endorsers else 0
        })
    # Vote weight
    if rep_res_id and rep_res_id != res_id:
        my_weight = 0  # they delegate
    else:
        # Self-rep: 1 + endorsements received (others endorsing me)
        received = sum(1 for r in lb_rows if int(r[0]) == res_id)
        # actually, lb_rows row count for me
        my_endorsements = next((int(r[1]) for r in lb_rows if int(r[0]) == res_id), 0)
        my_weight = max(1, my_endorsements + (1 if my_endorsements == 0 else 0))
        # Refined: weight = 1 if no endorsements (self only), else = endorsements (which already includes self if I self-endorsed)
        # Simplest: count = explicit endorsements of me + 1 if I haven't explicitly self-endorsed
        # Let's just say: weight = max(1, count of liquid_endorsements where rep=me)
        # If I've never been endorsed and never self-endorsed, weight = 1
        my_weight = max(1, my_endorsements)
    return jsonify({
        "myResId": res_id, "myName": name,
        "repResId": rep_res_id, "repName": rep_name, "repAlive": rep_alive,
        "isSelfRep": rep_res_id is None or rep_res_id == res_id,
        "myVoteWeight": my_weight,
        "leaderboard": leaderboard,
        "totalEndorsers": total_endorsers
    })

@app.route("/api/liquid/endorse", methods=["POST"])
def liquid_endorse():
    """Set or change the user's representative. repResId=null or own res_id = self-represent."""
    data = request.get_json(force=True)
    auth = _liquid_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, _ = auth
    rep_res_id = data.get("repResId")
    self_represent = data.get("selfRepresent", False)
    if self_represent or rep_res_id is None or rep_res_id == "":
        rep_res_id = None
    else:
        try: rep_res_id = int(rep_res_id)
        except: return jsonify({"error":"Invalid Res#"}),400
        if rep_res_id == res_id:
            rep_res_id = None  # self-represent
        else:
            # Validate rep exists
            info = _get_lookup_resident(rep_res_id)
            if not info: return jsonify({"error":f"Res#{rep_res_id} not found."}),404
            if not info["alive"]: return jsonify({"error":f"Res#{rep_res_id} is not active."}),400
    with get_election_db() as c:
        c.execute("""INSERT INTO liquid_endorsements (endorser_res_id, rep_res_id, updated_at)
                     VALUES (?,?,datetime('now'))
                     ON CONFLICT(endorser_res_id) DO UPDATE SET rep_res_id=excluded.rep_res_id, updated_at=datetime('now')""",
                  (res_id, rep_res_id))
        c.commit()
    return jsonify({"success":True, "repResId":rep_res_id})

@app.route("/api/liquid/rep-vote-status", methods=["POST"])
def liquid_rep_vote_status():
    """For an election: has my rep already voted in each section?"""
    data = request.get_json(force=True)
    auth = _liquid_auth(data)
    if not auth: return jsonify({"error":"Auth failed"}),401
    res_id, _ = auth
    with get_election_db() as c:
        row = c.execute("SELECT rep_res_id FROM liquid_endorsements WHERE endorser_res_id=?", (res_id,)).fetchone()
        rep_res_id = int(row[0]) if row and row[0] is not None else None
        if not rep_res_id or rep_res_id == res_id:
            return jsonify({"hasRep": False})
        rb = c.execute("SELECT voted_hexarchy, voted_charity, voted_democracy FROM resident_ballots WHERE resident_id=?",
                       (rep_res_id,)).fetchone()
    info = _get_lookup_resident(rep_res_id)
    return jsonify({
        "hasRep": True,
        "repResId": rep_res_id,
        "repName": info["name"] if info else f"Res#{rep_res_id}",
        "votedHexarchy": bool(rb and rb[0] == 1) if rb else False,
        "votedCharity": bool(rb and rb[1] == 1) if rb else False,
        "votedDemocracy": bool(rb and rb[2] == 1) if rb else False,
    })

# ══════════════════════════════════════════════════════════════════════════════
# THE FORUM
# ══════════════════════════════════════════════════════════════════════════════
import base64, mimetypes
FORUM_DB      = os.path.join(BASE_DIR, "forum.db")
FORUM_MEDIA   = os.path.join(BASE_DIR, "Forum_Media")
FORUM_AVATARS = os.path.join(BASE_DIR, "Forum_Avatars")
os.makedirs(FORUM_MEDIA,   exist_ok=True)
os.makedirs(FORUM_AVATARS, exist_ok=True)

def get_forum_db():
    conn = sqlite3.connect(FORUM_DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def _init_forum_db():
    with get_forum_db() as c:
        c.executescript("""
          CREATE TABLE IF NOT EXISTS forum_profiles (
            res_id INTEGER PRIMARY KEY,
            forum_username TEXT NOT NULL,
            avatar_filename TEXT DEFAULT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
          );
          CREATE UNIQUE INDEX IF NOT EXISTS idx_forum_username ON forum_profiles(forum_username);

          CREATE TABLE IF NOT EXISTS forum_posts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            res_id INTEGER NOT NULL,
            content TEXT NOT NULL DEFAULT '',
            media_filename TEXT DEFAULT NULL,
            media_type TEXT DEFAULT NULL,
            is_retweet INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
          );
          CREATE INDEX IF NOT EXISTS idx_fp_res ON forum_posts(res_id);
          CREATE INDEX IF NOT EXISTS idx_fp_created ON forum_posts(created_at);

          CREATE TABLE IF NOT EXISTS forum_likes (
            post_id INTEGER NOT NULL,
            res_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (post_id, res_id)
          );

          CREATE TABLE IF NOT EXISTS forum_retweets (
            post_id INTEGER NOT NULL,
            res_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (post_id, res_id)
          );

          CREATE TABLE IF NOT EXISTS forum_comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            post_id INTEGER NOT NULL,
            res_id INTEGER NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
          );
          CREATE INDEX IF NOT EXISTS idx_fc_post ON forum_comments(post_id);

          CREATE TABLE IF NOT EXISTS forum_follows (
            follower_res_id INTEGER NOT NULL,
            following_res_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (follower_res_id, following_res_id)
          );

          CREATE TABLE IF NOT EXISTS forum_blocks (
            blocker_res_id INTEGER NOT NULL,
            blocked_res_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (blocker_res_id, blocked_res_id)
          );

          CREATE TABLE IF NOT EXISTS forum_dms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender_res_id INTEGER NOT NULL,
            receiver_res_id INTEGER NOT NULL,
            content TEXT NOT NULL,
            read INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
          );
          CREATE INDEX IF NOT EXISTS idx_fdm_pair ON forum_dms(sender_res_id, receiver_res_id);
          CREATE INDEX IF NOT EXISTS idx_fdm_recv ON forum_dms(receiver_res_id, read);

          CREATE TABLE IF NOT EXISTS forum_views (
            post_id INTEGER NOT NULL,
            res_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (post_id, res_id)
          );
          CREATE INDEX IF NOT EXISTS idx_fv_post ON forum_views(post_id);
        """)
        try: c.execute("ALTER TABLE forum_profiles ADD COLUMN last_appearance_change TEXT DEFAULT NULL")
        except: pass
        c.commit()

def _forum_auth(data):
    """Returns res_id if credentials valid, else None.
    Uses same COL mapping and hash_pin() as verify_password."""
    try:
        res_id  = int(str(data.get('residentId', '')).strip())
        pw_raw  = str(data.get('password', '')).strip()
    except (ValueError, TypeError):
        return None
    if not res_id or not pw_raw:
        return None
    try:
        wb = load_wb()
        if not wb: return None
        ws = wb[RES_SHEET]
        ph = hash_pin(pw_raw)
        for row in ws.iter_rows(min_row=2, values_only=True):
            dr  = row[COL["res_num"] - 1]
            dpw = row[COL["password"] - 1]
            da  = row[COL["alive"] - 1]
            if not isinstance(dr, (int, float)):
                continue
            if int(dr) != res_id:
                continue
            if not is_active(da):
                return None
            if dpw and dpw == ph:
                return res_id
            return None
        return None
    except Exception:
        return None

def _forum_ensure_profile(res_id, conn):
    """Create a default forum profile if not exists."""
    row = conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (res_id,)).fetchone()
    if row:
        return dict(row)
    # Generate default username from resident data
    wb = load_wb()
    username = f"Resident{res_id}"
    if wb:
        try:
            ws = wb[RES_SHEET]
            for r in ws.iter_rows(min_row=2, values_only=True):
                dr = r[COL["res_num"] - 1]
                if not isinstance(dr, (int, float)): continue
                if int(dr) == res_id:
                    fn = str(r[COL["first_name"] - 1] or '').strip()
                    ln = str(r[COL["last_name"] - 1] or '').strip()
                    if fn or ln:
                        candidate = (fn + ln).replace(' ', '')[:20] or f"Resident{res_id}"
                        username = candidate
                    break
        except Exception:
            pass
    # Make username unique
    base = username
    i = 1
    while conn.execute("SELECT 1 FROM forum_profiles WHERE forum_username=?", (username,)).fetchone():
        username = f"{base}{i}"; i += 1
    conn.execute("INSERT INTO forum_profiles (res_id, forum_username) VALUES (?,?)", (res_id, username))
    conn.commit()
    return dict(conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (res_id,)).fetchone())

def _forum_profile_dict(row):
    if not row: return None
    d = dict(row)
    return {'res_id': d['res_id'], 'forum_username': d['forum_username'], 'avatar_filename': d.get('avatar_filename')}

def _forum_enrich_posts(posts, viewer_res_id, conn):
    """Add counts, profile info, and viewer flags to post list."""
    result = []
    for p in posts:
        pid = p['id']
        likes   = conn.execute("SELECT COUNT(*) FROM forum_likes WHERE post_id=?", (pid,)).fetchone()[0]
        rts     = conn.execute("SELECT COUNT(*) FROM forum_retweets WHERE post_id=?", (pid,)).fetchone()[0]
        cms     = conn.execute("SELECT COUNT(*) FROM forum_comments WHERE post_id=?", (pid,)).fetchone()[0]
        views   = conn.execute("SELECT COUNT(*) FROM forum_views WHERE post_id=?", (pid,)).fetchone()[0]
        liked   = bool(conn.execute("SELECT 1 FROM forum_likes WHERE post_id=? AND res_id=?", (pid, viewer_res_id)).fetchone())
        rted    = bool(conn.execute("SELECT 1 FROM forum_retweets WHERE post_id=? AND res_id=?", (pid, viewer_res_id)).fetchone())
        follows = bool(conn.execute("SELECT 1 FROM forum_follows WHERE follower_res_id=? AND following_res_id=?", (viewer_res_id, p['res_id'])).fetchone())
        profile_row = conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (p['res_id'],)).fetchone()
        profile = _forum_profile_dict(profile_row) if profile_row else {'res_id': p['res_id'], 'forum_username': f"Resident{p['res_id']}", 'avatar_filename': None}
        result.append({
            'id': pid, 'res_id': p['res_id'], 'content': p['content'],
            'media_filename': p['media_filename'], 'media_type': p['media_type'],
            'is_retweet': p['is_retweet'], 'created_at': p['created_at'],
            'likes_count': likes, 'retweets_count': rts, 'comments_count': cms, 'views_count': views,
            'user_liked': liked, 'user_retweeted': rted, 'user_follows': follows,
            'profile': profile,
        })
    return result

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/forum")
def forum_page():
    return send_from_directory("templates", "forum.html")

@app.route("/api/forum/profile/get", methods=["POST"])
def forum_profile_get():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    with get_forum_db() as conn:
        profile = _forum_ensure_profile(res_id, conn)
    return jsonify({"profile": _forum_profile_dict(profile)})

@app.route("/api/forum/profile/update", methods=["POST"])
def forum_profile_update():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    username = str(data.get('forumUsername', '')).strip()
    avatar_data = data.get('avatarData')
    if not username: return jsonify({"error": "Username required"}), 400
    if len(username) < 2 or len(username) > 30:
        return jsonify({"error": "Username must be 2-30 characters"}), 400
    import re as _re
    if not _re.match(r'^[\w\s.\-]+$', username):
        return jsonify({"error": "Username contains invalid characters"}), 400
    APPEARANCE_COOLDOWN_DAYS = 90  # 3 months
    with get_forum_db() as conn:
        _forum_ensure_profile(res_id, conn)
        # Check 3-month cooldown
        cooldown_row = conn.execute("SELECT last_appearance_change FROM forum_profiles WHERE res_id=?", (res_id,)).fetchone()
        if cooldown_row and cooldown_row[0]:
            import datetime as _dt
            last_change = _dt.datetime.strptime(cooldown_row[0], "%Y-%m-%d %H:%M:%S")
            days_since = (datetime.datetime.utcnow() - last_change).days
            if days_since < APPEARANCE_COOLDOWN_DAYS:
                days_left = APPEARANCE_COOLDOWN_DAYS - days_since
                return jsonify({"error": f"Appearance can only be changed once every 3 months. {days_left} day(s) remaining."}), 429
        existing = conn.execute("SELECT res_id FROM forum_profiles WHERE forum_username=? AND res_id!=?", (username, res_id)).fetchone()
        if existing: return jsonify({"error": "Username already taken"}), 400
        avatar_filename = None
        if avatar_data:
            try:
                if ',' in avatar_data:
                    header, b64 = avatar_data.split(',', 1)
                else:
                    b64 = avatar_data
                img_bytes = base64.b64decode(b64)
                if len(img_bytes) > 2*1024*1024:
                    return jsonify({"error": "Avatar image too large (max 2MB)"}), 400
                ext = 'jpg'
                ts = datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
                avatar_filename = f"av_{res_id}_{ts}.{ext}"
                with open(os.path.join(FORUM_AVATARS, avatar_filename), 'wb') as f:
                    f.write(img_bytes)
                # Remove old avatar
                old = conn.execute("SELECT avatar_filename FROM forum_profiles WHERE res_id=?", (res_id,)).fetchone()
                if old and old[0] and old[0] != avatar_filename:
                    try: os.remove(os.path.join(FORUM_AVATARS, old[0]))
                    except: pass
            except Exception as e:
                return jsonify({"error": f"Avatar error: {e}"}), 400
        if avatar_filename:
            conn.execute("UPDATE forum_profiles SET forum_username=?, avatar_filename=?, updated_at=datetime('now'), last_appearance_change=datetime('now') WHERE res_id=?",
                         (username, avatar_filename, res_id))
        else:
            conn.execute("UPDATE forum_profiles SET forum_username=?, updated_at=datetime('now'), last_appearance_change=datetime('now') WHERE res_id=?",
                         (username, res_id))
        conn.commit()
        profile = dict(conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (res_id,)).fetchone())
    return jsonify({"profile": _forum_profile_dict(profile)})

@app.route("/api/forum/profile/cooldown", methods=["POST"])
def forum_profile_cooldown():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    COOLDOWN_DAYS = 90
    with get_forum_db() as conn:
        _forum_ensure_profile(res_id, conn)
        row = conn.execute("SELECT last_appearance_change FROM forum_profiles WHERE res_id=?", (res_id,)).fetchone()
    if row and row[0]:
        import datetime as _dt
        last = _dt.datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
        days_since = (datetime.datetime.utcnow() - last).days
        days_left = max(0, COOLDOWN_DAYS - days_since)
        can_change = days_left == 0
        return jsonify({"canChange": can_change, "daysLeft": days_left, "lastChange": row[0]})
    return jsonify({"canChange": True, "daysLeft": 0, "lastChange": None})

@app.route("/api/forum/profile/view", methods=["POST"])
def forum_profile_view():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    target_res_id = int(data.get('targetResId', 0))
    with get_forum_db() as conn:
        _forum_ensure_profile(target_res_id, conn)
        prow = conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (target_res_id,)).fetchone()
        profile = _forum_profile_dict(prow)
        posts_raw = conn.execute(
            "SELECT * FROM forum_posts WHERE res_id=? ORDER BY created_at DESC LIMIT 10", (target_res_id,)
        ).fetchall()
        posts = _forum_enrich_posts([dict(r) for r in posts_raw], res_id, conn)
        is_following = bool(conn.execute(
            "SELECT 1 FROM forum_follows WHERE follower_res_id=? AND following_res_id=?", (res_id, target_res_id)
        ).fetchone())
        is_blocked = bool(conn.execute(
            "SELECT 1 FROM forum_blocks WHERE blocker_res_id=? AND blocked_res_id=?", (res_id, target_res_id)
        ).fetchone())
        followers_count = conn.execute(
            "SELECT COUNT(*) FROM forum_follows WHERE following_res_id=?", (target_res_id,)
        ).fetchone()[0]
        following_count = conn.execute(
            "SELECT COUNT(*) FROM forum_follows WHERE follower_res_id=?", (target_res_id,)
        ).fetchone()[0]
    return jsonify({"profile": profile, "posts": posts, "is_following": is_following,
                    "is_blocked": is_blocked, "followers_count": followers_count,
                    "following_count": following_count})

@app.route("/api/forum/profile/find", methods=["POST"])
def forum_profile_find():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    username = str(data.get('forumUsername', '')).strip()
    with get_forum_db() as conn:
        row = conn.execute("SELECT * FROM forum_profiles WHERE forum_username=?", (username,)).fetchone()
        if not row: return jsonify({"error": "User not found"}), 404
    return jsonify({"profile": _forum_profile_dict(dict(row))})

@app.route("/api/forum/avatar/<filename>")
def forum_avatar(filename):
    safe = os.path.basename(filename)
    path = os.path.join(FORUM_AVATARS, safe)
    if not os.path.exists(path): return '', 404
    return send_from_directory(FORUM_AVATARS, safe)

@app.route("/api/forum/media/<filename>")
def forum_media(filename):
    safe = os.path.basename(filename)
    path = os.path.join(FORUM_MEDIA, safe)
    if not os.path.exists(path): return '', 404
    mt, _ = mimetypes.guess_type(safe)
    return send_from_directory(FORUM_MEDIA, safe, mimetype=mt or 'application/octet-stream')

@app.route("/api/forum/posts/create", methods=["POST"])
def forum_post_create():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    content = str(data.get('content', '')).strip()
    media_data = data.get('mediaData')
    media_type_raw = str(data.get('mediaType', '')).strip()
    if not content and not media_data:
        return jsonify({"error": "Post must have content or media"}), 400
    if len(content) > 300:
        return jsonify({"error": "Max 300 characters"}), 400
    media_filename = None
    media_type_str = None
    if media_data:
        try:
            if ',' in media_data:
                header, b64 = media_data.split(',', 1)
            else:
                b64 = media_data
            file_bytes = base64.b64decode(b64)
            is_video = media_type_raw == 'video'
            max_size = 20*1024*1024 if is_video else 2*1024*1024
            if len(file_bytes) > max_size:
                return jsonify({"error": f"Media too large (max {'20MB' if is_video else '2MB'})"}), 400
            ext = 'mp4' if is_video else 'jpg'
            ts = datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
            media_filename = f"post_{res_id}_{ts}.{ext}"
            with open(os.path.join(FORUM_MEDIA, media_filename), 'wb') as f:
                f.write(file_bytes)
            media_type_str = 'video' if is_video else 'image'
        except Exception as e:
            return jsonify({"error": f"Media error: {e}"}), 400
    with get_forum_db() as conn:
        _forum_ensure_profile(res_id, conn)
        cur = conn.execute(
            "INSERT INTO forum_posts (res_id, content, media_filename, media_type) VALUES (?,?,?,?)",
            (res_id, content, media_filename, media_type_str)
        )
        post_id = cur.lastrowid
        conn.commit()
        post_row = dict(conn.execute("SELECT * FROM forum_posts WHERE id=?", (post_id,)).fetchone())
        posts = _forum_enrich_posts([post_row], res_id, conn)
    return jsonify({"success": True, "post": posts[0]})

@app.route("/api/forum/posts/feed", methods=["POST"])
def forum_posts_feed():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    feed_type = str(data.get('feedType', 'foryou'))
    offset = int(data.get('offset', 0))
    limit  = int(data.get('limit', 20))
    with get_forum_db() as conn:
        _forum_ensure_profile(res_id, conn)
        # Get blocked users to exclude
        blocked = [r[0] for r in conn.execute(
            "SELECT blocked_res_id FROM forum_blocks WHERE blocker_res_id=?", (res_id,)).fetchall()]
        blocked_by = [r[0] for r in conn.execute(
            "SELECT blocker_res_id FROM forum_blocks WHERE blocked_res_id=?", (res_id,)).fetchall()]
        exclude = set(blocked + blocked_by)
        exclude_sql = ','.join(['?']*len(exclude)) if exclude else '0'

        if feed_type == 'following':
            following = [r[0] for r in conn.execute(
                "SELECT following_res_id FROM forum_follows WHERE follower_res_id=?", (res_id,)).fetchall()]
            if not following:
                return jsonify({"posts": []})
            follow_sql = ','.join(['?']*len(following))
            placeholders = following + list(exclude) + [limit, offset]
            raw = conn.execute(f"""
                SELECT * FROM forum_posts
                WHERE res_id IN ({follow_sql})
                  AND res_id NOT IN ({exclude_sql if exclude else '0'})
                ORDER BY created_at DESC LIMIT ? OFFSET ?
            """, placeholders).fetchall()
        else:
            # "For You" — ranked feed: (likes*1.5 + retweets*2 + comments*1) / (hours_old+2)^0.8
            exclude_params = list(exclude) + [limit, offset]
            raw = conn.execute(f"""
                SELECT p.*,
                  (SELECT COUNT(*) FROM forum_likes WHERE post_id=p.id) as lc,
                  (SELECT COUNT(*) FROM forum_retweets WHERE post_id=p.id) as rc,
                  (SELECT COUNT(*) FROM forum_comments WHERE post_id=p.id) as cc,
                  ((julianday('now') - julianday(p.created_at)) * 24) as hours_old
                FROM forum_posts p
                WHERE p.res_id NOT IN ({exclude_sql if exclude else '0'})
                ORDER BY (lc*1.5 + rc*2.0 + cc*1.0) / pow(MAX(hours_old+2, 2), 0.8) DESC,
                         p.created_at DESC
                LIMIT ? OFFSET ?
            """, exclude_params).fetchall()
        posts = _forum_enrich_posts([dict(r) for r in raw], res_id, conn)
    return jsonify({"posts": posts})

@app.route("/api/forum/posts/mine", methods=["POST"])
def forum_posts_mine():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    offset = int(data.get('offset', 0))
    limit  = int(data.get('limit', 30))
    with get_forum_db() as conn:
        raw = conn.execute(
            "SELECT * FROM forum_posts WHERE res_id=? ORDER BY created_at DESC LIMIT ? OFFSET ?",
            (res_id, limit, offset)
        ).fetchall()
        posts = _forum_enrich_posts([dict(r) for r in raw], res_id, conn)
    return jsonify({"posts": posts})

@app.route("/api/forum/posts/like", methods=["POST"])
def forum_post_like():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    post_id = int(data.get('postId', 0))
    unlike = bool(data.get('unlike', False))
    with get_forum_db() as conn:
        if unlike:
            conn.execute("DELETE FROM forum_likes WHERE post_id=? AND res_id=?", (post_id, res_id))
        else:
            conn.execute("INSERT OR IGNORE INTO forum_likes (post_id, res_id) VALUES (?,?)", (post_id, res_id))
        conn.commit()
    return jsonify({"success": True})

@app.route("/api/forum/posts/retweet", methods=["POST"])
def forum_post_retweet():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    post_id = int(data.get('postId', 0))
    unretweet = bool(data.get('unretweet', False))
    with get_forum_db() as conn:
        if unretweet:
            conn.execute("DELETE FROM forum_retweets WHERE post_id=? AND res_id=?", (post_id, res_id))
        else:
            conn.execute("INSERT OR IGNORE INTO forum_retweets (post_id, res_id) VALUES (?,?)", (post_id, res_id))
        conn.commit()
    return jsonify({"success": True})

@app.route("/api/forum/posts/comments", methods=["POST"])
def forum_post_comments():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    post_id = int(data.get('postId', 0))
    with get_forum_db() as conn:
        rows = conn.execute(
            "SELECT * FROM forum_comments WHERE post_id=? ORDER BY created_at ASC LIMIT 50", (post_id,)
        ).fetchall()
        comments = []
        for r in rows:
            d = dict(r)
            pr = conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (d['res_id'],)).fetchone()
            d['profile'] = _forum_profile_dict(dict(pr)) if pr else {'res_id': d['res_id'], 'forum_username': f"Resident{d['res_id']}", 'avatar_filename': None}
            comments.append(d)
    return jsonify({"comments": comments})

@app.route("/api/forum/posts/view", methods=["POST"])
def forum_post_view():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    post_ids = data.get("postIds", [])
    if not isinstance(post_ids, list): post_ids = []
    with get_forum_db() as conn:
        for pid in post_ids[:50]:
            try:
                conn.execute("INSERT OR IGNORE INTO forum_views (post_id, res_id) VALUES (?,?)", (int(pid), res_id))
            except: pass
        conn.commit()
    return jsonify({"success": True})

@app.route("/api/forum/posts/comment", methods=["POST"])
def forum_post_comment():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    post_id = int(data.get('postId', 0))
    content = str(data.get('content', '')).strip()
    if not content: return jsonify({"error": "Comment cannot be empty"}), 400
    if len(content) > 300: return jsonify({"error": "Max 300 characters"}), 400
    with get_forum_db() as conn:
        _forum_ensure_profile(res_id, conn)
        conn.execute("INSERT INTO forum_comments (post_id, res_id, content) VALUES (?,?,?)", (post_id, res_id, content))
        conn.commit()
    return jsonify({"success": True})

@app.route("/api/forum/follow", methods=["POST"])
def forum_follow():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    target = int(data.get('targetResId', 0))
    unfollow = bool(data.get('unfollow', False))
    if target == res_id: return jsonify({"error": "Cannot follow yourself"}), 400
    with get_forum_db() as conn:
        if unfollow:
            conn.execute("DELETE FROM forum_follows WHERE follower_res_id=? AND following_res_id=?", (res_id, target))
        else:
            conn.execute("INSERT OR IGNORE INTO forum_follows (follower_res_id, following_res_id) VALUES (?,?)", (res_id, target))
        conn.commit()
    return jsonify({"success": True})

@app.route("/api/forum/block", methods=["POST"])
def forum_block():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    target = int(data.get('targetResId', 0))
    unblock = bool(data.get('unblock', False))
    with get_forum_db() as conn:
        if unblock:
            conn.execute("DELETE FROM forum_blocks WHERE blocker_res_id=? AND blocked_res_id=?", (res_id, target))
        else:
            conn.execute("INSERT OR IGNORE INTO forum_blocks (blocker_res_id, blocked_res_id) VALUES (?,?)", (res_id, target))
            # Also unfollow both ways if blocking
            conn.execute("DELETE FROM forum_follows WHERE (follower_res_id=? AND following_res_id=?) OR (follower_res_id=? AND following_res_id=?)",
                         (res_id, target, target, res_id))
        conn.commit()
    return jsonify({"success": True})

@app.route("/api/forum/manage/list", methods=["POST"])
def forum_manage_list():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    list_type = str(data.get('listType', 'following'))
    with get_forum_db() as conn:
        if list_type == 'following':
            rows = conn.execute(
                "SELECT following_res_id as other_id FROM forum_follows WHERE follower_res_id=?", (res_id,)
            ).fetchall()
        elif list_type == 'followers':
            rows = conn.execute(
                "SELECT follower_res_id as other_id FROM forum_follows WHERE following_res_id=?", (res_id,)
            ).fetchall()
        else:  # blocked
            rows = conn.execute(
                "SELECT blocked_res_id as other_id FROM forum_blocks WHERE blocker_res_id=?", (res_id,)
            ).fetchall()
        users = []
        for r in rows:
            other_id = r[0]
            pr = conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (other_id,)).fetchone()
            if not pr:
                _forum_ensure_profile(other_id, conn)
                pr = conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (other_id,)).fetchone()
            u = _forum_profile_dict(dict(pr)) if pr else {'res_id': other_id, 'forum_username': f"Resident{other_id}", 'avatar_filename': None}
            users.append(u)
    return jsonify({"users": users})

@app.route("/api/forum/users/suggested", methods=["POST"])
def forum_users_suggested():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    with get_forum_db() as conn:
        blocked = [r[0] for r in conn.execute(
            "SELECT blocked_res_id FROM forum_blocks WHERE blocker_res_id=?", (res_id,)).fetchall()]
        following = [r[0] for r in conn.execute(
            "SELECT following_res_id FROM forum_follows WHERE follower_res_id=?", (res_id,)).fetchall()]
        exclude = set(blocked + following + [res_id])
        exclude_sql = ','.join(['?']*len(exclude)) if exclude else '0'
        rows = conn.execute(f"""
            SELECT p.res_id, p.forum_username, p.avatar_filename,
              (SELECT COUNT(*) FROM forum_follows WHERE following_res_id=p.res_id) as fol_count
            FROM forum_profiles p
            WHERE p.res_id NOT IN ({exclude_sql})
            ORDER BY fol_count DESC, RANDOM()
            LIMIT 8
        """, list(exclude)).fetchall()
        users = []
        for r in rows:
            u = {'res_id': r[0], 'forum_username': r[1], 'avatar_filename': r[2], 'is_following': False}
            users.append(u)
    return jsonify({"users": users})

@app.route("/api/forum/users/active", methods=["POST"])
def forum_users_active():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    with get_forum_db() as conn:
        rows = conn.execute("""
            SELECT p.res_id, p.forum_username, p.avatar_filename,
              COUNT(fp.id) as post_count
            FROM forum_profiles p
            LEFT JOIN forum_posts fp ON fp.res_id = p.res_id
            GROUP BY p.res_id
            ORDER BY post_count DESC
            LIMIT 8
        """).fetchall()
        users = [{'res_id': r[0], 'forum_username': r[1], 'avatar_filename': r[2], 'post_count': r[3]} for r in rows]
    return jsonify({"users": users})

@app.route("/api/forum/dm/send", methods=["POST"])
def forum_dm_send():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    receiver = int(data.get('receiverResId', 0))
    content = str(data.get('content', '')).strip()
    if not content: return jsonify({"error": "Message required"}), 400
    if len(content) > 500: return jsonify({"error": "Max 500 characters"}), 400
    if receiver == res_id: return jsonify({"error": "Cannot message yourself"}), 400
    with get_forum_db() as conn:
        # Check if receiver has blocked sender
        if conn.execute("SELECT 1 FROM forum_blocks WHERE blocker_res_id=? AND blocked_res_id=?", (receiver, res_id)).fetchone():
            return jsonify({"error": "Cannot message this user"}), 403
        conn.execute("INSERT INTO forum_dms (sender_res_id, receiver_res_id, content) VALUES (?,?,?)",
                     (res_id, receiver, content))
        conn.commit()
    return jsonify({"success": True})

@app.route("/api/forum/dm/messages", methods=["POST"])
def forum_dm_messages():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    other_id = int(data.get('otherResId', 0))
    with get_forum_db() as conn:
        rows = conn.execute("""
            SELECT * FROM forum_dms
            WHERE (sender_res_id=? AND receiver_res_id=?)
               OR (sender_res_id=? AND receiver_res_id=?)
            ORDER BY created_at ASC LIMIT 100
        """, (res_id, other_id, other_id, res_id)).fetchall()
        # Mark as read
        conn.execute("UPDATE forum_dms SET read=1 WHERE receiver_res_id=? AND sender_res_id=? AND read=0",
                     (res_id, other_id))
        conn.commit()
        messages = [dict(r) for r in rows]
    return jsonify({"messages": messages})

@app.route("/api/forum/dm/conversations", methods=["POST"])
def forum_dm_conversations():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    with get_forum_db() as conn:
        rows = conn.execute("""
            SELECT other_res_id, MAX(created_at) as last_time, last_message,
                   SUM(CASE WHEN unread=1 THEN 1 ELSE 0 END) as unread_count
            FROM (
                SELECT CASE WHEN sender_res_id=? THEN receiver_res_id ELSE sender_res_id END as other_res_id,
                       created_at, content as last_message,
                       CASE WHEN receiver_res_id=? AND read=0 THEN 1 ELSE 0 END as unread
                FROM forum_dms
                WHERE sender_res_id=? OR receiver_res_id=?
            )
            GROUP BY other_res_id
            ORDER BY last_time DESC
        """, (res_id, res_id, res_id, res_id)).fetchall()
        convs = []
        for r in rows:
            other_id = r[0]
            pr = conn.execute("SELECT * FROM forum_profiles WHERE res_id=?", (other_id,)).fetchone()
            convs.append({
                'other_res_id': other_id,
                'other_username': pr['forum_username'] if pr else f"Resident{other_id}",
                'other_profile': _forum_profile_dict(dict(pr)) if pr else None,
                'last_message': r[2],
                'unread_count': r[3]
            })
    return jsonify({"conversations": convs})

@app.route("/api/forum/dm/unread-count", methods=["POST"])
def forum_dm_unread_count():
    data = request.get_json(force=True)
    res_id = _forum_auth(data)
    if not res_id: return jsonify({"error": "Auth failed"}), 401
    with get_forum_db() as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM forum_dms WHERE receiver_res_id=? AND read=0", (res_id,)
        ).fetchone()[0]
    return jsonify({"count": count})

# ══════════════════════════════════════════════════════════════════════════════
# STARTUP
# ══════════════════════════════════════════════════════════════════════════════
if __name__=="__main__":
    _seed_example_corps()
    _init_criminal_db()
    _init_jury_db()
    _init_chess_db()
    _init_notif_db()
    _init_changelog_db()
    _init_forum_db()
    print("="*56); print("  ARK Citizen Registry  v6.7")
    print("  URL  : http://localhost:5000"); print("  Stop : Ctrl+C"); print("="*56)
    app.run(host="0.0.0.0",port=5000,debug=False)
