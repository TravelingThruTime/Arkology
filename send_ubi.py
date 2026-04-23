"""
send_ubi.py — Ark Monthly UBI / Net Payment Sender
Reads the Ark_Database_v6-1.xlsx, finds all living residents with:
  - Positive column M (Payment) value
  - A BCH address in column BA
Then either:
  (a) Prints a dry-run summary, or
  (b) Sends transactions via Electron Cash CLI

Usage:
  python send_ubi.py           → dry run (preview only)
  python send_ubi.py --send    → actually send via electron-cash CLI

REQUIRES: Electron Cash installed and unlocked wallet open.
  electron-cash must be accessible on PATH, OR set ELECTRON_CASH_PATH below.

WARNING: Only run this ONCE per month. Double-clicking twice sends double pay.
"""

import os, sys, datetime, subprocess, math

ELECTRON_CASH_PATH = r"C:\Program Files\Electron Cash\Electron-Cash.exe"  # Verified path — update if Electron Cash is installed elsewhere
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Ark_Database_v6-1.xlsx")
LOG_FILE   = os.path.join(os.path.dirname(__file__), "ubi_payment_log.txt")

DRY_RUN = "--send" not in sys.argv

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl not installed. Run SETUP_FIRST_TIME.bat first.")
    input("\nPress Enter to close...")
    sys.exit(1)

def load_payments(a19):
    """Read Excel and compute UBI payments from net_tax * A19 (NOT cached col M).
    Only includes alive residents who passed ref check + immigration interview.
    """
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    # Column indices (0-based for values_only iteration)
    COL_RESNUM = 1   # B
    COL_FIRST  = 5   # F
    COL_LAST   = 4   # E
    COL_NET    = 13  # N — net tax (negative = owes, positive = receives UBI)
    COL_ALIVE  = 15  # P
    COL_BCH    = 49  # AX — BCH Address
    COL_REF_CHK= 95  # CR — Ref Checked (col 96)
    COL_IMMIG  = 99  # CV — Immigration Interview (col 100)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Res"]

    payments = []
    skipped_ref = []
    skipped_immig = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        res_num = row[COL_RESNUM]
        alive   = row[COL_ALIVE]
        bch     = row[COL_BCH]
        net_tax = row[COL_NET]        # raw net tax — compute payment as net_tax * A19
        first   = row[COL_FIRST] or ""
        last    = row[COL_LAST]  or ""
        ref_chk = row[COL_REF_CHK] if len(row) > COL_REF_CHK else None
        immig   = row[COL_IMMIG]   if len(row) > COL_IMMIG   else None

        if not isinstance(res_num, (int, float)): continue
        if alive != 1: continue
        if not bch or not str(bch).strip().startswith("bitcoincash:"): continue
        if not isinstance(net_tax, (int, float)) or net_tax <= 0: continue

        # Skip residents whose references have NOT been checked
        if ref_chk != 1:
            skipped_ref.append({"res_num": int(res_num), "name": f"{first} {last}".strip()})
            continue

        # Skip residents who have NOT passed the immigration interview
        if immig != 1:
            skipped_immig.append({"res_num": int(res_num), "name": f"{first} {last}".strip()})
            continue

        # Compute actual payment from live A19 × net_tax (avoids stale cached col M)
        payment = round(float(net_tax) * a19, 2)
        if payment <= 0:
            continue

        payments.append({
            "res_num": int(res_num),
            "name":    f"{first} {last}".strip(),
            "bch":     str(bch).strip(),
            "amount":  payment,
        })

    if skipped_ref:
        print(f"\n  [WARN] {len(skipped_ref)} resident(s) skipped — references not checked:")
        for s in skipped_ref:
            print(f"         Res#{s['res_num']} {s['name']}")
        print()

    if skipped_immig:
        print(f"\n  [WARN] {len(skipped_immig)} resident(s) skipped — immigration interview not passed:")
        for s in skipped_immig:
            print(f"         Res#{s['res_num']} {s['name']}")
        print()

    return payments

def get_bch_rate():
    """Fetch live BCH/USD rate from CoinGecko."""
    import urllib.request, json as _json
    try:
        url = "https://api.coingecko.com/api/v3/simple/price?ids=bitcoin-cash&vs_currencies=usd"
        req = urllib.request.Request(url, headers={"User-Agent": "ArkRegistry/1.0"})
        with urllib.request.urlopen(req, timeout=8) as r:
            data = _json.loads(r.read())
        return float(data["bitcoin-cash"]["usd"])
    except:
        return 43000  # fallback


def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(msg + "\n")

def send_via_electron_cash(bch_address, amount_bch):
    """Send amount_bch BCH to bch_address using Electron Cash CLI (two-step: payto → broadcast)."""
    try:
        # Step 1: Create signed transaction
        cmd_payto = [ELECTRON_CASH_PATH, "payto", bch_address, str(amount_bch)]
        result_payto = subprocess.run(cmd_payto, capture_output=True, text=True, timeout=60)
        if result_payto.returncode != 0:
            return False, result_payto.stderr.strip() or result_payto.stdout.strip()
        signed_tx = result_payto.stdout.strip()
        if not signed_tx:
            return False, "payto returned empty transaction"
        # Step 2: Broadcast the signed transaction
        cmd_broadcast = [ELECTRON_CASH_PATH, "broadcast", signed_tx]
        result_bc = subprocess.run(cmd_broadcast, capture_output=True, text=True, timeout=60)
        if result_bc.returncode == 0:
            return True, result_bc.stdout.strip()
        return False, result_bc.stderr.strip() or result_bc.stdout.strip()
    except FileNotFoundError:
        return False, (
            f"Electron Cash not found at '{ELECTRON_CASH_PATH}'.\n"
            f"  Set ELECTRON_CASH_PATH at the top of send_ubi.py to the full path.\n"
            f"  Example: ELECTRON_CASH_PATH = r'C:\\Program Files\\Electron Cash\\electron-cash.exe'"
        )
    except Exception as e:
        return False, f"Error: {e}"

def load_charity_and_employee_payments(a19):
    """Load charities and employees with BCH addresses, scaled by A19."""
    from openpyxl import load_workbook as _lw
    extra = []
    try:
        wb = _lw(EXCEL_FILE, data_only=True)
        year = datetime.date.today().year
        budget_col = 16 + (year - 2025)

        # Charities
        if "Charity" in wb.sheetnames:
            ws_c = wb["Charity"]
            for r in range(2, ws_c.max_row + 1):
                cid  = ws_c.cell(r, 2).value
                name = ws_c.cell(r, 3).value
                bch  = str(ws_c.cell(r, 7).value or "").strip()
                budget = ws_c.cell(r, budget_col).value
                if not cid or not name: continue
                if not bch.startswith("bitcoincash:"): continue
                if not isinstance(budget, (int, float)) or budget <= 0: continue
                pay = round(float(budget) * a19, 2)
                if pay > 0:
                    extra.append({"type": "Charity", "res_num": None,
                                  "name": str(name), "bch": bch, "amount": pay})

        # Gov Employees (with BCH via Res#)
        if "Gov Employees" in wb.sheetnames and "Res" in wb.sheetnames:
            ws_emp = wb["Gov Employees"]; ws_res = wb["Res"]
            res_bch = {}
            for row in ws_res.iter_rows(min_row=2, values_only=True):
                rn = row[1]; bch = row[49]  # col B and col AX
                if isinstance(rn, (int, float)) and bch and str(bch).strip().startswith("bitcoincash:"):
                    res_bch[int(rn)] = str(bch).strip()
            for row in ws_emp.iter_rows(min_row=2, values_only=True):
                emp_num, res_num, last, first, position, salary, d_hired, active, hexarchy = (row+(None,)*9)[:9]
                if not isinstance(salary, (int,float)) or salary <= 0: continue
                if str(active or "").strip().lower() not in ("yes","y","true","1"): continue
                if not isinstance(res_num, (int,float)): continue
                ri = int(res_num)
                if ri not in res_bch: continue
                pay = round(float(salary) * a19, 2)  # Scale salary by A19 like CHECK_BCH_BALANCE
                if pay > 0:
                    nm = f"{first or ''} {last or ''}".strip() or f"Res#{ri}"
                    extra.append({"type": "Employee", "res_num": ri,
                                  "name": nm, "bch": res_bch[ri], "amount": pay})
        wb.close()
    except Exception as e:
        print(f"  [WARN] Could not load charity/employee payments: {e}")
    return extra


def get_a19():
    """Read A19 (balancing multiplier) from spreadsheet."""
    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(EXCEL_FILE, data_only=True)
        a19 = float(wb["Res"].cell(19, 1).value or 0.75)
        wb.close()
        return a19
    except:
        return 0.75


def main():
    print()
    print("=" * 60)
    print("  ARK UBI / PAYMENT SENDER")
    print(f"  Date: {datetime.date.today().strftime('%B %d, %Y')}")
    print(f"  Mode: {'DRY RUN (no transactions sent)' if DRY_RUN else '*** LIVE — SENDING TRANSACTIONS ***'}")
    print("=" * 60)
    print()

    a19 = get_a19()
    print(f"  Balancing multiplier (A19): {a19:.6f}  ({a19*100:.2f}%)")
    print()

    payments = load_payments(a19)
    extra_payments = load_charity_and_employee_payments(a19)

    all_payments = payments + extra_payments
    if not all_payments:
        print("[INFO] No eligible residents/charities/employees found with BCH.")
        input("\nPress Enter to close...")
        return

    res_total   = sum(p["amount"] for p in payments)
    extra_total = sum(p["amount"] for p in extra_payments)
    grand_total = res_total + extra_total

    print(f"  UBI recipients:  {len(payments)}  (${res_total:,.2f})")
    print(f"  Charities/Empl:  {len(extra_payments)}  (${extra_total:,.2f})")
    print(f"  Grand total:     ${grand_total:,.2f}")
    print()
    print(f"  {'Type':<8} {'Res#':<6} {'Name':<24} {'BCH Address':<44} {'Amount':>10}")
    print(f"  {'-'*8} {'-'*6} {'-'*24} {'-'*44} {'-'*10}")
    for p in payments:
        addr_short = p["bch"][:20] + "..." + p["bch"][-8:] if len(p["bch"]) > 32 else p["bch"]
        print(f"  {'Resident':<8} {p['res_num']:<6} {p['name']:<24} {addr_short:<44} ${p['amount']:>9.2f}")
    for p in extra_payments:
        addr_short = p["bch"][:20] + "..." + p["bch"][-8:] if len(p["bch"]) > 32 else p["bch"]
        rn = str(p["res_num"]) if p["res_num"] else "—"
        print(f"  {p['type']:<8} {rn:<6} {p['name']:<24} {addr_short:<44} ${p['amount']:>9.2f}")
    print()

    if DRY_RUN:
        print("  *** DRY RUN — no transactions sent ***")
        print("  To actually send, run:  python send_ubi.py --send")
        print("  Or double-click:        SEND_UBI_LIVE.bat")
        print()
        input("Press Enter to close...")
        return

    # ── BCH balance pre-check ──────────────────────────────────────────────────
    BCH_SYSTEM_ADDRESS = "bitcoincash:qqvvwjw8rns37te6r4a0jzvtu86s3kx6lsykymqj5k"
    import urllib.request, json as _json
    bch_rate = get_bch_rate()
    total_usd = grand_total
    total_bch_needed = total_usd / bch_rate if bch_rate > 0 else 0

    print(f"  BCH rate: ${bch_rate:,.2f} USD/BCH")
    print(f"  Total USD to pay: ${total_usd:,.2f}  ({len(all_payments)} recipients)")
    print(f"  Total BCH needed: {total_bch_needed:.5f} BCH")

    wallet_balance_bch = 0
    try:
        url = f"https://blockchair.com/bitcoin-cash/dashboards/address/{BCH_SYSTEM_ADDRESS}?limit=1"
        req = urllib.request.Request(url, headers={"User-Agent": "ArkRegistry/1.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = _json.loads(r.read())
        addr = data.get("data", {}).get(BCH_SYSTEM_ADDRESS, {}).get("address", {})
        wallet_balance_bch = addr.get("balance", 0) / 1e8
        print(f"  Wallet balance:   {wallet_balance_bch:.5f} BCH  (${wallet_balance_bch * bch_rate:,.2f})")
    except Exception as e:
        print(f"  [WARN] Could not fetch wallet balance: {e}")
        print("  Proceeding without balance verification...")

    # Scale payments if balance is insufficient
    scale = 1.0
    FEE_RESERVE = 0.99  # assume 1% tx fees
    if wallet_balance_bch > 0 and total_bch_needed > 0:
        available = wallet_balance_bch * FEE_RESERVE
        if available < total_bch_needed:
            scale = available / total_bch_needed
            scaled_total = total_usd * scale
            print()
            print(f"  ⚠  INSUFFICIENT BALANCE — scaling all payments to {scale*100:.1f}%")
            print(f"     Original total: ${total_usd:,.2f}  →  Scaled total: ${scaled_total:,.2f}")
            for p in all_payments:
                p["amount"] = round(p["amount"] * scale, 2)
                p["bch_amount"] = round(p["amount"] / bch_rate, 5)
        else:
            for p in all_payments:
                p["bch_amount"] = round(p["amount"] / bch_rate, 5)
    else:
        for p in all_payments:
            p["bch_amount"] = round(p["amount"] / bch_rate if bch_rate > 0 else 0, 5)

    print()

    # LIVE MODE — confirm
    # Pre-check Electron Cash path before asking for confirmation
    import shutil
    ec_found = shutil.which(ELECTRON_CASH_PATH) is not None or os.path.isfile(ELECTRON_CASH_PATH)
    if not ec_found:
        print(f"  ⚠  WARNING: Electron Cash not found at: '{ELECTRON_CASH_PATH}'")
        print(f"  Edit the ELECTRON_CASH_PATH line at the top of send_ubi.py")
        print(f"  Example: ELECTRON_CASH_PATH = r'C:\\Program Files\\Electron Cash\\electron-cash.exe'")
        print()

    print()
    confirm = input(f"  CONFIRM: Send {len(all_payments)} transactions totaling ${grand_total:,.2f} USD? (yes/no): ").strip().lower()
    if confirm != "yes":
        print("  Cancelled.")
        input("\nPress Enter to close...")
        return

    print()
    log(f"\n{'='*60}")
    log(f"UBI PAYMENT RUN — {datetime.datetime.now().isoformat()}")
    log(f"{'='*60}")

    success_count = 0
    for p in all_payments:
        bch_amt = p.get("bch_amount", round(p["amount"] / 43000, 5))
        ok, msg = send_via_electron_cash(p["bch"], bch_amt)
        status = "OK   " if ok else "FAIL "
        line = f"  [{status}] #{p['res_num']} {p['name']:20} ${p['amount']:8.2f}  {msg}"
        log(line)
        if ok: success_count += 1

    log(f"\nResult: {success_count}/{len(all_payments)} sent successfully.")
    print()
    print(f"  Done. {success_count}/{len(all_payments)} transactions sent.")
    print(f"  Log saved to: {LOG_FILE}")
    print()
    input("Press Enter to close...")

if __name__ == "__main__":
    main()
